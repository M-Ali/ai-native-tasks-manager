import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook
wb = load_workbook("output/reports/PSO_Pricing_Strategy_Workbook.xlsx")
ws = wb["README"]
# Look at cell A1 which showed garbled em-dash
for row in ws.iter_rows(min_row=1, max_row=20):
    for cell in row:
        v = cell.value
        if v and isinstance(v, str) and len(v) > 3:
            codepoints = [f"U+{ord(c):04X}({c})" for c in v[:30]]
            print(f"{cell.coordinate}: {''.join(codepoints[:20])}")
            break
    else:
        continue
    break
