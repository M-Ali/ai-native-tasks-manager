# -*- coding: utf-8 -*-
"""
PSO Lubricants -- Pricing Strategy Excel Workbook v2
Clean rebuild: all source strings ASCII-safe, encoding fixed at generation time.
Sheets: README | Competitor Data | F1-F8 | Summary
"""
import sqlite3
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH    = Path("db/prices.db")
LUBES_FILE = Path("../Lubes Data Final.xlsx")
OUT_PATH   = Path("output/reports/PSO_Pricing_Strategy_Workbook_v2.xlsx")
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# Palette
C_DARK  = "023520"
C_MED   = "00703C"
C_GOLD  = "C8960C"
C_LIGHT = "E8F5EE"
C_WHITE = "FFFFFF"
C_GRAY  = "F0F0F0"
C_BLUE  = "1A5276"

# Domain constants
PSO_TIER = {
    "PSO Carient Ultra": "super_premium",
    "PSO Carient FS":    "premium",
    "PSO Carient Plus":  "mainstream",
    "PSO Carient SPRO":  "economy",
    "PSO DEO Max":       "super_premium",
    "PSO DEO 8000":      "super_premium",
    "PSO DEO 6000":      "premium",
    "PSO DEO 5000":      "premium",
    "PSO DEO 3000":      "mainstream",
    "PSO Dieselube":     "economy",
    "PSO Blaze Xtreme":  "premium",
    "PSO Blaze 4T":      "mainstream",
}
TIER_LABEL = {
    "super_premium": "Super Premium",
    "premium":       "Premium",
    "mainstream":    "Mainstream",
    "economy":       "Economy",
}
SHELL_TIER_BRANDS = {
    "super_premium": ["Shell Helix Ultra", "Shell"],
    "premium":       ["Shell Helix HX7", "ZIC X9"],
    "mainstream":    ["ZIC X7", "Total Quartz 9000"],
    "economy":       ["Kixx", "Kixx G1"],
}
PSO_BRAND_DISCOUNT = {
    "super_premium": 0.08, "premium": 0.06, "mainstream": 0.04, "economy": 0.02,
}
API_SPEC_PREMIUM = {
    "super_premium": 0.28, "premium": 0.18, "mainstream": 0.08, "economy": 0.00,
}
PPA_MULTIPLIER = {
    0.8: 1.55, 1.0: 1.40, 1.5: 1.30, 2.0: 1.20, 3.0: 1.12,
    3.5: 1.10, 4.0: 1.00, 5.0: 0.97, 7.0: 0.92, 10.0: 0.88, 20.0: 0.78,
}
BASE_OIL_GROUP = {
    "0W-20": "III", "5W-20": "III", "5W-30": "III", "5W-40": "III",
    "10W-30": "II", "10W-40": "II", "15W-40": "II", "20W-50": "I",
}
BASE_OIL_COST  = {"I": 280, "II": 420, "III": 680}
ADDITIVE_COST  = {
    "0W-20": 320, "5W-20": 300, "5W-30": 280, "5W-40": 260,
    "10W-30": 160, "10W-40": 140, "15W-40": 120, "20W-50": 90,
}
MIN_MARKUP = 2.8
TIER_MARGIN = {
    "super_premium": 0.45, "premium": 0.38, "mainstream": 0.30, "economy": 0.22,
}
CHANNEL_DISCOUNT = {
    "Retail (Pump)": 0.00, "Workshop": 0.12, "Distributor": 0.15, "Fleet Contract": 0.20,
}
REGIONAL_SIGNAL = {
    "pcmo": {"Central":  0.02, "North": -0.03, "South": 0.00},
    "hdeo": {"Central":  0.05, "North": -0.02, "South": 0.00},
    "mco":  {"Central":  0.00, "North": -0.08, "South": 0.00},
}
BRAND_GRADES = {
    "PSO Carient Ultra": ["0W-20", "5W-20", "5W-30", "5W-40"],
    "PSO Carient FS":    ["5W-30", "5W-40", "10W-40"],
    "PSO Carient Plus":  ["10W-40", "15W-40", "20W-50"],
    "PSO Carient SPRO":  ["15W-40", "20W-50"],
    "PSO DEO Max":       ["10W-40", "15W-40"],
    "PSO DEO 8000":      ["10W-40", "15W-40"],
    "PSO DEO 6000":      ["15W-40"],
    "PSO DEO 5000":      ["10W-40", "15W-40"],
    "PSO DEO 3000":      ["15W-40", "20W-50"],
    "PSO Dieselube":     ["15W-40", "20W-50"],
    "PSO Blaze 4T":      ["10W-40", "20W-50"],
    "PSO Blaze Xtreme":  ["10W-40"],
}
BRAND_MAP = {
    "Carient SPRO":  ("PSO Carient SPRO",  "pcmo"),
    "Carient FS":    ("PSO Carient FS",    "pcmo"),
    "Carient Ultra": ("PSO Carient Ultra", "pcmo"),
    "Carient Plus":  ("PSO Carient Plus",  "pcmo"),
    "Blaze 4T":      ("PSO Blaze 4T",      "mco"),
    "Blaze Xtreme":  ("PSO Blaze Xtreme",  "mco"),
    "DEO 3000":      ("PSO DEO 3000",      "hdeo"),
    "DEO 5000":      ("PSO DEO 5000",      "hdeo"),
    "DEO 6000":      ("PSO DEO 6000",      "hdeo"),
    "DEO 8000":      ("PSO DEO 8000",      "hdeo"),
    "DEO Max":       ("PSO DEO Max",       "hdeo"),
    "Dieselube":     ("PSO Dieselube",     "hdeo"),
}

# Methodology explanations (pure ASCII)
METHODOLOGY = {
    "F1": {
        "n": 8,
        "how": [
            "Identifies the Shell tier equivalent for each PSO brand: "
            "Carient Ultra = Shell Helix Ultra (Super Premium), Carient Plus = ZIC X7/Total Quartz 9000 (Mainstream), "
            "Carient SPRO = Kixx G1 (Economy).",
            "Pulls the Daraz.pk median price/L for that Shell benchmark brand x grade from the competitor dataset.",
            "Applies a PSO brand-perception discount: 8% Super Premium, 6% Premium, 4% Mainstream, 2% Economy.",
            "Formula: F1 Price = Shell_Tier_Median x (1 - Brand_Discount%)",
        ],
        "rationale": [
            "Shell is the cognitive price anchor in Pakistan's lubricants market. "
            "Consumers use Shell prices as a mental reference when evaluating any other brand.",
            "Pricing relative to Shell gives PSO a clear, defensible story: 'same quality tier, better value.' "
            "This is easier to communicate than abstract per-litre comparisons.",
            "The brand discount is NOT permanent. It reflects today's gap in awareness and OEM approval depth. "
            "In markets where PSO has invested in workshops, the gap is already closer to 2-3%.",
            "Using the Shell MEDIAN (not minimum) avoids being dragged down by short-term Shell promotional prices.",
        ],
        "implications": [
            "As PSO brand equity grows through OEM approvals and ATL/BTL investment, the discount should narrow: "
            "Super Premium target is 3-4% below Shell by 2028, not 8%. Each 1% narrowing recovers ~Rs 30-35/L.",
            "If Shell raises prices due to base oil cost pressures, PSO has immediate cover to follow -- "
            "the percentage gap is maintained, so the absolute price difference actually grows.",
            "For Carient Ultra 0W-20 and 5W-30, the 8% discount may already be too deep given product quality. "
            "Consider narrowing to 5% in FY26 and redirecting recovered margin into workshop visibility programmes.",
            "Monitor Shell pricing quarterly via the Daraz.pk scrape. Any Shell price movement triggers an immediate F1 recalculation.",
        ],
    },
    "F2": {
        "n": 7,
        "how": [
            "Collects all non-PSO Daraz.pk listings for each grade x pack combination from the scraped dataset.",
            "Computes the market median (50th percentile price per litre) -- the price where half the "
            "competitors charge more and half charge less.",
            "Positions PSO at exactly 3% below median: F2 = Market_Median x 0.97.",
            "When fewer than 3 exact-pack listings exist, scope widens to all packs for that grade (flagged as 'Grade-wide').",
        ],
        "rationale": [
            "The market median is the most transparent benchmark -- directly observable by any buyer "
            "comparing options online or in a workshop. It cannot be disputed.",
            "3% below median (not 20% below) avoids the 'cheap brand' trap. FMCG research shows extreme "
            "discounting signals inferior quality more than value.",
            "Being 3% below median means PSO wins the tie-break: same spec, lower price, wins the shelf.",
            "F2 carries the highest weight (30%) because it is the most current, market-facing signal and "
            "automatically absorbs competitor price movements without requiring model recalibration.",
        ],
        "implications": [
            "Recalculate F2 quarterly using fresh Daraz scrape data. If competitor prices rise, "
            "F2 will show PSO can raise prices without losing competitive position.",
            "The 3% discount is a policy choice. If PSO builds stronger brand equity, reduce it to 1-2%, "
            "capturing significant margin without volume loss.",
            "For 'Grade-wide' scope SKUs, the F2 result is less reliable -- prioritise getting pack-specific "
            "competitor data before finalising list prices for those SKUs.",
            "F2 sets the baseline for F3 (pack architecture) and F6 (geographic). An error in F2 propagates "
            "downstream: data quality in the competitor scrape is the single most critical input.",
        ],
    },
    "F3": {
        "n": 7,
        "how": [
            "Designates the 4L pack as the anchor (multiplier = 1.00x). All other pack sizes carry "
            "a multiplier relative to 4L anchor price per litre.",
            "The 4L market median (from competitor data for each grade) is the anchor price/L.",
            "Each pack size price = 4L anchor x PPA multiplier. Example: 1L pack = 4L anchor x 1.40 (40% premium per litre).",
            "Multipliers derived from international retail lubricants practice: smaller packs = convenience premium, "
            "larger packs = bulk discount.",
        ],
        "rationale": [
            "Without a structured PPA, brands accidentally price 1L packs lower per litre than 3L, "
            "confusing price-aware buyers who calculate unit costs.",
            "The 1L premium (40% above 4L per litre) is justified by: single engine-change convenience, "
            "retail display slot cost, impulse-purchase pricing, and high turnover at petrol station forecourts.",
            "A consistent price ladder prevents pack cannibalism: if 3L packs were priced at par with 4L "
            "per litre, buyers would always choose 3L, destroying 4L volumes and margin.",
            "PSO's range spans 1L to 4L across same brand/grade. Buyers comparing cross-pack will notice "
            "inconsistency, and it damages credibility.",
        ],
        "implications": [
            "Biggest revenue opportunity: 1L PCMO pack at petrol stations, where buyers are least price-sensitive "
            "and most convenience-driven. The 40% per-litre premium is fully justifiable and should be enforced.",
            "Review current 1L and 3L price points against the PPA ladder. "
            "The next list price revision is the right moment to re-anchor the entire range.",
            "Any new pack size launch (e.g., 800ml sachet) must have its PPA multiplier defined BEFORE launch. "
            "Without pre-setting the multiplier, the new pack will almost certainly be mispriced.",
            "For MCO (motorcycle oil), the 1L pack dominates sales. The 1.40x multiplier should be enforced -- "
            "MCO buyers are highly brand-loyal once a workshop mechanic recommends a product.",
        ],
    },
    "F4": {
        "n": 12,
        "how": [
            "Builds the minimum viable price from cost structure upward, not downward from the market.",
            "Step 1 -- Cost build-up: Base oil import parity (Group I Rs 280/L, II Rs 420/L, III Rs 680/L) "
            "+ Additive package (grade-specific) + Packaging (by pack size) = Total Manufacturing Cost.",
            "Step 2 -- Pocket price floor: Total Cost / (1 - Target Margin%). "
            "Targets: 45% Super Premium, 38% Premium, 30% Mainstream, 22% Economy.",
            "Step 3 -- List price floor: Pocket Price Floor / (1 - 20% Fleet Discount). "
            "Fleet is the deepest channel, so the list price must cover this to maintain margin everywhere.",
            "F4 is a hard floor -- if the weighted synthesis falls below F4, the recommendation is raised to F4.",
        ],
        "rationale": [
            "The McKinsey Price Waterfall reveals how revenue 'leaks' from list price to the price that "
            "actually reaches PSO's P&L ('pocket price'). Without this, small concessions destroy profitability collectively.",
            "PSO's fleet channel is especially exposed: 20% off a list price set too low means PSO is "
            "subsidising fleet customers from its own margin -- an unsustainable practice.",
            "Target margins (45% Super Premium to 22% Economy) reflect international lubricant industry norms "
            "adjusted for Pakistan's distribution structure, dealer margins, and working capital costs.",
            "F4 also sets the minimum for fleet negotiation. Any deal deeper than 20% off list requires "
            "explicit board approval because it breaches the pocket price floor.",
        ],
        "implications": [
            "F4 fires most often for economy-tier HDEO (20W-50, 15W-40 mineral) where Korean/Chinese imports "
            "on Daraz drive market prices below PSO's cost floor. Do NOT match these prices.",
            "Instead of cutting price below F4: (a) exit commodity packs where margin is structurally impossible, "
            "(b) reposition to 'value premium' with better packaging, or "
            "(c) explore private-label manufacturing for fleet-only packs.",
            "If F4 floor > F7 floor for any grade: the margin target is correctly set above pure cost floor -- good. "
            "If F4 floor < F7 floor: the margin assumption is too low and needs upward revision.",
            "Re-run F4 whenever base oil import prices shift more than 10% -- approximately every 6 months "
            "in a stable environment, more frequently during crude oil price shocks.",
        ],
    },
    "F5": {
        "n": 7,
        "how": [
            "Anchors to the economy-tier (Kixx/Kixx G1) Daraz.pk median for each grade as the baseline -- "
            "what the same viscosity grade costs with the minimum viable API specification.",
            "Applies an API/ACEA specification uplift based on PSO's tier: "
            "+28% for API SP fully synthetic, +18% for API SN+ semi-synthetic, +8% for API SN mineral, 0% for economy.",
            "Formula: F5 Price = Economy_Baseline x (1 + Spec_Uplift%)",
            "Uplift percentages derived from international market research on OEM approval premiums in "
            "comparable emerging markets (India, Vietnam, Indonesia).",
        ],
        "rationale": [
            "Engine oil pricing must reflect specification level, not just viscosity grade. "
            "A 5W-40 API SP (fully synthetic) has fundamentally different chemistry and engine protection "
            "vs a 5W-40 API SN mineral -- they are not comparable products.",
            "In Pakistan, pump attendants and workshop mechanics rarely explain spec differences. "
            "F5 sets the price SIGNAL that communicates quality: consumers associate higher price with higher spec.",
            "F5 is weighted at 10% because spec literacy in Pakistan is still developing. "
            "As car parc modernises (more Euro-4/5 engines, turbo/hybrid vehicles), "
            "the spec premium will become more important and the weight should increase.",
            "F5 prevents underpricing of premium products: Carient Ultra 5W-40 (API SP) should never "
            "approach Kixx G1 5W-40 (API SN) pricing, regardless of competitive pressure.",
        ],
        "implications": [
            "Investing in OEM approvals (VW 504.00, BMW Longlife-04, Toyota WS) directly increases the "
            "justifiable spec premium from 28% toward 35-40% for those SKUs. "
            "A single major OEM approval generates years of pricing power.",
            "For economy HDEO (DEO 3000, Dieselube), F5 uplift is 0% -- these compete purely on price. "
            "Any price above the F5 baseline must be supported by brand or distribution advantages, not spec.",
            "Monitor economy baseline (Kixx pricing) closely. If Korean brands discount aggressively, "
            "the F5 baseline drops and PSO's spec premium erodes unless brand investment is maintained.",
            "Display API specification prominently on packaging and POS materials. "
            "Consumers who understand spec buy on spec, not just price -- this shifts F5 weight from 10% toward 20%.",
        ],
    },
    "F6": {
        "n": 11,
        "how": [
            "Uses PSO's own FY25 lubricants volume data (Lubes Data Final.xlsx) to compute regional YoY growth signals.",
            "Regions growing faster than the national average can support a price premium (Central DEO: +5%). "
            "Declining regions require a discount to defend volume (North MCO: -8%).",
            "Formula: Regional Price = F2_National_Price x (1 + Regional_Signal%)",
            "South (Karachi/Hyderabad) is the national baseline (0%). "
            "Adjustments are quarterly promotional price variations -- NOT permanent list price changes.",
        ],
        "rationale": [
            "Pakistan's lubricants market is deeply regional: Central Punjab drives HDEO (agricultural/transport), "
            "Karachi dominates PCMO, KPK/Northern corridor has distinct MCO patterns.",
            "A flat national price leaves revenue on the table in strong-growth regions (demand is relatively inelastic) "
            "and loses volume in weak regions (buyers switch to cheaper local alternatives).",
            "Geographic pricing is standard practice for CPG and lubricants companies in South Asia. "
            "Castrol, Shell, and Total all maintain regional trade price lists in Pakistan.",
            "Using PSO's own volume data (not external estimates) ties the signal to PSO's distribution reality -- "
            "it reflects where PSO is genuinely growing or losing ground.",
        ],
        "implications": [
            "For HDEO in Central Punjab (+5%): translates to ~Rs 75-90/L higher price on 15W-40 4L. "
            "Recoverable through regional promotional pricing without disturbing the national list.",
            "For MCO in North (-8%): this is a defensive measure. North MCO decline suggests PSO is losing "
            "to local blenders. Pricing alone will not recover volume -- conduct a distribution audit.",
            "Grey market risk: if South-to-North price gap on the same SKU exceeds Rs 50-70/L, "
            "distributors will buy cheap in South and sell in North. Cap regional variation at 5% on high-volume SKUs.",
            "Update regional signals every 6 months using the latest PSO sales data. "
            "The FY25 data will be stale by Q3 FY26 and must be refreshed.",
        ],
    },
    "F7": {
        "n": 8,
        "how": [
            "Takes the base oil import parity cost for the grade's required group "
            "(Group I: Rs 280/L, Group II: Rs 420/L, Group III: Rs 680/L).",
            "Adds the grade-specific additive package cost and a fixed blending/overhead allocation of Rs 50/L.",
            "Multiplies total manufacturing cost by 2.8x -- the standard retail lubricants markup in Pakistan "
            "(equivalent to ~64% gross margin).",
            "Formula: F7 Floor = (Base_Oil_Cost + Additive_Cost + Rs 50) x 2.8",
            "F7 is a standalone sanity check and audit tool -- NOT a synthesis input.",
        ],
        "rationale": [
            "Base oil is the primary raw material in any lubricant (55-75% of total product cost). "
            "Its price tracks crude oil with a 6-8 week lag, creating predictable cost movements.",
            "The 2.8x markup is the industry standard for retail lubricants in Pakistan, derived from: "
            "distributor margin (15%), dealer/pump margin (8%), PSO overhead (12%), and net profit target (10%).",
            "F7 provides an independent floor that does not depend on competitor data (F2) or margin targets (F4). "
            "Even if competitive data is stale or targets are mis-set, F7 catches structural underpricing.",
            "F7 is especially valuable during crude oil price shocks: when base oil import costs spike, "
            "F7 immediately shows which PSO products are priced below manufacturing cost.",
        ],
        "implications": [
            "Any PSO SKU priced below its F7 floor is sold at a structural loss. "
            "This triggers an immediate price revision or product exit decision.",
            "If F7 > F4 for any SKU, the F4 margin target is too low -- revise it upward for that tier.",
            "Monitor base oil import prices monthly (PARCO, NRL, Korean import indices). "
            "A 10% crude oil increase typically means 8-12% increase in F7 floor values within 6-8 weeks.",
            "Group III (5W-40, 5W-30, 0W-20) is the most volatile and entirely imported. "
            "Consider hedging 30-60 days of base oil inventory to protect Super Premium pricing from cost shocks.",
        ],
    },
    "F8": {
        "n": 9,
        "how": [
            "Takes the final recommended list price from the weighted synthesis (Summary sheet) as the starting point.",
            "Applies a structured discount cascade for each trade channel: "
            "Retail Pump = 0%, Workshop = -12%, Distributor = -15%, Fleet Contract = -20%.",
            "Formula: Channel Price = List Price x (1 - Discount%)",
            "The fleet contract price is the critical floor -- it must never fall below the F4 cost waterfall floor.",
        ],
        "rationale": [
            "Different trade channels have fundamentally different cost structures, volume commitments, and credit terms. "
            "A single price for all channels either overcharges workshops (losing trade) or undercharges fleet (losing margin).",
            "Workshop -12%: they stock product and drive recommendations but have no volume commitment. "
            "12% is the minimum trade allowance to secure shelf placement and mechanic recommendation.",
            "Distributor -15%: they take inventory risk, provide credit to dealers, manage last-mile logistics. "
            "15% covers their operating costs and provides a viable business model.",
            "Fleet -20%: they commit to annual volumes and provide predictable revenue but negotiate aggressively "
            "and pay in 45-90 day cycles. 20% compensates for working capital cost and volume commitment.",
            "A structured discount matrix prevents channel conflict: each account type knows their price "
            "and cannot arbitrage because the discount is tied to account type, not transaction size.",
        ],
        "implications": [
            "Critical constraint: Fleet contract price (list x 0.80) must always exceed the F4 pocket price floor. "
            "If not, PSO is subsidising fleet customers from its own margin -- unsustainable.",
            "No channel should receive more than 20% off list without board-level approval. "
            "Any 'special deal' deeper than fleet discount erodes the entire price architecture.",
            "Workshops are PSO's most strategic brand advocacy channel. The 12% discount is a relationship investment. "
            "Consider tiering: 12% standard, 14% for top-volume workshops, 16% for PSO-branded workshop partners.",
            "Fleet contracts must be reviewed annually against the UPDATED synthesis price, not the prior year's list. "
            "If list prices rise 8% due to base oil inflation but fleet contracts are locked at prior list, "
            "PSO absorbs the entire cost increase at its most margin-dilutive channel.",
        ],
    },
}

# ==================================================================
#  Style helpers
# ==================================================================

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, size=9, color="111111"):
    return Font(bold=bold, size=size, color=color, name="Calibri")
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left(wrap=True): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
def _right(): return Alignment(horizontal="right", vertical="center")
def cref(row, col_n): return f"${get_column_letter(col_n)}${row}"

def title_row(ws, row, text, n_cols, bg=C_DARK, size=13):
    ws.row_dimensions[row].height = 28
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=size, color=C_WHITE, name="Calibri")
    c.fill = _fill(bg)
    c.alignment = _left(wrap=False)

def section_row(ws, row, text, n_cols, bg=C_MED):
    ws.row_dimensions[row].height = 18
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=9, color=C_WHITE, name="Calibri")
    c.fill = _fill(bg)
    c.alignment = _left(wrap=False)

def desc_row(ws, row, text, n_cols, bg="FFFFFF", h=40):
    ws.row_dimensions[row].height = h
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=8.5, italic=True, color="555555", name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    if bg != "FFFFFF":
        c.fill = _fill(bg)

def header_row(ws, row, headers, start_col=1, bg=C_DARK):
    ws.row_dimensions[row].height = 17
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(bold=True, size=8.5, color=C_WHITE, name="Calibri")
        c.fill = _fill(bg)
        c.alignment = _center()
        c.border = _border()

def data_cell(ws, row, col_n, value, fmt=None, bold=False, bg=None, align="right"):
    c = ws.cell(row=row, column=col_n, value=value)
    c.font = Font(bold=bold, size=8.5, name="Calibri")
    c.border = _border()
    c.alignment = _left() if align == "left" else _right()
    if fmt == "pkr":    c.number_format = "#,##0"
    elif fmt == "pct":  c.number_format = "0%"
    elif fmt == "pct1": c.number_format = "0.0%"
    elif fmt == "dec":  c.number_format = "0.00"
    if bg: c.fill = _fill(bg)
    return c

def formula_cell(ws, row, col_n, formula, fmt="pkr", bg=None, bold=False):
    c = ws.cell(row=row, column=col_n, value=formula)
    c.font = Font(bold=bold, size=8.5, color=C_BLUE if not bg else "111111", name="Calibri")
    c.border = _border()
    c.alignment = _right()
    if fmt == "pkr":   c.number_format = "#,##0"
    elif fmt == "pct": c.number_format = "0%"
    elif fmt == "pct1":c.number_format = "0.0%"
    elif fmt == "dec": c.number_format = "0.00"
    if bg: c.fill = _fill(bg)
    return c

def blank_row(ws, row, h=8):
    ws.row_dimensions[row].height = h

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_methodology(ws, start_row, n_cols, how=None, rationale=None, implications=None,
                    how_it_works=None):
    how_it_works = how_it_works or how or []
    rationale    = rationale    or []
    implications = implications or []
    # shadow outer name so write_block closure works
    _how_it_works = how_it_works; _rationale = rationale; _implications = implications
    blank_row(ws, start_row, h=10)
    r = start_row + 1

    def write_block(ws, row, label, bullets, header_bg, body_bg, text_color=C_WHITE):
        ws.row_dimensions[row].height = 17
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=True, size=9, color=text_color, name="Calibri")
        c.fill = _fill(header_bg)
        c.alignment = _left(wrap=False)
        text = "\n".join(f"  *  {b}" for b in bullets)
        lines = sum(max(1, len(b) // 110 + 1) for b in bullets) + 1
        ws.row_dimensions[row + 1].height = max(22, lines * 13)
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=n_cols)
        cb = ws.cell(row=row + 1, column=1, value=text)
        cb.font = Font(size=9, color="111111", name="Calibri")
        cb.fill = _fill(body_bg)
        cb.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        return row + 2

    r = write_block(ws, r, "HOW IT WORKS", how_it_works, C_DARK, "F4FAF7")
    ws.row_dimensions[r].height = 5; r += 1
    r = write_block(ws, r, "RATIONALE  --  Why this framework is used", rationale, C_MED, "F0FAF5")
    ws.row_dimensions[r].height = 5; r += 1
    r = write_block(ws, r, "IMPLICATIONS  --  What this means for PSO pricing decisions",
                    implications, C_GOLD, "FEF9ED", text_color="111111")

# ==================================================================
#  Data loaders
# ==================================================================

def load_comp_df():
    con = sqlite3.connect(DB_PATH)
    rows = con.execute("""
        SELECT brand_detected, grade_detected, pack_size_l, oil_type,
               price_per_litre, price, title, platform
        FROM scraped_products
        WHERE price_per_litre > 100 AND grade_detected IS NOT NULL
          AND brand_detected IS NOT NULL
          AND DATE(scraped_at) = (SELECT DATE(MAX(scraped_at)) FROM scraped_products)
        ORDER BY oil_type, grade_detected, brand_detected, pack_size_l
    """).fetchall()
    con.close()
    return pd.DataFrame(rows, columns=["Brand","Grade","Pack_L","Oil_Type",
                                        "Price_L","Total_Price","Title","Platform"])

def load_pso_skus():
    import openpyxl as xl
    wb = xl.load_workbook(str(LUBES_FILE), data_only=True)
    ws = wb["SKU Wise"]
    skus, current = [], None
    section_hdrs = {
        "Vol vs Val (PCMO)", "Vol vs Val (MCO)",
        "Vol vs Val (DEO)", "Vol vs Val (Industrial)"
    }
    for row in ws.iter_rows(values_only=True):
        v0 = str(row[0]).strip() if row[0] else ""
        if v0 in section_hdrs: current = None; continue
        if v0 in ("Brand", "None"): continue
        if v0: current = v0
        if current not in BRAND_MAP: continue
        brand, oil_type = BRAND_MAP[current]
        try:
            pack_l = float(row[1]) if row[1] else None
            vol    = float(row[3]) if row[3] else 0.0
        except Exception: continue
        if not pack_l or pack_l > 4.0 or vol <= 0: continue
        for grade in BRAND_GRADES.get(brand, []):
            skus.append((brand, grade, pack_l, oil_type))
    seen = set(); unique = []
    for s in skus:
        if s not in seen: seen.add(s); unique.append(s)
    return unique

def mkt_stats(df, grade, pack_l=None, excl_pso=True):
    m = df["Grade"] == grade
    if pack_l: m &= df["Pack_L"] == pack_l
    if excl_pso: m &= ~df["Brand"].str.contains("PSO|Carient", na=False)
    vals = df.loc[m, "Price_L"].dropna()
    widened = False
    if vals.empty and pack_l:
        m2 = df["Grade"] == grade
        if excl_pso: m2 &= ~df["Brand"].str.contains("PSO|Carient", na=False)
        vals = df.loc[m2, "Price_L"].dropna()
        widened = True
    if vals.empty: return None, None, None, 0, widened
    return (round(float(vals.min()),0), round(float(vals.median()),0),
            round(float(vals.max()),0), len(vals), widened)

def tier_med(df, tier, grade, pack_l=None):
    brands = SHELL_TIER_BRANDS.get(tier, [])
    m = df["Grade"] == grade
    if pack_l: m &= df["Pack_L"] == pack_l
    m &= df["Brand"].isin(brands)
    vals = df.loc[m, "Price_L"].dropna()
    if vals.empty and pack_l:
        m2 = (df["Grade"] == grade) & df["Brand"].isin(brands)
        vals = df.loc[m2, "Price_L"].dropna()
    return round(float(vals.median()),0) if not vals.empty else None

# ==================================================================
#  Sheet builders
# ==================================================================

def build_readme(wb):
    ws = wb.create_sheet("README", 0)
    N = 8
    title_row(ws, 1, "PSO LUBRICANTS -- PRICING STRATEGY WORKBOOK v2", N, size=14)
    desc_row(ws, 2, "Data: Daraz.pk scrape (365 listings) + PSO Lubes Data Final.xlsx | July 2026 | 8 Pricing Frameworks | 45 SKUs", N)
    blank_row(ws, 3)
    section_row(ws, 4, "SHEET GUIDE", N)
    sheets_guide = [
        ("README",                "This overview and navigation guide"),
        ("Competitor Data",       "365 motor oil listings scraped from Daraz.pk -- basis for all market benchmarks"),
        ("F1 - Value Tiering",    "Shell-equivalent benchmarking. F1 = Shell tier median x (1 - brand discount). Weight: 30%"),
        ("F2 - Market Ref Price", "Market median price x 0.97 positioning. Weight: 30%"),
        ("F3 - Pack Architecture","Price ladder by pack size. F3 = 4L anchor x PPA multiplier. Weight: 20%"),
        ("F4 - Cost Waterfall",   "McKinsey cost build-up. Hard floor -- overrides synthesis if violated."),
        ("F5 - Spec Premium",     "API/ACEA specification uplift. F5 = Economy baseline x (1 + spec%). Weight: 10%"),
        ("F6 - Geographic",       "Regional price adjustments from PSO volume data. Quarterly promotional overlay."),
        ("F7 - Base Oil Floor",   "Sanity check floor: F7 = (Base oil + Additive + Rs 50) x 2.8x markup"),
        ("F8 - Channel Pricing",  "Channel cascade: Retail 0% / Workshop -12% / Distributor -15% / Fleet -20%"),
        ("Summary",               "All 45 SKUs with every framework price, weighted synthesis, and final recommendation"),
    ]
    header_row(ws, 5, ["Sheet", "Description"])
    for i, (sheet, desc) in enumerate(sheets_guide, 6):
        ws.row_dimensions[i].height = 20
        data_cell(ws, i, 1, sheet, align="left", bold=True, bg=C_LIGHT if i%2==0 else None)
        data_cell(ws, i, 2, desc, align="left", bg=C_LIGHT if i%2==0 else None)

    blank_row(ws, len(sheets_guide)+7)
    section_row(ws, len(sheets_guide)+8, "FORMULA REFERENCE", N)
    formulas = [
        ("F1 Value Tiering",   "= Shell_Tier_Median x (1 - Brand_Discount%)",
         "8%/6%/4%/2% discount by tier; Shell benchmark from Daraz medians"),
        ("F2 Market Ref",      "= Market_Median x 0.97",
         "3% below market median = best value, not cheapest"),
        ("F3 Pack Arch",       "= 4L_Market_Anchor x PPA_Multiplier",
         "Multiplier: 1L=1.40x, 3L=1.12x, 4L=1.00x (anchor)"),
        ("F4 Waterfall",       "= (BO + Additive + Pkg) / (1 - Margin%) / (1 - 20%)",
         "Pocket price floor then divide by fleet discount"),
        ("F5 Spec Premium",    "= Economy_Baseline x (1 + Spec_Uplift%)",
         "Uplift: Super Prem 28%, Premium 18%, Mainstream 8%"),
        ("F6 Geographic",      "= F2_National x (1 + Regional_Signal%)",
         "Central +2-5%, North -2-8%, South = baseline"),
        ("F7 Base Oil Floor",  "= (BO_Cost + Additive + 50) x 2.8",
         "2.8x = industry retail lubricants markup multiple"),
        ("F8 Channel",         "= List_Price x (1 - Channel_Discount%)",
         "0% / 12% / 15% / 20% off list by channel"),
        ("Synthesis",          "= SUM(F_i x Weight_i) / SUM(Weight_i)",
         "F1=30%, F2=30%, F3=20%, F5=10%, F4=10%"),
        ("Final Rec",          "= MAX(Synthesis, F4_Floor)",
         "Hard floor ensures margin is never destroyed"),
    ]
    r = len(sheets_guide) + 9
    header_row(ws, r, ["Framework", "Formula Pattern", "Notes"])
    r += 1
    for fw, formula, note in formulas:
        ws.row_dimensions[r].height = 18
        data_cell(ws, r, 1, fw, align="left", bold=True)
        data_cell(ws, r, 2, formula, align="left", bg="EEF7FF")
        data_cell(ws, r, 3, note, align="left")
        r += 1

    set_widths(ws, [28, 65, 55])
    ws.freeze_panes = "A2"


def build_competitor_data(wb, df):
    ws = wb.create_sheet("Competitor Data")
    N = 9
    title_row(ws, 1, "COMPETITOR PRICING DATA -- Daraz.pk Scrape (July 2026)", N)
    desc_row(ws, 2,
             "365 motor oil listings scraped from Daraz.pk, normalised to Price/Litre (PKR) for cross-pack comparison. "
             "Source of truth for all F1, F2, F3, and F5 framework calculations. "
             "PSO listings are highlighted but excluded from market benchmark statistics.", N, h=35)
    blank_row(ws, 3)
    header_row(ws, 4, ["#","Brand","Grade","Pack (L)","Oil Type",
                        "Price/L (PKR)","Total Price (PKR)","Product Title","Platform"])
    for idx, row in df.iterrows():
        r = idx + 5
        ws.row_dimensions[r].height = 15
        is_pso = "PSO" in str(row["Brand"]) or "Carient" in str(row["Brand"])
        bg = "FFF8E6" if is_pso else None
        data_cell(ws, r, 1, idx+1, align="right", bg=bg)
        data_cell(ws, r, 2, row["Brand"], align="left", bg=bg)
        data_cell(ws, r, 3, row["Grade"], align="left", bg=bg)
        data_cell(ws, r, 4, row["Pack_L"], fmt="dec", bg=bg)
        data_cell(ws, r, 5, row["Oil_Type"].upper(), align="left", bg=bg)
        data_cell(ws, r, 6, row["Price_L"], fmt="pkr", bg=bg)
        data_cell(ws, r, 7, row["Total_Price"], fmt="pkr", bg=bg)
        data_cell(ws, r, 8, row["Title"], align="left", bg=bg)
        data_cell(ws, r, 9, row["Platform"].title(), align="left", bg=bg)

    last = len(df) + 5
    blank_row(ws, last)
    section_row(ws, last+1, "MARKET STATISTICS BY GRADE (excluding PSO)", N)
    header_row(ws, last+2, ["Grade","Oil Type","Listings","Min/L","Median/L","Max/L","Std Dev"])
    r = last + 3
    for grade in ["0W-20","5W-20","5W-30","5W-40","10W-40","15W-40","20W-50"]:
        mn, med, mx, cnt, _ = mkt_stats(df, grade)
        m2 = (df["Grade"]==grade) & ~df["Brand"].str.contains("PSO|Carient", na=False)
        std = round(float(df.loc[m2,"Price_L"].std()), 0) if cnt > 1 else None
        oil = "PCMO" if grade in ["0W-20","5W-20","5W-30","5W-40","10W-40"] else "HDEO/PCMO"
        data_cell(ws, r, 1, grade, align="left")
        data_cell(ws, r, 2, oil, align="left")
        data_cell(ws, r, 3, cnt)
        data_cell(ws, r, 4, mn, fmt="pkr")
        data_cell(ws, r, 5, med, fmt="pkr", bold=True)
        data_cell(ws, r, 6, mx, fmt="pkr")
        data_cell(ws, r, 7, std, fmt="pkr")
        r += 1
    set_widths(ws, [5, 22, 9, 8, 10, 12, 14, 65, 10])
    ws.freeze_panes = "A5"


def build_f1(wb, df, skus):
    ws = wb.create_sheet("F1 - Value Tiering")
    N = 8
    title_row(ws, 1, "F1 -- VALUE-BASED TIERING  (Weight: 30%)", N)
    desc_row(ws, 2,
             "PSO Carient is benchmarked against the Shell tier equivalent (Super Premium = Helix Ultra, "
             "Premium = HX7/ZIC X9, Mainstream = ZIC X7/Total 9000, Economy = Kixx). "
             "A brand-perception discount is applied because PSO brand equity is currently below Shell. "
             "FORMULA: F1 Price = Shell_Tier_Median x (1 - Brand_Discount%)", N, h=50)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE 1 -- Shell Tier Benchmark Prices (Daraz.pk medians, PKR/L)", N)
    grades = ["0W-20","5W-20","5W-30","5W-40","10W-40","15W-40","20W-50"]
    header_row(ws, 5, ["Tier","Shell Benchmark Brands"] + grades)
    r = 6
    for tier, brands in SHELL_TIER_BRANDS.items():
        data_cell(ws, r, 1, TIER_LABEL[tier], align="left", bold=True)
        data_cell(ws, r, 2, ", ".join(brands), align="left")
        for i, g in enumerate(grades, 3):
            p = tier_med(df, tier, g)
            data_cell(ws, r, i, p, fmt="pkr")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "PARAMETER TABLE 2 -- PSO Brand Perception Discount vs Shell", N)
    r += 1
    header_row(ws, r, ["Tier","Discount %","Rationale"])
    r += 1
    disc_map = {}
    for tier, disc in PSO_BRAND_DISCOUNT.items():
        rat = {
            "super_premium": "Significant gap -- fully synthetic brand communication is nascent in Pakistan",
            "premium":       "Moderate gap -- HX7-class perception; closing as fleet specs improve",
            "mainstream":    "Small gap -- mineral lubricants mostly price-competitive",
            "economy":       "Minimal gap -- budget buyers are price-led, brand matters less",
        }[tier]
        disc_map[tier] = (r, 2)
        data_cell(ws, r, 1, TIER_LABEL[tier], align="left", bold=True)
        data_cell(ws, r, 2, disc, fmt="pct")
        data_cell(ws, r, 3, rat, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "CALCULATION TABLE -- F1 Price per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Pack (L)","Tier",
                        "Shell Benchmark (PKR/L)","Brand Discount","F1 Price (PKR/L)"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        tier = PSO_TIER.get(brand, "mainstream")
        shell_p = tier_med(df, tier, grade, pack_l)
        disc = PSO_BRAND_DISCOUNT.get(tier, 0.04)
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, pack_l, fmt="dec")
        data_cell(ws, r, 5, TIER_LABEL[tier], align="left")
        data_cell(ws, r, 6, shell_p, fmt="pkr")
        data_cell(ws, r, 7, disc, fmt="pct")
        if shell_p:
            formula_cell(ws, r, 8, f"={cref(r,6)}*(1-{cref(r,7)})", fmt="pkr")
        else:
            data_cell(ws, r, 8, "N/A", align="right")
        r += 1

    add_methodology(ws, r, N, **{k: METHODOLOGY["F1"][k] for k in ("how","rationale","implications")})
    set_widths(ws, [28, 9, 8, 8, 15, 20, 15, 18])
    ws.freeze_panes = "A6"


def build_f2(wb, df, skus):
    ws = wb.create_sheet("F2 - Market Ref Price")
    N = 8
    title_row(ws, 1, "F2 -- COMPETITIVE REFERENCE PRICE  (Weight: 30%)", N)
    desc_row(ws, 2,
             "F2 anchors PSO prices to the independent market median -- the price at which "
             "half the competitors charge more and half charge less. PSO is positioned 3% below median "
             "(best-value-in-class, not cheapest). "
             "FORMULA: F2 Price = Market_Median x 0.97", N, h=45)
    blank_row(ws, 3)

    section_row(ws, 4, "MARKET STATISTICS TABLE (Daraz.pk, July 2026 -- PSO/Carient excluded)", N)
    header_row(ws, 5, ["Grade","Pack (L)","Listings","Min (PKR/L)","Median (PKR/L)",
                        "Max (PKR/L)","Scope","F2 = Median x 0.97"])
    r = 6
    stat_rows = {}
    combos = sorted(set((g, p) for _, g, p, _ in skus))
    for grade, pack_l in combos:
        mn, med, mx, cnt, widened = mkt_stats(df, grade, pack_l)
        stat_rows[(grade, pack_l)] = (r, 5)
        data_cell(ws, r, 1, grade, align="left")
        data_cell(ws, r, 2, pack_l, fmt="dec")
        data_cell(ws, r, 3, cnt)
        data_cell(ws, r, 4, mn, fmt="pkr")
        data_cell(ws, r, 5, med, fmt="pkr", bold=True)
        data_cell(ws, r, 6, mx, fmt="pkr")
        data_cell(ws, r, 7, "Grade-wide" if widened else "Exact pack", align="left")
        formula_cell(ws, r, 8, f"={cref(r,5)}*0.97", fmt="pkr") if med else data_cell(ws, r, 8, "N/A")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "BRAND-LEVEL RESULTS -- F2 Price per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Pack (L)","Market Median (PKR/L)","F2 CRP (PKR/L)","vs Median"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        _, med, _, _, _ = mkt_stats(df, grade, pack_l)
        f2 = round(med * 0.97, 0) if med else None
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, pack_l, fmt="dec")
        data_cell(ws, r, 5, med, fmt="pkr")
        data_cell(ws, r, 6, f2, fmt="pkr")
        data_cell(ws, r, 7, -0.03 if f2 else None, fmt="pct1")
        r += 1

    add_methodology(ws, r, N, **{k: METHODOLOGY["F2"][k] for k in ("how","rationale","implications")})
    set_widths(ws, [28, 9, 8, 8, 18, 18, 12, 0])
    ws.freeze_panes = "A6"


def build_f3(wb, df, skus):
    ws = wb.create_sheet("F3 - Pack Architecture")
    N = 7
    title_row(ws, 1, "F3 -- PRICE-PACK ARCHITECTURE  (Weight: 20%)", N)
    desc_row(ws, 2,
             "The 4L pack is the anchor (multiplier = 1.00x). All other pack sizes are priced relative to "
             "the 4L market median. This creates a legible, consistent price ladder. "
             "FORMULA: F3 Price = 4L_Market_Median x PPA_Multiplier", N, h=40)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE -- PPA Multipliers (international retail lubricant norms)", N)
    header_row(ws, 5, ["Pack Size (L)","Multiplier vs 4L Anchor","Premium/(Discount)","Rationale"])
    r = 6
    mult_rows = {}
    for pack_l, mult in PPA_MULTIPLIER.items():
        mult_rows[pack_l] = (r, 2)
        rat = {
            0.8: "Top-up sachet -- convenience + dispensing premium",
            1.0: "Retail single-change -- most common passenger car oil change",
            1.5: "Partial change plus reserve",
            2.0: "DIY market -- two-thirds engine change",
            3.0: "Near-full PCMO change (3L sump common)",
            3.5: "Slight discount -- niche size",
            4.0: "ANCHOR -- full 4L engine sump change",
            5.0: "Slight bulk discount -- light commercial",
            7.0: "Bulk / workshop supply",
            10.0:"Distributor / fleet bulk",
            20.0:"20L drum -- workshop / small fleet",
        }.get(pack_l, "")
        data_cell(ws, r, 1, pack_l, fmt="dec", bold=(pack_l==4.0))
        data_cell(ws, r, 2, mult, fmt="dec", bold=(pack_l==4.0),
                  bg=C_LIGHT if pack_l==4.0 else None)
        data_cell(ws, r, 3, f"{(mult-1)*100:+.0f}%", align="right")
        data_cell(ws, r, 4, rat, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "4L ANCHOR PRICES BY GRADE (from competitor market medians)", N)
    r += 1
    header_row(ws, r, ["Grade","4L Market Median (PKR/L)","Data Points","Note"])
    r += 1
    anchor_rows = {}
    for grade in sorted(set(g for _, g, _, _ in skus)):
        _, med4, _, cnt4, widened = mkt_stats(df, grade, 4.0)
        anchor_rows[grade] = (r, 2)
        data_cell(ws, r, 1, grade, align="left", bold=True)
        data_cell(ws, r, 2, med4, fmt="pkr", bold=True)
        data_cell(ws, r, 3, cnt4)
        data_cell(ws, r, 4, "Widened to all packs" if widened else "Exact 4L", align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "CALCULATION TABLE -- F3 Price per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Pack (L)","4L Anchor (PKR/L)","Multiplier","F3 Price (PKR/L)"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        mult = PPA_MULTIPLIER.get(pack_l, 1.0)
        _, med4, _, _, _ = mkt_stats(df, grade, 4.0)
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, pack_l, fmt="dec")
        data_cell(ws, r, 5, med4, fmt="pkr")
        data_cell(ws, r, 6, mult, fmt="dec")
        formula_cell(ws, r, 7, f"={cref(r,5)}*{cref(r,6)}", fmt="pkr") if med4 else data_cell(ws, r, 7, "N/A")
        r += 1

    add_methodology(ws, r, N, **{k: METHODOLOGY["F3"][k] for k in ("how","rationale","implications")})
    set_widths(ws, [28, 9, 8, 8, 20, 12, 18])
    ws.freeze_panes = "A6"


def build_f4(wb, df, skus):
    ws = wb.create_sheet("F4 - Cost Waterfall")
    N = 12
    title_row(ws, 1, "F4 -- McKINSEY PRICE WATERFALL  (Hard Floor Constraint)", N)
    desc_row(ws, 2,
             "Calculates the minimum viable list price from cost structure upward. "
             "NOT a pricing target -- a hard floor. If F1/F2/F3 synthesis falls below F4, "
             "the recommendation is raised to F4. "
             "FORMULA: Pocket Floor = Total_Cost / (1 - Margin%)  -->  List Floor = Pocket / (1 - 20%)", N, h=50)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE 1 -- Base Oil Cost by Group (PKR/L, Pakistan import parity 2025)", N)
    header_row(ws, 5, ["Base Oil Group","Cost (PKR/L)","Typical Grades","Source"])
    r = 6
    for grp, cost in BASE_OIL_COST.items():
        grades_in = [g for g, gr in BASE_OIL_GROUP.items() if gr == grp]
        src = {"I": "SN150/BS150 -- mainly domestic",
               "II": "VHVI -- Korean/Middle East import",
               "III": "PAO/GTL -- European/Korean import"}[grp]
        data_cell(ws, r, 1, f"Group {grp}", align="left", bold=True)
        data_cell(ws, r, 2, cost, fmt="pkr")
        data_cell(ws, r, 3, ", ".join(grades_in), align="left")
        data_cell(ws, r, 4, src, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "PARAMETER TABLE 2 -- Additive Package Cost by Grade (PKR/L)", N)
    r += 1
    header_row(ws, r, ["Grade","Base Oil Group","Additive Cost (PKR/L)","Notes"])
    r += 1
    for grade, add_cost in ADDITIVE_COST.items():
        bog = BASE_OIL_GROUP.get(grade, "II")
        note = ("ACEA A5/SN+ -- expensive DI package" if bog == "III"
                else "SN/SL additive package" if bog == "II"
                else "SM/SL economy additive")
        data_cell(ws, r, 1, grade, align="left")
        data_cell(ws, r, 2, f"Group {bog}", align="left")
        data_cell(ws, r, 3, add_cost, fmt="pkr")
        data_cell(ws, r, 4, note, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "PARAMETER TABLE 3 -- Packaging Cost by Pack Size (PKR/L)", N)
    r += 1
    header_row(ws, r, ["Pack Size","Pkg Cost (PKR/L)","Rationale"])
    r += 1
    for label, cost, rat in [("<=1L", 60, "Small packs -- higher per-L moulding & labelling cost"),
                              ("1L-5L", 35, "Standard HDPE bottle"),
                              (">5L drum", 20, "Drum / gallon -- lower per-L container cost")]:
        data_cell(ws, r, 1, label, align="left")
        data_cell(ws, r, 2, cost, fmt="pkr")
        data_cell(ws, r, 3, rat, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "PARAMETER TABLE 4 -- Gross Margin Targets by Tier", N)
    r += 1
    header_row(ws, r, ["Tier","Target Pocket Margin %","Rationale"])
    r += 1
    margin_rat = {
        "super_premium": "Fully synthetic -- high spec, brand premium, lower vol --> needs 45% to fund A&P",
        "premium":       "Semi-synthetic -- balanced vol/value, 38% industry norm for mid-tier",
        "mainstream":    "Mineral + partial synthetic -- volume-driven, 30% sustains distribution margin",
        "economy":       "Commodity mineral -- price-led, 22% floor preserves channel viability",
    }
    for tier, margin in TIER_MARGIN.items():
        data_cell(ws, r, 1, TIER_LABEL[tier], align="left", bold=True)
        data_cell(ws, r, 2, margin, fmt="pct")
        data_cell(ws, r, 3, margin_rat[tier], align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "WATERFALL CALCULATION -- per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Grade","Pack (L)","Tier","BO Cost","Additive","Pkg",
                        "Total Cost","Target Margin","Pocket Floor","Fleet Disc","List FLOOR"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        tier = PSO_TIER.get(brand, "mainstream")
        bog  = BASE_OIL_GROUP.get(grade, "II")
        bo_c = BASE_OIL_COST.get(bog, 420)
        add_c = ADDITIVE_COST.get(grade, 140)
        pkg_c = 60 if pack_l <= 1 else 35 if pack_l <= 5 else 20
        margin = TIER_MARGIN.get(tier, 0.30)
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, grade, align="left")
        data_cell(ws, r, 3, pack_l, fmt="dec")
        data_cell(ws, r, 4, TIER_LABEL[tier], align="left")
        data_cell(ws, r, 5, bo_c, fmt="pkr")
        data_cell(ws, r, 6, add_c, fmt="pkr")
        data_cell(ws, r, 7, pkg_c, fmt="pkr")
        formula_cell(ws, r, 8, f"={cref(r,5)}+{cref(r,6)}+{cref(r,7)}", fmt="pkr")
        data_cell(ws, r, 9, margin, fmt="pct")
        formula_cell(ws, r, 10, f"={cref(r,8)}/(1-{cref(r,9)})", fmt="pkr")
        data_cell(ws, r, 11, 0.20, fmt="pct")
        formula_cell(ws, r, 12, f"={cref(r,10)}/(1-{cref(r,11)})", fmt="pkr", bold=True)
        r += 1

    add_methodology(ws, r, N, **{k: METHODOLOGY["F4"][k] for k in ("how","rationale","implications")})
    set_widths(ws, [26, 8, 8, 15, 10, 10, 8, 12, 13, 13, 12, 16])
    ws.freeze_panes = "A14"


def build_f5(wb, df, skus):
    ws = wb.create_sheet("F5 - Spec Premium")
    N = 7
    title_row(ws, 1, "F5 -- OEM / API SPECIFICATION PREMIUM  (Weight: 10%)", N)
    desc_row(ws, 2,
             "Anchors to economy-tier (Kixx/Kixx G1) market price per grade as baseline. "
             "Applies spec uplift: 28% fully synthetic (API SP), 18% semi-synthetic (API SN+), "
             "8% mineral (API SN), 0% economy. "
             "FORMULA: F5 Price = Economy_Baseline x (1 + Spec_Uplift%)", N, h=45)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE 1 -- Economy Baseline Prices by Grade (Kixx/Kixx G1 medians, PKR/L)", N)
    header_row(ws, 5, ["Grade","Economy Baseline (PKR/L)","Listings","Notes"])
    r = 6
    eco_rows = {}
    for grade in ["0W-20","5W-20","5W-30","5W-40","10W-40","15W-40","20W-50"]:
        eco_brands = SHELL_TIER_BRANDS["economy"]
        m = (df["Grade"] == grade) & df["Brand"].isin(eco_brands)
        vals = df.loc[m, "Price_L"].dropna()
        baseline = round(float(vals.median()), 0) if not vals.empty else mkt_stats(df, grade)[0]
        cnt = len(vals)
        eco_rows[grade] = (r, 2)
        data_cell(ws, r, 1, grade, align="left")
        data_cell(ws, r, 2, baseline, fmt="pkr", bold=True)
        data_cell(ws, r, 3, cnt)
        data_cell(ws, r, 4, "Kixx median" if not vals.empty else "Market min (economy proxy)", align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "PARAMETER TABLE 2 -- Spec Uplift by Tier", N)
    r += 1
    header_row(ws, r, ["Tier","API/ACEA Spec","Uplift %","Rationale"])
    r += 1
    spec_desc = {
        "super_premium": ("API SP / ACEA A5", 0.28,
                          "Fully synthetic Group III -- OEM fuel economy approvals command 28% premium"),
        "premium":       ("API SN+ / ACEA A3", 0.18,
                          "Semi-synthetic Group II/III blend -- common fleet spec approval, 18% above economy"),
        "mainstream":    ("API SN / ACEA A3", 0.08,
                          "Mineral / part-synthetic Group I/II -- basic OEM approval, 8% above economy"),
        "economy":       ("API SN / SM", 0.00, "Mineral Group I -- baseline, no spec uplift"),
    }
    spec_rows = {}
    for tier, (spec, uplift, rat) in spec_desc.items():
        spec_rows[tier] = (r, 3)
        data_cell(ws, r, 1, TIER_LABEL[tier], align="left", bold=True)
        data_cell(ws, r, 2, spec, align="left")
        data_cell(ws, r, 3, uplift, fmt="pct")
        data_cell(ws, r, 4, rat, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "CALCULATION TABLE -- F5 Price per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Tier","Economy Baseline","Spec Uplift","F5 Price"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        tier = PSO_TIER.get(brand, "mainstream")
        m2 = (df["Grade"] == grade) & df["Brand"].isin(SHELL_TIER_BRANDS["economy"])
        vals = df.loc[m2, "Price_L"].dropna()
        eco_b = round(float(vals.median()), 0) if not vals.empty else mkt_stats(df, grade)[0]
        uplift = API_SPEC_PREMIUM.get(tier, 0)
        f5 = round(eco_b * (1 + uplift), 0) if eco_b else None
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, TIER_LABEL[tier], align="left")
        data_cell(ws, r, 5, eco_b, fmt="pkr")
        data_cell(ws, r, 6, uplift, fmt="pct")
        formula_cell(ws, r, 7, f"={cref(r,5)}*(1+{cref(r,6)})", fmt="pkr", bold=True) if eco_b else data_cell(ws, r, 7, "N/A")
        r += 1

    add_methodology(ws, r, N, **{k: METHODOLOGY["F5"][k] for k in ("how","rationale","implications")})
    set_widths(ws, [28, 9, 8, 15, 20, 12, 18])
    ws.freeze_panes = "A6"


def build_f6(wb, df, skus):
    ws = wb.create_sheet("F6 - Geographic")
    N = 11
    title_row(ws, 1, "F6 -- GEOGRAPHIC PRICE SEGMENTATION", N)
    desc_row(ws, 2,
             "Uses PSO FY25 volume data to identify regional growth signals. "
             "Growing regions support a small premium; declining regions need a discount. "
             "Applied as quarterly promotional pricing, not permanent list changes. "
             "FORMULA: Regional Price = F2_National x (1 + Regional_Signal%)", N, h=45)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE -- Regional Growth Signals (source: PSO Lubes Data Final.xlsx FY25 vs FY24)", N)
    header_row(ws, 5, ["Oil Category","Region","Volume Signal","Price Adj %","Price Directive"])
    r = 6
    for oil_type, regions in REGIONAL_SIGNAL.items():
        for region, adj in regions.items():
            vol_str = {
                ("hdeo","Central"): "+18.6% YoY DEO Central volume",
                ("mco","North"):    "-27.3% YoY MCO North volume",
                ("pcmo","North"):   "Soft YoY PCMO North",
            }.get((oil_type, region), "Regional baseline")
            directive = (f"Raise by {abs(adj*100):.0f}% vs national" if adj > 0
                         else f"Discount by {abs(adj*100):.0f}% vs national" if adj < 0
                         else "No adjustment (South = national baseline)")
            data_cell(ws, r, 1, oil_type.upper(), align="left")
            data_cell(ws, r, 2, region, align="left")
            data_cell(ws, r, 3, vol_str, align="left")
            data_cell(ws, r, 4, adj, fmt="pct1")
            data_cell(ws, r, 5, directive, align="left")
            r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "CALCULATION TABLE -- Regional Prices per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Pack (L)","F2 National (PKR/L)",
                        "South Adj","South Price","Central Adj","Central Price","North Adj","North Price"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        _, med, _, _, _ = mkt_stats(df, grade, pack_l)
        f2 = round(med * 0.97, 0) if med else None
        sig = REGIONAL_SIGNAL.get(oil_type.lower(), {})
        s_adj = sig.get("South", 0)
        c_adj = sig.get("Central", 0)
        n_adj = sig.get("North", 0)
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, pack_l, fmt="dec")
        data_cell(ws, r, 5, f2, fmt="pkr")
        data_cell(ws, r, 6, s_adj, fmt="pct1")
        formula_cell(ws, r, 7,  f"={cref(r,5)}*(1+{cref(r,6)})", fmt="pkr")
        data_cell(ws, r, 8, c_adj, fmt="pct1")
        formula_cell(ws, r, 9,  f"={cref(r,5)}*(1+{cref(r,8)})", fmt="pkr")
        data_cell(ws, r, 10, n_adj, fmt="pct1")
        formula_cell(ws, r, 11, f"={cref(r,5)}*(1+{cref(r,10)})", fmt="pkr")
        r += 1

    add_methodology(ws, r, N, **{k: METHODOLOGY["F6"][k] for k in ("how","rationale","implications")})
    set_widths(ws, [26, 9, 8, 8, 18, 10, 16, 12, 16, 10, 16])
    ws.freeze_panes = "A6"


def build_f7(wb, df, skus):
    ws = wb.create_sheet("F7 - Base Oil Floor")
    N = 8
    title_row(ws, 1, "F7 -- BASE OIL INDEX PRICE FLOOR  (Sanity Check)", N)
    desc_row(ws, 2,
             "Minimum viable price derived purely from base oil import costs. "
             "If PSO's price falls below F7, it implies subsidised base oil or accounting error. "
             "Markup 2.8x is the industry standard retail lubricants multiple (~64% gross margin). "
             "FORMULA: F7 Floor = (Base_Oil_Cost + Additive_Cost + Rs 50 overhead) x 2.8", N, h=50)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE -- Cost Inputs", N)
    header_row(ws, 5, ["Parameter","Value","Unit","Notes"])
    r = 6
    for param, val, unit, note in [
        ("Retail markup multiple", 2.8, "x", "Industry norm: 2.8x mfg cost = ~64% gross margin"),
        ("Blending/overhead add-on", 50, "PKR/L", "Blending, QC, filling, fixed overhead allocation"),
        ("Group I base oil cost", 280, "PKR/L", "Domestic SN150/BS150, Pakistan market 2025"),
        ("Group II base oil cost", 420, "PKR/L", "VHVI import -- Korean/Middle East, Pakistan 2025"),
        ("Group III base oil cost", 680, "PKR/L", "PAO/GTL import -- Europe/Korea, Pakistan 2025"),
    ]:
        data_cell(ws, r, 1, param, align="left", bold=True)
        data_cell(ws, r, 2, val, fmt="dec")
        data_cell(ws, r, 3, unit, align="left")
        data_cell(ws, r, 4, note, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "CALCULATION TABLE -- F7 Floor per Grade", N)
    r += 1
    header_row(ws, r, ["Grade","Base Oil Group","BO Cost","Additive","Overhead",
                        "Total Mfg Cost","x2.8 Markup","F7 Floor (PKR/L)"])
    r += 1
    for grade in sorted(set(g for _, g, _, _ in skus)):
        bog  = BASE_OIL_GROUP.get(grade, "II")
        bo_c = BASE_OIL_COST.get(bog, 420)
        add_c = ADDITIVE_COST.get(grade, 140)
        data_cell(ws, r, 1, grade, align="left", bold=True)
        data_cell(ws, r, 2, f"Group {bog}", align="left")
        data_cell(ws, r, 3, bo_c, fmt="pkr")
        data_cell(ws, r, 4, add_c, fmt="pkr")
        data_cell(ws, r, 5, 50, fmt="pkr")
        formula_cell(ws, r, 6, f"={cref(r,3)}+{cref(r,4)}+{cref(r,5)}", fmt="pkr")
        data_cell(ws, r, 7, 2.8, fmt="dec")
        formula_cell(ws, r, 8, f"={cref(r,6)}*{cref(r,7)}", fmt="pkr", bold=True)
        r += 1

    add_methodology(ws, r, N, **{k: METHODOLOGY["F7"][k] for k in ("how","rationale","implications")})
    set_widths(ws, [26, 14, 14, 14, 12, 16, 12, 18])
    ws.freeze_panes = "A6"


def build_f8(wb, df, skus, final_prices):
    ws = wb.create_sheet("F8 - Channel Pricing")
    N = 9
    title_row(ws, 1, "F8 -- CHANNEL PRICING MATRIX", N)
    desc_row(ws, 2,
             "Recommended list price cascades to four trade channels via structured discount waterfall. "
             "Channel prices must always stay above the F4 waterfall floor, even for fleet. "
             "FORMULA: Channel Price = List_Price x (1 - Channel_Discount%)", N, h=40)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE -- Channel Discount Structure", N)
    header_row(ws, 5, ["Channel","Discount Off List","Rationale","Typical Customer"])
    r = 6
    ch_info = [
        ("Retail (Pump)", 0.00,
         "Full list price -- pump is PSO's most valuable channel for margin",
         "Individual car owner at petrol station"),
        ("Workshop", 0.12,
         "12% trade allowance -- compensates stocking and application labour",
         "Auto workshop, lube shop, tyre shop"),
        ("Distributor", 0.15,
         "15% margin for regional distributors who take volume risk",
         "Regional wholesale distributor"),
        ("Fleet Contract", 0.20,
         "20% for committed annual volume -- volume + predictability trade-off",
         "Trucks, buses, ride-hailing fleets (>50 vehicles)"),
    ]
    for ch, disc, rat, cust in ch_info:
        data_cell(ws, r, 1, ch, align="left", bold=True)
        data_cell(ws, r, 2, disc, fmt="pct")
        data_cell(ws, r, 3, rat, align="left")
        data_cell(ws, r, 4, cust, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "CALCULATION TABLE -- Channel Prices per PSO SKU (based on Final Recommended List Price)", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Pack (L)","List Price (PKR/L)",
                        "Retail Pump","Workshop (-12%)","Distributor (-15%)","Fleet (-20%)"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        list_p = final_prices.get((brand, grade, pack_l))
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, pack_l, fmt="dec")
        data_cell(ws, r, 5, list_p, fmt="pkr", bold=True, bg=C_LIGHT)
        formula_cell(ws, r, 6, f"={cref(r,5)}*1.00", fmt="pkr")
        formula_cell(ws, r, 7, f"={cref(r,5)}*0.88", fmt="pkr")
        formula_cell(ws, r, 8, f"={cref(r,5)}*0.85", fmt="pkr")
        formula_cell(ws, r, 9, f"={cref(r,5)}*0.80", fmt="pkr")
        r += 1

    add_methodology(ws, r, N, **{k: METHODOLOGY["F8"][k] for k in ("how","rationale","implications")})
    set_widths(ws, [26, 9, 8, 8, 18, 16, 18, 18, 16])
    ws.freeze_panes = "A7"


def build_summary(wb, df, skus):
    ws = wb.create_sheet("Summary")
    N = 20
    title_row(ws, 1, "PSO LUBRICANTS -- PRICING STRATEGY SUMMARY  (Frameworks F1-F8, July 2026)", N)
    desc_row(ws, 2,
             "Final recommended price per litre for all PSO SKUs. "
             "Synthesis = weighted average of F1 (30%) + F2 (30%) + F3 (20%) + F5 (10%) + F4 (10%), "
             "with F4 waterfall as hard floor. "
             "Confidence: HIGH if >=4 frameworks have data, MEDIUM if >=2, LOW otherwise.", N, h=40)
    blank_row(ws, 3)
    header_row(ws, 4, [
        "Brand","Oil Type","Grade","Pack (L)","Tier",
        "Mkt Min","Mkt Med","Mkt Max",
        "F1 VBT\n(30%)","F2 CRP\n(30%)","F3 PPA\n(20%)","F4 Floor\n(hard)","F5 Spec\n(10%)",
        "F7 BOI\nFloor","Synthesis\n(wtd avg)","F4 Override?",
        "Final Rec\nPKR/L","Final Rec\nPkg PKR","Confidence","vs Market"
    ])
    final_prices = {}
    r = 5
    for brand, grade, pack_l, oil_type in skus:
        tier = PSO_TIER.get(brand, "mainstream")
        mn, med, mx, cnt, _ = mkt_stats(df, grade, pack_l)

        shell_p = tier_med(df, tier, grade, pack_l)
        disc    = PSO_BRAND_DISCOUNT.get(tier, 0.04)
        f1      = round(shell_p * (1 - disc), 0) if shell_p else None

        f2 = round(med * 0.97, 0) if med else None

        _, med4, _, _, _ = mkt_stats(df, grade, 4.0)
        mult = PPA_MULTIPLIER.get(pack_l, 1.0)
        f3   = round(med4 * mult, 0) if med4 else None

        bog   = BASE_OIL_GROUP.get(grade, "II")
        bo_c  = BASE_OIL_COST.get(bog, 420)
        add_c = ADDITIVE_COST.get(grade, 140)
        pkg_c = 60 if pack_l <= 1 else 35 if pack_l <= 5 else 20
        marg  = TIER_MARGIN.get(tier, 0.30)
        f4    = round((bo_c + add_c + pkg_c) / (1 - marg) / 0.80, 0)

        eco_m  = (df["Grade"] == grade) & df["Brand"].isin(SHELL_TIER_BRANDS["economy"])
        eco_v  = df.loc[eco_m, "Price_L"].dropna()
        eco_b  = round(float(eco_v.median()), 0) if not eco_v.empty else mn
        uplift = API_SPEC_PREMIUM.get(tier, 0)
        f5     = round(eco_b * (1 + uplift), 0) if eco_b else None

        f7 = round((bo_c + add_c + 50) * MIN_MARKUP, 0)

        fw_vals = [(f1, 0.30), (f2, 0.30), (f3, 0.20), (f5, 0.10), (f4, 0.10)]
        avail   = [(v, w) for v, w in fw_vals if v]
        synth   = round(sum(v*w for v,w in avail) / sum(w for _,w in avail), 0) if avail else None
        overridden = synth is not None and synth < f4
        rec        = max(synth, f4) if synth else f4
        confidence = ("HIGH" if sum(1 for v,_ in fw_vals if v) >= 4
                      else "MEDIUM" if sum(1 for v,_ in fw_vals if v) >= 2 else "LOW")
        final_prices[(brand, grade, pack_l)] = rec
        vs_med = round((rec - med) / med * 100, 1) if rec and med else None
        signal = ("PREMIUM" if vs_med and vs_med > 10
                  else "AT MARKET" if vs_med and vs_med > -10
                  else "VALUE POSITION" if vs_med is not None else "INSUFF. DATA")
        sig_bg = {"PREMIUM": "FFF8E6", "AT MARKET": C_LIGHT,
                  "VALUE POSITION": "FFE6E6"}.get(signal, C_GRAY)

        data_cell(ws, r,  1, brand, align="left")
        data_cell(ws, r,  2, oil_type.upper(), align="left")
        data_cell(ws, r,  3, grade, align="left")
        data_cell(ws, r,  4, pack_l, fmt="dec")
        data_cell(ws, r,  5, TIER_LABEL[tier], align="left")
        data_cell(ws, r,  6, mn, fmt="pkr")
        data_cell(ws, r,  7, med, fmt="pkr")
        data_cell(ws, r,  8, mx, fmt="pkr")
        data_cell(ws, r,  9, f1, fmt="pkr")
        data_cell(ws, r, 10, f2, fmt="pkr")
        data_cell(ws, r, 11, f3, fmt="pkr")
        data_cell(ws, r, 12, f4, fmt="pkr")
        data_cell(ws, r, 13, f5, fmt="pkr")
        data_cell(ws, r, 14, f7, fmt="pkr")
        data_cell(ws, r, 15, synth, fmt="pkr")
        data_cell(ws, r, 16, "YES -- raised to F4" if overridden else "No",
                  align="left", bg="FFE6E6" if overridden else None)
        data_cell(ws, r, 17, rec, fmt="pkr", bold=True, bg=C_LIGHT)
        data_cell(ws, r, 18, round(rec * pack_l, 0) if rec else None, fmt="pkr", bold=True, bg=C_LIGHT)
        data_cell(ws, r, 19, confidence, align="left",
                  bg={"HIGH":"D5F5E3","MEDIUM":C_LIGHT,"LOW":"FFE6E6"}.get(confidence))
        data_cell(ws, r, 20,
                  f"{vs_med:+.1f}%" if vs_med is not None else "N/A",
                  align="left", bg=sig_bg)
        r += 1

    blank_row(ws, r)
    desc_row(ws, r+1,
             "SYNTHESIS: Weighted average where available. F1+F2+F3 = 80% of weight. "
             "F4 is both a 10%-weight input AND a hard floor -- if weighted avg < F4, rec is raised to F4. "
             "F6 (geographic) and F8 (channel) apply AFTER synthesis as overlays.", N, h=30)
    desc_row(ws, r+2,
             "MARKET SIGNAL: PREMIUM = PSO >10% above market median. "
             "AT MARKET = within +-10%. VALUE POSITION = PSO >10% below median (revenue upside).", N, bg=C_LIGHT, h=25)

    set_widths(ws, [26, 9, 8, 7, 14, 9, 9, 9, 10, 10, 10, 10, 10, 10, 10, 16, 12, 12, 10, 13])
    ws.freeze_panes = "A5"
    return final_prices


# ==================================================================
#  Main
# ==================================================================

def main():
    print("Loading data...")
    df   = load_comp_df()
    skus = load_pso_skus()
    print(f"  Competitor listings: {len(df)}")
    print(f"  PSO SKUs: {len(skus)}")

    wb = Workbook()
    wb.remove(wb.active)

    print("Building sheets...")
    build_readme(wb)
    print("  README")
    build_competitor_data(wb, df)
    print("  Competitor Data")
    final_prices = build_summary(wb, df, skus)
    print("  Summary")
    build_f1(wb, df, skus);          print("  F1")
    build_f2(wb, df, skus);          print("  F2")
    build_f3(wb, df, skus);          print("  F3")
    build_f4(wb, df, skus);          print("  F4")
    build_f5(wb, df, skus);          print("  F5")
    build_f6(wb, df, skus);          print("  F6")
    build_f7(wb, df, skus);          print("  F7")
    build_f8(wb, df, skus, final_prices); print("  F8")

    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
    print(f"SKUs: {len(skus)} | Competitor listings: {len(df)}")


if __name__ == "__main__":
    main()
