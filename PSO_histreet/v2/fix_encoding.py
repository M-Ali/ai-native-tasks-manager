"""
Fix mojibake in PSO_Pricing_Strategy_Workbook.xlsx.
UTF-8 bytes were stored as Latin-1 characters.
Repair: encode each string back to Latin-1 bytes, decode as UTF-8.
Safe: ASCII strings round-trip identically. Correctly-stored Unicode
(em dash etc.) fails Latin-1 encode and is left unchanged.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from pathlib import Path
from openpyxl import load_workbook

PATH = Path("output/reports/PSO_Pricing_Strategy_Workbook.xlsx")
OUT  = Path("output/reports/PSO_Pricing_Strategy_Workbook_fixed.xlsx")

def try_fix(s):
    if s.startswith("="):          # leave formulas alone
        return s
    try:
        return s.encode("cp1252").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return s                   # already correct or not fixable

wb = load_workbook(PATH)
fixed_count = 0

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (v and isinstance(v, str)):
                continue
            repaired = try_fix(v)
            if repaired != v:
                cell.value = repaired
                fixed_count += 1

wb.save(OUT)
print(f"Fixed {fixed_count} cells.")
print(f"Saved: {OUT}")
