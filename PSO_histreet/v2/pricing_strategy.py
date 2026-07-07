"""
PSO Lubricants — SKU-Level Pricing Strategy Engine
Applies 8 international pricing frameworks to every PSO Carient SKU.
Output: pricing_strategy.csv  (one row per Brand × Grade × Pack)

Frameworks applied:
  F1  Value-Based Tiering         (Shell/Castrol tier equivalence)
  F2  Competitive Reference Price (market median ± band)
  F3  Price-Pack Architecture     (per-litre premium ladder by pack size)
  F4  McKinsey Price Waterfall    (pocket price health check)
  F5  OEM / API Spec Premium      (spec uplift over base grade)
  F6  Geographic Segmentation     (regional growth signals from v1 data)
  F7  Base Oil Index Floor        (cost-floor by base oil group)
  F8  Channel Pricing             (retail / workshop / fleet / distributor)

Final recommendation = weighted synthesis of F1–F7 with F8 channel overlay.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import sqlite3
import warnings
import numpy as np
import pandas as pd
from pathlib import Path

warnings.filterwarnings("ignore")

DB_PATH   = Path("db/prices.db")
OUT_CSV   = Path("output/reports/PSO_Pricing_Strategy.csv")
PORTFOLIO = Path("../PSO Portfolio.xlsx")
LUBES     = Path("../Lubes Data Final.xlsx")
OUT_CSV.parent.mkdir(parents=True, exist_ok=True)

# ── Constants ─────────────────────────────────────────────────────

# Tier hierarchy (lower index = more premium)
TIER_RANK = {"super_premium": 0, "premium": 1, "mainstream": 2, "economy": 3}

# Shell tier equivalents for PSO brands
PSO_TIER = {
    # PCMO
    "PSO Carient Ultra": "super_premium",
    "PSO Carient FS":    "premium",
    "PSO Carient Plus":  "mainstream",
    "PSO Carient SPRO":  "economy",
    # HDEO
    "PSO DEO Max":       "super_premium",
    "PSO DEO 8000":      "super_premium",
    "PSO DEO 6000":      "premium",
    "PSO DEO 5000":      "premium",
    "PSO DEO 3000":      "mainstream",
    "PSO Dieselube":     "economy",
    # MCO
    "PSO Blaze Xtreme":  "premium",
    "PSO Blaze 4T":      "mainstream",
}

SHELL_TIER = {
    "super_premium": ["Shell Helix Ultra"],
    "premium":       ["Shell Helix HX8", "Shell Helix HX7"],
    "mainstream":    ["Shell Helix HX5"],
    "economy":       ["Shell Helix HX3"],
}

# PPA anchor pack = 4L; multipliers relative to 4L price/L
PPA_MULTIPLIER = {
    0.8:  1.55,   # 800ml sachets — high per-unit premium
    1.0:  1.40,   # 1L — retail premium
    1.5:  1.30,
    2.0:  1.20,
    3.0:  1.12,
    3.5:  1.10,
    4.0:  1.00,   # ANCHOR
    5.0:  0.97,
    7.0:  0.92,
    10.0: 0.88,
    20.0: 0.78,
}

# Base oil groups by grade
BASE_OIL_GROUP = {
    "0W-20":  "III",   "5W-20":  "III",   "5W-30":  "III",
    "5W-40":  "III",   "10W-30": "II",    "10W-40": "II",
    "15W-40": "II",    "20W-50": "I",
}

# Base oil import cost floor (PKR/L) — approximate Pakistan market
BASE_OIL_COST = {"I": 280, "II": 420, "III": 680}

# Typical additive cost (PKR/L) by grade quality
ADDITIVE_COST = {
    "0W-20": 320, "5W-20": 300, "5W-30": 280, "5W-40": 260,
    "10W-30": 160, "10W-40": 140, "15W-40": 120, "20W-50": 90,
}

# Minimum viable markup over total cost (blending + additive + packaging + margin)
MIN_MARKUP = 2.8   # 2.8x cost = ~64% gross margin (industry standard for retail lubes)

# OEM/API spec premium over equivalent conventional grade
API_SPEC_PREMIUM = {
    "super_premium": 0.28,   # +28% for API SP / ACEA A3 fully synthetic
    "premium":       0.18,   # +18% for API SN+ / semi-synthetic
    "mainstream":    0.08,   # +8%  for API SN mineral
    "economy":       0.00,
}

# Channel discounts off list price
CHANNEL_DISCOUNT = {
    "retail_pump":    0.00,
    "workshop":       0.12,
    "distributor":    0.15,
    "fleet_contract": 0.20,
}

# Regional growth signals from v1 analysis (PSO Lubes Data Final)
REGIONAL_SIGNAL = {
    "DEO":  {"Central": +0.05,  "North": -0.02, "South": +0.00},
    "PCMO": {"Central": +0.02,  "North": -0.03, "South": +0.00},
    "MCO":  {"Central": +0.00,  "North": -0.08, "South": +0.00},
    "HDEO": {"Central": +0.05,  "North": -0.02, "South": +0.00},
}

# PSO brand discount vs Shell (brand perception gap, Pakistan market)
PSO_BRAND_DISCOUNT = {
    "super_premium": 0.08,
    "premium":       0.06,
    "mainstream":    0.04,
    "economy":       0.02,
}

# Grades per PSO brand — sourced from product knowledge
# SKU Wise sheet only carries pack sizes; grades come from product specs
BRAND_GRADES = {
    "PSO Carient Ultra": ["0W-20", "5W-20", "5W-30", "5W-40"],
    "PSO Carient FS":    ["5W-30", "5W-40", "10W-40"],
    "PSO Carient Plus":  ["10W-40", "15W-40", "20W-50"],
    "PSO Carient SPRO":  ["15W-40", "20W-50"],
    "PSO DEO Max":       ["10W-40", "15W-40"],
    "PSO DEO 8000":      ["10W-40", "15W-40"],
    "PSO DEO 6000":      ["15W-40"],
    "PSO DEO 5000":      ["10W-40", "15W-40"],
    "PSO DEO 3000":      ["15W-40", "20W-50"],
    "PSO Dieselube":     ["15W-40", "20W-50"],
    "PSO Blaze 4T":      ["10W-40", "20W-50"],
    "PSO Blaze Xtreme":  ["10W-40"],
}

# Maps raw brand names in SKU Wise sheet → canonical PSO brand name + oil type
BRAND_MAP = {
    "Carient SPRO":  ("PSO Carient SPRO",  "pcmo"),
    "Carient FS":    ("PSO Carient FS",    "pcmo"),
    "Carient Ultra": ("PSO Carient Ultra", "pcmo"),
    "Carient Plus":  ("PSO Carient Plus",  "pcmo"),
    "Blaze 4T":      ("PSO Blaze 4T",      "mco"),
    "Blaze Xtreme":  ("PSO Blaze Xtreme",  "mco"),
    "DEO 3000":      ("PSO DEO 3000",      "hdeo"),
    "DEO 5000":      ("PSO DEO 5000",      "hdeo"),
    "DEO 6000":      ("PSO DEO 6000",      "hdeo"),
    "DEO 8000":      ("PSO DEO 8000",      "hdeo"),
    "DEO Max":       ("PSO DEO Max",       "hdeo"),
    "Dieselube":     ("PSO Dieselube",     "hdeo"),
}

LUBES_FILE = Path("../Lubes Data Final.xlsx")
MAX_PACK_L = 4.0   # only retail packs ≤ 4L


def load_pso_skus_from_file() -> list[tuple]:
    """
    Reads SKU Wise sheet from Lubes Data Final.xlsx.
    Returns list of (brand, grade, pack_l, oil_type) for all SKUs
    with FY25 volume > 0 and pack_l <= MAX_PACK_L.
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(LUBES_FILE), data_only=True)
    ws = wb["SKU Wise"]
    rows = list(ws.iter_rows(values_only=True))

    skus = []
    current_brand_raw = None
    section_headers = {"Vol vs Val (PCMO)", "Vol vs Val (MCO)",
                       "Vol vs Val (DEO)", "Vol vs Val (Industrial)"}

    for row in rows:
        v0 = str(row[0]).strip() if row[0] else ""

        # Section / column headers — skip and (for section headers) reset brand
        if v0 in section_headers:
            current_brand_raw = None
            continue
        if v0 in ("Brand", "None"):
            continue

        # Update current brand when column A has a non-empty value
        if v0:
            current_brand_raw = v0

        # No known brand yet, or brand not in our map — skip data row
        if current_brand_raw not in BRAND_MAP:
            continue

        brand, oil_type = BRAND_MAP[current_brand_raw]

        # col1 = pack size (numeric litres), col3 = FY25 volume
        try:
            pack_raw = row[1]
            fy25_vol = row[3]
            if pack_raw is None:
                continue
            pack_l = float(pack_raw)
            fy25_vol = float(fy25_vol) if fy25_vol else 0.0
        except (TypeError, ValueError):
            continue

        # Only retail packs with actual FY25 sales
        if pack_l > MAX_PACK_L or fy25_vol <= 0:
            continue

        grades = BRAND_GRADES.get(brand, [])
        for grade in grades:
            skus.append((brand, grade, pack_l, oil_type))

    # Deduplicate while preserving order
    seen = set()
    unique = []
    for s in skus:
        if s not in seen:
            seen.add(s)
            unique.append(s)

    return unique


# ── Data loaders ──────────────────────────────────────────────────

def load_competitor_prices() -> pd.DataFrame:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    rows = con.execute("""
        SELECT brand_detected, grade_detected, pack_size_l, oil_type,
               price_per_litre, price
        FROM scraped_products
        WHERE price_per_litre IS NOT NULL
          AND grade_detected IS NOT NULL
          AND brand_detected IS NOT NULL
          AND DATE(scraped_at) = (SELECT DATE(MAX(scraped_at)) FROM scraped_products)
    """).fetchall()
    con.close()
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame([dict(r) for r in rows])
    df = df[df["price_per_litre"] > 100]   # strip zero/junk prices
    return df


def load_pso_portfolio_margins() -> dict:
    """
    Load PSO brand-level margin/L from PSO Portfolio.xlsx.
    Returns dict: {brand_key: margin_per_litre_fy25}
    """
    margins = {}
    try:
        wb = __import__("openpyxl").load_workbook(PORTFOLIO, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        current_seg = current_cat = current_brand = None
        TOP_SEGMENTS = {"RETAIL FUELS", "LUBRICANTS", "LPG", "AVIATION", "POWER"}
        SUB_CATS = {"PCMO", "DEO", "MCO", "INDUSTRIAL"}

        for row in rows:
            v0 = str(row[0]).strip().upper() if row[0] else ""
            v1 = str(row[1]).strip().upper() if row[1] else ""

            if v0 in TOP_SEGMENTS:
                current_seg = v0
                continue
            if v1 in SUB_CATS:
                current_cat = v1
                continue

            # Brand row: has a name in col1 and numeric data
            has_data = any(
                isinstance(row[c], (int, float)) and row[c] not in (None, 0)
                for c in range(2, 13)
            )
            if has_data and row[1]:
                current_brand = str(row[1]).strip()

            # Margin columns: col10=margin FY24, col11=margin FY25, col12=margin CY26
            if current_brand and has_data:
                margin_fy25 = row[11] if len(row) > 11 and isinstance(row[11], (int, float)) else None
                if margin_fy25:
                    key = f"{current_cat}|{current_brand}".upper()
                    margins[key] = margin_fy25

    except Exception as e:
        print(f"  [warn] Could not load Portfolio margins: {e}")
    return margins


# ── Competitor price helpers ───────────────────────────────────────

def get_comp_prices(comp_df: pd.DataFrame, grade: str, pack_l: float,
                    brands: list[str] = None) -> pd.Series:
    mask = (comp_df["grade_detected"] == grade)
    if pack_l:
        mask &= (comp_df["pack_size_l"] == pack_l)
    if brands:
        mask &= comp_df["brand_detected"].isin(brands)
    return comp_df.loc[mask, "price_per_litre"].dropna()


def get_shell_tier_price(comp_df: pd.DataFrame, tier: str, grade: str, pack_l: float) -> float | None:
    tier_brands = SHELL_TIER.get(tier, [])
    vals = get_comp_prices(comp_df, grade, pack_l, brands=tier_brands)
    if vals.empty:
        # Try without pack filter
        vals = get_comp_prices(comp_df, grade, None, brands=tier_brands)
    return float(vals.median()) if not vals.empty else None


def get_market_stats(comp_df: pd.DataFrame, grade: str, pack_l: float,
                     exclude_pso: bool = True) -> dict:
    mask = (comp_df["grade_detected"] == grade)
    if pack_l:
        mask &= (comp_df["pack_size_l"] == pack_l)
    if exclude_pso:
        mask &= ~comp_df["brand_detected"].str.contains("PSO|Carient", na=False)
    vals = comp_df.loc[mask, "price_per_litre"].dropna()
    if vals.empty:
        # Widen to all packs for same grade
        mask2 = (comp_df["grade_detected"] == grade)
        if exclude_pso:
            mask2 &= ~comp_df["brand_detected"].str.contains("PSO|Carient", na=False)
        vals = comp_df.loc[mask2, "price_per_litre"].dropna()

    if vals.empty:
        return {"min": None, "med": None, "max": None, "count": 0, "widened": True}
    return {
        "min": round(float(vals.min()), 0),
        "med": round(float(vals.median()), 0),
        "max": round(float(vals.max()), 0),
        "count": len(vals),
        "widened": pack_l not in comp_df.loc[mask, "pack_size_l"].values if pack_l else False,
    }


# ── Framework functions ────────────────────────────────────────────

def f1_value_based_tiering(tier: str, grade: str, pack_l: float,
                            comp_df: pd.DataFrame) -> tuple[float | None, str]:
    """Price at Shell tier equivalent minus PSO brand discount."""
    shell_price = get_shell_tier_price(comp_df, tier, grade, pack_l)
    if shell_price is None:
        return None, f"No Shell {tier} data found for {grade} {pack_l}L on Daraz."

    discount = PSO_BRAND_DISCOUNT.get(tier, 0.05)
    rec = round(shell_price * (1 - discount), 0)
    rationale = (
        f"Shell {tier} equivalent median = Rs {shell_price:,.0f}/L. "
        f"PSO brand perception discount = {discount*100:.0f}% "
        f"(narrows as Carient brand equity grows). "
        f"Recommended: Rs {rec:,.0f}/L."
    )
    return rec, rationale


def f2_competitive_reference_price(grade: str, pack_l: float,
                                    comp_df: pd.DataFrame) -> tuple[float | None, str]:
    """Price within +/-10% of market median."""
    stats = get_market_stats(comp_df, grade, pack_l)
    if stats["med"] is None:
        return None, f"Insufficient market data for {grade} {pack_l}L."

    med = stats["med"]
    rec = round(med * 0.97, 0)   # 3% below median — competitive but not cheapest
    gap_note = "(widened to all packs)" if stats.get("widened") else f"{pack_l}L"
    rationale = (
        f"Market {gap_note}: min Rs {stats['min']:,.0f}/L, "
        f"median Rs {stats['med']:,.0f}/L, max Rs {stats['max']:,.0f}/L "
        f"({stats['count']} listings). "
        f"CRP target: 3% below median = Rs {rec:,.0f}/L. "
        f"Positions PSO as best-value-in-class, not cheapest."
    )
    return rec, rationale


def f3_price_pack_architecture(brand: str, grade: str, pack_l: float,
                                comp_df: pd.DataFrame) -> tuple[float | None, str]:
    """Derive price from 4L anchor × PPA multiplier."""
    # Get 4L anchor from competitor market for this grade
    anchor_stats = get_market_stats(comp_df, grade, 4.0)
    anchor_ppl = anchor_stats["med"]

    if anchor_ppl is None:
        # Fall back to any available pack
        anchor_stats = get_market_stats(comp_df, grade, None)
        anchor_ppl = anchor_stats["med"]

    if anchor_ppl is None:
        return None, f"No market anchor available for {grade}."

    multiplier = PPA_MULTIPLIER.get(pack_l, 1.0)
    rec = round(anchor_ppl * multiplier, 0)
    direction = "premium" if multiplier > 1 else "discount"
    rationale = (
        f"4L market anchor = Rs {anchor_ppl:,.0f}/L. "
        f"PPA multiplier for {pack_l}L pack = {multiplier:.2f}x "
        f"({abs(1-multiplier)*100:.0f}% {direction} vs 4L). "
        f"Recommended: Rs {rec:,.0f}/L. "
        f"Ensures per-litre price ladder is legible to buyers."
    )
    return rec, rationale


def f4_price_waterfall(brand: str, grade: str, pack_l: float,
                        tier: str, margins: dict) -> tuple[float | None, str]:
    """
    McKinsey waterfall: derive minimum viable list price from
    pocket price floor (cost + margin).
    """
    bog = BASE_OIL_GROUP.get(grade, "II")
    cost_bo = BASE_OIL_COST.get(bog, 400)
    cost_add = ADDITIVE_COST.get(grade, 150)
    cost_pkg = 60 if pack_l <= 1 else 35 if pack_l <= 5 else 20
    total_cost = cost_bo + cost_add + cost_pkg

    # Target pocket margin (after all channel discounts)
    target_margin_pct = {"super_premium": 0.45, "premium": 0.38,
                         "mainstream": 0.30, "economy": 0.22}.get(tier, 0.30)

    pocket_price_floor = round(total_cost / (1 - target_margin_pct), 0)

    # List price must cover the deepest channel discount (fleet = -20%)
    max_discount = max(CHANNEL_DISCOUNT.values())
    list_price_floor = round(pocket_price_floor / (1 - max_discount), 0)

    rationale = (
        f"Cost build-up: base oil (Group {bog}) Rs {cost_bo}/L + "
        f"additive Rs {cost_add}/L + packaging Rs {cost_pkg}/L = Rs {total_cost}/L. "
        f"Target pocket margin ({tier}): {target_margin_pct*100:.0f}%. "
        f"Pocket price floor: Rs {pocket_price_floor:,.0f}/L. "
        f"After fleet discount (-{max_discount*100:.0f}%), "
        f"minimum list price: Rs {list_price_floor:,.0f}/L. "
        f"Any price below this destroys margin at the fleet channel."
    )
    return list_price_floor, rationale


def f5_spec_premium(brand: str, grade: str, tier: str,
                    comp_df: pd.DataFrame) -> tuple[float | None, str]:
    """
    Price premium for API/ACEA specification vs economy baseline.
    Anchors to economy-tier (SPRO) market price × spec uplift.
    """
    # Economy baseline for same grade
    economy_brands = SHELL_TIER.get("economy", [])
    base_vals = get_comp_prices(comp_df, grade, None, brands=economy_brands)
    if base_vals.empty:
        # Use market-wide min as proxy for economy
        stats = get_market_stats(comp_df, grade, None)
        base_price = stats["min"]
    else:
        base_price = float(base_vals.median())

    if base_price is None:
        return None, f"No economy-tier baseline for {grade}."

    uplift = API_SPEC_PREMIUM.get(tier, 0.0)
    rec = round(base_price * (1 + uplift), 0)

    spec_labels = {
        "super_premium": "API SP / ACEA A5 (fully synthetic, Group III+)",
        "premium":       "API SN+ / ACEA A3 (semi-synthetic, Group II/III blend)",
        "mainstream":    "API SN / ACEA A3 (mineral / part-synthetic, Group I/II)",
        "economy":       "API SN / SM (mineral, Group I)",
    }
    spec = spec_labels.get(tier, "")
    rationale = (
        f"Economy baseline ({grade}) = Rs {base_price:,.0f}/L. "
        f"{brand} spec: {spec}. "
        f"International OEM approval premium for this spec: +{uplift*100:.0f}%. "
        f"Spec-justified price: Rs {rec:,.0f}/L. "
        f"Premium only realised if spec is communicated at point of sale."
    )
    return rec, rationale


def f6_geographic_segmentation(brand: str, oil_type: str,
                                f2_price: float | None) -> dict:
    """
    Returns region-specific price adjustments based on v1 growth data.
    """
    cat_key = oil_type.upper()
    if cat_key == "PCMO":
        cat_key = "PCMO"
    elif cat_key in ("HDEO",):
        cat_key = "DEO"

    signals = REGIONAL_SIGNAL.get(cat_key, {})
    base = f2_price or 0
    result = {}
    rationale_parts = []

    for region, adj in signals.items():
        price = round(base * (1 + adj), 0) if base else None
        result[region] = price
        direction = "premium" if adj > 0 else "discount"
        vol_pct = abs(adj * 100)
        # Map adjustment back to regional data source
        if cat_key == "DEO" and region == "Central":
            source = "DEO Central volume +18.6% YoY (PSO Lubes Data Final)"
        elif cat_key == "MCO" and region == "North":
            source = "MCO North volume -27.3% YoY (PSO Lubes Data Final)"
        elif cat_key == "PCMO" and region == "North":
            source = "PCMO North soft YoY growth"
        else:
            source = "regional baseline"
        if adj != 0:
            rationale_parts.append(
                f"{region}: {'+' if adj>0 else ''}{adj*100:.0f}% vs national "
                f"(source: {source})"
            )

    rationale = (
        f"National list price (F2 CRP) = Rs {base:,.0f}/L. "
        + " | ".join(rationale_parts) + ". "
        "Suggest applying as quarterly promotional price, not permanent list change."
    )
    return result, rationale


def f7_base_oil_index(grade: str, tier: str) -> tuple[float, str]:
    """
    Minimum price floor derived from base oil import costs.
    """
    bog = BASE_OIL_GROUP.get(grade, "II")
    cost_bo = BASE_OIL_COST.get(bog, 400)
    cost_add = ADDITIVE_COST.get(grade, 150)
    total_mfg = cost_bo + cost_add + 50   # +50 blending/overhead

    floor = round(total_mfg * MIN_MARKUP, 0)
    rationale = (
        f"Grade {grade} requires Group {bog} base oil. "
        f"Import cost floor: Rs {cost_bo}/L. "
        f"Additive package: Rs {cost_add}/L. "
        f"Total manufacturing cost: Rs {total_mfg}/L. "
        f"At {MIN_MARKUP}x markup (industry standard for retail lubes): "
        f"BOI floor = Rs {floor:,.0f}/L. "
        f"Pricing below this implies margin destruction or subsidised base oil."
    )
    return floor, rationale


def f8_channel_pricing(list_price: float | None) -> tuple[dict, str]:
    """
    Derive channel-specific prices from list price.
    """
    if not list_price:
        return {}, "No list price available."

    channels = {}
    for ch, disc in CHANNEL_DISCOUNT.items():
        channels[ch] = round(list_price * (1 - disc), 0)

    rationale = (
        f"List price Rs {list_price:,.0f}/L. "
        f"Workshop (-{CHANNEL_DISCOUNT['workshop']*100:.0f}%): "
        f"Rs {channels['workshop']:,.0f}/L. "
        f"Distributor (-{CHANNEL_DISCOUNT['distributor']*100:.0f}%): "
        f"Rs {channels['distributor']:,.0f}/L. "
        f"Fleet contract (-{CHANNEL_DISCOUNT['fleet_contract']*100:.0f}%): "
        f"Rs {channels['fleet_contract']:,.0f}/L. "
        "Fleet price should never cross the F4 waterfall floor."
    )
    return channels, rationale


# ── Synthesis ─────────────────────────────────────────────────────

FRAMEWORK_WEIGHTS = {
    "f1_vbt":  0.30,   # Tier anchoring — strongest signal
    "f2_crp":  0.30,   # Market median — strongest signal
    "f3_ppa":  0.20,   # Pack ladder — structural
    "f5_spec": 0.10,   # Spec premium — directional
    "f4_pw":   0.10,   # Cost floor — hard constraint (overrides if violated)
}

def synthesize(prices: dict, f4_floor: float) -> tuple[float, str]:
    """Weighted average of available framework prices, with F4 as hard floor."""
    vals, weights = [], []
    for key, w in FRAMEWORK_WEIGHTS.items():
        p = prices.get(key)
        if p and p > 0:
            vals.append(p)
            weights.append(w)

    if not vals:
        return None, "Insufficient data for synthesis."

    total_w = sum(weights)
    weighted = sum(v * w for v, w in zip(vals, weights)) / total_w
    rec = round(weighted, 0)

    # Hard floor: F4 price waterfall
    violated = rec < f4_floor
    if violated:
        rec = f4_floor

    inputs_str = " | ".join(
        f"{k.upper()}=Rs {prices[k]:,.0f}" for k in FRAMEWORK_WEIGHTS if prices.get(k)
    )
    rationale = (
        f"Weighted synthesis ({inputs_str}). "
        f"Weights: F1=30%, F2=30%, F3=20%, F5=10%, F4=10%. "
        f"Raw weighted = Rs {weighted:,.0f}/L. "
        + (f"Adjusted UP to F4 floor Rs {f4_floor:,.0f}/L (margin protection)." if violated
           else f"Above F4 floor Rs {f4_floor:,.0f}/L — margin is protected.")
    )
    confidence = "HIGH" if len(vals) >= 4 else "MEDIUM" if len(vals) >= 2 else "LOW"
    return rec, rationale, confidence


# ── Main builder ──────────────────────────────────────────────────

def build():
    print("Loading competitor price data...")
    comp_df = load_competitor_prices()
    if comp_df.empty:
        print("ERROR: No competitor data in DB. Run scrape first.")
        return

    print(f"  {len(comp_df):,} competitor listings loaded.")
    print("Loading PSO Portfolio margins...")
    margins = load_pso_portfolio_margins()
    print(f"  {len(margins)} margin entries loaded.")

    records = []
    pso_skus = load_pso_skus_from_file()
    print(f"\nProcessing {len(pso_skus)} PSO SKUs (from Lubes Data Final.xlsx) across 8 frameworks...\n")

    for brand, grade, pack_l, oil_type in pso_skus:
        tier = PSO_TIER.get(brand, "mainstream")
        print(f"  {brand} | {grade} | {pack_l}L")

        row = {
            "Brand":    brand,
            "Grade":    grade,
            "Pack_L":   pack_l,
            "Oil_Type": oil_type.upper(),
            "Tier":     tier.replace("_", " ").title(),
        }

        # Market context
        mkt = get_market_stats(comp_df, grade, pack_l)
        row["Mkt_Min_PKR_L"]   = mkt["min"]
        row["Mkt_Med_PKR_L"]   = mkt["med"]
        row["Mkt_Max_PKR_L"]   = mkt["max"]
        row["Mkt_Listing_Count"] = mkt["count"]

        # F1 Value-Based Tiering
        f1_p, f1_r = f1_value_based_tiering(tier, grade, pack_l, comp_df)
        row["F1_VBT_Price_L"]     = f1_p
        row["F1_VBT_Rationale"]   = f1_r

        # F2 Competitive Reference Price
        f2_p, f2_r = f2_competitive_reference_price(grade, pack_l, comp_df)
        row["F2_CRP_Price_L"]     = f2_p
        row["F2_CRP_Rationale"]   = f2_r

        # F3 Price-Pack Architecture
        f3_p, f3_r = f3_price_pack_architecture(brand, grade, pack_l, comp_df)
        row["F3_PPA_Price_L"]     = f3_p
        row["F3_PPA_Rationale"]   = f3_r

        # F4 Price Waterfall (floor)
        f4_p, f4_r = f4_price_waterfall(brand, grade, pack_l, tier, margins)
        row["F4_Waterfall_Floor_L"]  = f4_p
        row["F4_Waterfall_Rationale"] = f4_r

        # F5 OEM/Spec Premium
        f5_p, f5_r = f5_spec_premium(brand, grade, tier, comp_df)
        row["F5_Spec_Price_L"]    = f5_p
        row["F5_Spec_Rationale"]  = f5_r

        # F6 Geographic Segmentation
        geo, f6_r = f6_geographic_segmentation(brand, oil_type, f2_p)
        row["F6_South_Price_L"]   = geo.get("South")
        row["F6_Central_Price_L"] = geo.get("Central")
        row["F6_North_Price_L"]   = geo.get("North")
        row["F6_Geo_Rationale"]   = f6_r

        # F7 Base Oil Index Floor
        f7_p, f7_r = f7_base_oil_index(grade, tier)
        row["F7_BOI_Floor_L"]     = f7_p
        row["F7_BOI_Rationale"]   = f7_r

        # Synthesis
        framework_prices = {
            "f1_vbt": f1_p, "f2_crp": f2_p, "f3_ppa": f3_p,
            "f4_pw":  f4_p, "f5_spec": f5_p,
        }
        synth = synthesize(framework_prices, f4_p or 0)
        rec_ppl, synth_r, confidence = synth

        row["Recommended_Price_L"]   = rec_ppl
        row["Recommended_Price_PKR"] = round(rec_ppl * pack_l, 0) if rec_ppl else None
        row["Confidence"]            = confidence
        row["Synthesis_Rationale"]   = synth_r

        # F8 Channel Pricing (applied to final recommended price)
        channels, f8_r = f8_channel_pricing(rec_ppl)
        row["F8_Retail_Pump_L"]      = channels.get("retail_pump")
        row["F8_Workshop_L"]         = channels.get("workshop")
        row["F8_Distributor_L"]      = channels.get("distributor")
        row["F8_Fleet_Contract_L"]   = channels.get("fleet_contract")
        row["F8_Channel_Rationale"]  = f8_r

        # vs market signal
        if rec_ppl and mkt["med"]:
            vs_med = (rec_ppl - mkt["med"]) / mkt["med"] * 100
            row["Rec_vs_Market_Med_Pct"] = round(vs_med, 1)
            row["Market_Signal"] = (
                "PREMIUM" if vs_med > 10 else
                "AT MARKET" if vs_med > -10 else
                "VALUE POSITION"
            )
        else:
            row["Rec_vs_Market_Med_Pct"] = None
            row["Market_Signal"] = "INSUFFICIENT DATA"

        records.append(row)

    df_out = pd.DataFrame(records)

    # Column order — readable in Excel
    col_order = [
        "Brand", "Grade", "Pack_L", "Oil_Type", "Tier",
        "Mkt_Min_PKR_L", "Mkt_Med_PKR_L", "Mkt_Max_PKR_L", "Mkt_Listing_Count",
        "F1_VBT_Price_L",       "F1_VBT_Rationale",
        "F2_CRP_Price_L",       "F2_CRP_Rationale",
        "F3_PPA_Price_L",       "F3_PPA_Rationale",
        "F4_Waterfall_Floor_L", "F4_Waterfall_Rationale",
        "F5_Spec_Price_L",      "F5_Spec_Rationale",
        "F6_South_Price_L", "F6_Central_Price_L", "F6_North_Price_L", "F6_Geo_Rationale",
        "F7_BOI_Floor_L",       "F7_BOI_Rationale",
        "Recommended_Price_L",  "Recommended_Price_PKR", "Confidence",
        "Synthesis_Rationale",
        "F8_Retail_Pump_L", "F8_Workshop_L", "F8_Distributor_L",
        "F8_Fleet_Contract_L",  "F8_Channel_Rationale",
        "Rec_vs_Market_Med_Pct", "Market_Signal",
    ]
    df_out = df_out[[c for c in col_order if c in df_out.columns]]

    df_out.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")   # utf-8-sig for Excel compat
    print(f"\nSaved: {OUT_CSV}")
    print(f"Rows: {len(df_out)} SKUs | Columns: {len(df_out.columns)}")

    # Console summary
    print("\n" + "="*90)
    print(f"{'Brand':<22} {'Grade':<8} {'Pack':>5}  {'Rec/L':>8}  {'Rec PKR':>9}  {'vs Mkt':>7}  Signal")
    print("="*90)
    for _, r in df_out.iterrows():
        rec = f"Rs {r['Recommended_Price_L']:,.0f}" if pd.notna(r.get("Recommended_Price_L")) else "  N/A   "
        rpkr = f"Rs {r['Recommended_Price_PKR']:,.0f}" if pd.notna(r.get("Recommended_Price_PKR")) else "    N/A  "
        vm = f"{r['Rec_vs_Market_Med_Pct']:+.1f}%" if pd.notna(r.get("Rec_vs_Market_Med_Pct")) else "    N/A"
        print(f"{str(r['Brand']):<22} {str(r['Grade']):<8} {str(r['Pack_L']):>5}L  {rec:>8}  {rpkr:>9}  {vm:>7}  {r.get('Market_Signal','')}")
    print("="*90)


if __name__ == "__main__":
    build()
