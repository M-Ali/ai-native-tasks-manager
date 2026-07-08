"""
Fix PSO_Pricing_Strategy_Workbook.xlsx:
  Remove self-referencing cross-sheet prefixes from formulas.
  e.g. ='F2 – Market Ref Price'!$E$14  →  =$E$14
"""
import re
from pathlib import Path
from openpyxl import load_workbook

IN  = Path("output/reports/PSO_Pricing_Strategy_Workbook.xlsx")
OUT = Path("output/reports/PSO_Pricing_Strategy_Workbook.xlsx")

wb = load_workbook(IN)
fixed = 0

for ws in wb.worksheets:
    sheet_name = ws.title
    # Pattern: ='<this sheet name>'!  or ="<this sheet name>"!
    pat = re.compile(
        r"='" + re.escape(sheet_name) + r"'!",
        re.IGNORECASE
    )
    for row in ws.iter_rows():
        for cell in row:
            if not (cell.value and isinstance(cell.value, str) and cell.value.startswith("=")):
                continue
            new = pat.sub("=", cell.value)
            if new != cell.value:
                cell.value = new
                fixed += 1

print(f"Fixed {fixed} self-referencing formulas.")
wb.save(OUT)
print(f"Saved: {OUT}")
