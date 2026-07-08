import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

wb = load_workbook("output/reports/PSO_Pricing_Strategy_Workbook_fixed.xlsx")
issues = []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (v and isinstance(v, str)):
                continue
            # Any Latin-1/cp1252 characters that shouldn't appear in clean text
            for ch in v:
                cp = ord(ch)
                if 0x80 <= cp <= 0x9F:   # cp1252 control range - sign of mojibake
                    issues.append((ws.title, cell.coordinate, v[:60]))
                    break

print(f"Remaining issues: {len(issues)}")
for sheet, coord, text in issues[:10]:
    print(f"  [{sheet}] {coord}: {text}")

# Spot check a known cell
ws2 = wb["README"]
print()
print("README A1:", ws2["A1"].value)
ws3 = wb["F1 – Value Tiering"]
print("F1 title:", ws3["A1"].value)
