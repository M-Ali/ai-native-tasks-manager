# -*- coding: utf-8 -*-
"""
Appends three-part methodology blocks (How It Works / Rationale / Implications)
below the calculation tables on each F1-F8 sheet.
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

PATH = Path("output/reports/PSO_Pricing_Strategy_Workbook.xlsx")

C_DARK = "023520"
C_MED  = "00703C"
C_GOLD = "C8960C"

def _fill(h):
    return PatternFill("solid", fgColor=h)

def write_header(ws, row, text, n_cols, bg, text_color="FFFFFF"):
    ws.row_dimensions[row].height = 17
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=9, color=text_color, name="Calibri")
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def write_body(ws, row, bullets, n_cols, bg="FAFAFA"):
    text = "\n".join(f"  •  {b}" for b in bullets)
    lines = sum(max(1, len(b) // 110 + 1) for b in bullets) + 1
    ws.row_dimensions[row].height = max(22, lines * 13)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=9, color="222222", name="Calibri")
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    return row + 1

def add_methodology(ws, start_row, n_cols, how_it_works, rationale, implications):
    ws.row_dimensions[start_row].height = 10  # gap
    r = start_row + 1
    write_header(ws, r, "HOW IT WORKS", n_cols, C_DARK)
    r = write_body(ws, r + 1, how_it_works, n_cols, "F4FAF7")
    ws.row_dimensions[r].height = 5; r += 1

    write_header(ws, r, "RATIONALE  —  Why this framework is used", n_cols, C_MED)
    r = write_body(ws, r + 1, rationale, n_cols, "F0FAF5")
    ws.row_dimensions[r].height = 5; r += 1

    write_header(ws, r, "IMPLICATIONS  —  What this means for PSO pricing decisions", n_cols, C_GOLD, "111111")
    r = write_body(ws, r + 1, implications, n_cols, "FEF9ED")

# ---- Methodology text (keyed by sheet-name fragment, matched loosely) ----

SHEETS = {
    "F1": dict(
        n=8,
        how=[
            "Identifies the Shell tier equivalent for each PSO brand: Carient Ultra = Shell Helix Ultra (Super Premium), "
            "Carient Plus = ZIC X7/Total Quartz 9000 (Mainstream), Carient SPRO = Kixx G1 (Economy), etc.",
            "Pulls the Daraz.pk median price/L for that Shell benchmark brand x grade combination from the competitor dataset.",
            "Applies a PSO brand-perception discount: 8% for Super Premium, 6% Premium, 4% Mainstream, 2% Economy.",
            "Formula: F1 Price = Shell_Tier_Median x (1 - Brand_Discount%)",
        ],
        rationale=[
            "Shell is the cognitive price anchor in Pakistan's lubricants market. Consumers use Shell prices as a mental "
            "reference when evaluating any other brand, even if they do not realise it.",
            "Pricing relative to Shell gives PSO a clear, defensible price story at the pump: 'same quality tier, better value.' "
            "This is easier to communicate than abstract per-litre comparisons.",
            "The brand discount is NOT a permanent feature. It reflects today's gap in awareness and OEM approval depth. "
            "In markets where PSO has invested in workshops and A&P, the effective gap is already closer to 2-3%.",
            "Using the Shell MEDIAN (not Shell minimum) avoids being dragged down by short-term Shell promotional prices.",
        ],
        implications=[
            "As PSO brand equity grows through OEM approvals and ATL/BTL investment, the discount should narrow: "
            "Super Premium target is 3-4% below Shell by 2028, not 8%. Each 1% narrowing on Carient Ultra alone "
            "recovers approximately Rs 30-35/L in margin.",
            "If Shell raises prices due to base oil cost pressures, PSO has immediate cover to follow: "
            "the percentage gap is maintained, so the absolute price difference actually grows, reinforcing 'best value.'",
            "For Carient Ultra 0W-20 and 5W-30, the 8% discount may already be too deep given product quality. "
            "Consider narrowing to 5% in FY26 and redirecting recovered margin into workshop visibility programmes.",
            "Monitor Shell pricing quarterly via the Daraz.pk scrape. Any Shell price movement triggers an immediate F1 recalculation.",
        ]
    ),
    "F2": dict(
        n=7,
        how=[
            "Collects all non-PSO Daraz.pk listings for each grade x pack combination from the scraped dataset.",
            "Computes the market median (50th percentile price per litre): the price at which half the competitors "
            "charge more and half charge less.",
            "Positions PSO at exactly 3% below median: F2 = Market_Median x 0.97.",
            "When fewer than 3 exact-pack listings exist, scope widens to all packs for that grade "
            "(flagged as 'Grade-wide' in the Scope column).",
        ],
        rationale=[
            "The market median is the most transparent benchmark: directly observable by any buyer comparing options "
            "online or in a workshop. It cannot be disputed.",
            "3% below median (not 20% below) avoids the 'cheap brand' perception trap. FMCG research shows extreme "
            "discounting signals inferior quality more than value. PSO should be 'better value,' not 'cheap.'",
            "Being 3% below median means PSO wins the tie-break when a buyer compares two similar-spec products "
            "at the same quality level: same spec, lower price, wins the shelf.",
            "F2 carries the highest weight (30%) because it is the most current, market-facing signal and automatically "
            "absorbs competitor price movements without requiring model recalibration.",
        ],
        implications=[
            "Recalculate F2 quarterly using fresh Daraz scrape data. If competitor prices rise due to base oil pressures, "
            "F2 will show PSO can raise prices without losing competitive position.",
            "The 3% discount is a policy choice. If PSO builds stronger brand equity, reduce it to 1-2%, "
            "capturing significant margin without volume loss. A 1% change on a 45-SKU range is material.",
            "For 'Grade-wide' scope SKUs, the F2 result is less reliable. Prioritise getting pack-specific competitor "
            "data before finalising list prices for those SKUs.",
            "F2 sets the baseline for both F3 (pack architecture) and F6 (geographic). An error in F2 propagates "
            "downstream: data quality in the competitor scrape is the single most important input to this model.",
        ]
    ),
    "F3": dict(
        n=7,
        how=[
            "Designates the 4L pack as the anchor (multiplier = 1.00x). All other pack sizes carry a multiplier "
            "relative to the 4L anchor price per litre.",
            "The 4L market median (from competitor data for each grade) is the anchor price/L.",
            "Each pack size price = 4L anchor x PPA multiplier. Example: 1L pack = 4L anchor x 1.40 (40% per-litre premium).",
            "Multipliers are derived from international retail lubricants practice: smaller packs carry a "
            "convenience premium; larger packs carry a bulk discount.",
        ],
        rationale=[
            "Without a structured PPA, brands accidentally price 1L packs lower per litre than 3L packs "
            "(or vice versa), confusing price-aware buyers who calculate unit costs.",
            "The 1L premium (40% above 4L per litre) is justified by: single engine-change convenience, "
            "retail display slot cost, impulse-purchase pricing psychology, and high turnover at petrol station forecourts.",
            "A consistent price ladder prevents pack cannibalism: if 3L packs were priced at par with 4L per litre, "
            "buyers would always choose 3L, destroying 4L volumes and margin.",
            "PSO's range spans 1L to 4L across the same brand/grade. Buyers who visit multiple channels "
            "compare cross-pack, and inconsistency destroys credibility.",
        ],
        implications=[
            "The biggest revenue opportunity is the 1L PCMO pack at petrol stations, where buyers are least "
            "price-sensitive and most convenience-driven. The 40% per-litre premium is fully justifiable and should be enforced.",
            "Review current 1L and 3L price points against the PPA ladder. The next list price revision is the "
            "right moment to re-anchor the entire range. Do not introduce mid-cycle exceptions.",
            "Any new pack size launch (e.g., 800ml sachet) must have its PPA multiplier defined BEFORE launch. "
            "Without pre-setting the multiplier, the new pack will almost certainly be mispriced relative to the range.",
            "For MCO (motorcycle oil), the 1L pack dominates sales. The 1.40x multiplier should be enforced here "
            "since MCO buyers are highly brand-loyal once a workshop mechanic recommends a product.",
        ]
    ),
    "F4": dict(
        n=12,
        how=[
            "Builds the minimum viable price from cost structure upward, not downward from the market.",
            "Step 1 - Cost build-up: Base oil import parity (Group I Rs 280/L, II Rs 420/L, III Rs 680/L) "
            "+ Additive package (grade-specific) + Packaging (per-litre by pack size) = Total Manufacturing Cost.",
            "Step 2 - Pocket price floor: Total Cost / (1 - Target Margin%). Targets: 45% Super Premium, "
            "38% Premium, 30% Mainstream, 22% Economy. This is the minimum price that covers PSO's cost after all deductions.",
            "Step 3 - List price floor: Pocket Price Floor / (1 - 20% Fleet Discount). Fleet is the deepest channel, "
            "so the list price must cover this to maintain margin everywhere.",
            "F4 is a hard floor: if the weighted synthesis falls below F4, the recommendation is raised to F4.",
        ],
        rationale=[
            "The McKinsey Price Waterfall reveals how revenue 'leaks' from list price to the price that actually reaches "
            "PSO's P&L ('pocket price'). Without this, individually small concessions destroy profitability collectively.",
            "PSO's fleet channel is especially exposed: 20% off a list price set too low means PSO is subsidising "
            "fleet customers' engine oil from its own margin - an unsustainable practice.",
            "The target margins (45% Super Premium to 22% Economy) reflect international lubricant industry norms "
            "adjusted for Pakistan's distribution structure, dealer margins, and working capital costs.",
            "F4 also sets the minimum for fleet negotiation. Any fleet deal deeper than 20% off list "
            "requires explicit board approval because it breaches the pocket price floor.",
        ],
        implications=[
            "F4 fires most often for economy-tier HDEO (20W-50, 15W-40 mineral) where Korean/Chinese imports "
            "on Daraz drive market prices below PSO's cost floor. Do NOT match these prices.",
            "Instead of cutting price below F4: (a) exit commodity packs where margin is structurally impossible, "
            "(b) reposition to 'value premium' with better packaging and communication, or "
            "(c) explore private-label manufacturing for fleet-only packs at different cost structure.",
            "If F4 floor > F7 floor for any grade, the margin target is correctly set above the pure cost floor - good. "
            "If F4 floor < F7 floor, the margin assumption is too low and needs upward revision.",
            "Re-run F4 whenever base oil import prices shift more than 10% - approximately every 6 months "
            "in a stable environment, more frequently during crude oil price shocks.",
        ]
    ),
    "F5": dict(
        n=7,
        how=[
            "Anchors to the economy-tier (Kixx/Kixx G1) Daraz.pk median for each grade as the 'floor' baseline - "
            "what the same viscosity grade costs with the minimum viable API specification.",
            "Applies an API/ACEA specification uplift based on PSO's tier: +28% for API SP fully synthetic, "
            "+18% for API SN+ semi-synthetic, +8% for API SN mineral, 0% for economy.",
            "Formula: F5 Price = Economy_Baseline x (1 + Spec_Uplift%)",
            "Uplift percentages are derived from international market research on OEM approval premiums in comparable "
            "emerging markets (India, Vietnam, Indonesia).",
        ],
        rationale=[
            "Engine oil pricing must reflect specification level, not just viscosity grade. A 5W-40 API SP (fully synthetic) "
            "has fundamentally different chemistry, manufacturing cost, and engine protection vs a 5W-40 API SN mineral.",
            "In Pakistan, pump attendants and workshop mechanics rarely explain spec differences. "
            "F5 sets the price SIGNAL that communicates quality: consumers learn to associate higher price with higher spec, "
            "even without reading the API label.",
            "F5 is weighted at only 10% because spec literacy in Pakistan is still developing. "
            "As car parc modernises (more Euro-4/5 engines, more turbo/hybrid vehicles requiring SP/SN+), "
            "the spec premium will become more important and the weight should increase.",
            "F5 also prevents underpricing of premium products: Carient Ultra 5W-40 (API SP) should never "
            "approach Kixx G1 5W-40 (API SN) pricing, regardless of competitive pressure.",
        ],
        implications=[
            "Investing in OEM approvals (VW 504.00, BMW Longlife-04, Toyota WS) directly increases the "
            "justifiable spec premium from 28% toward 35-40% for those specific SKUs. "
            "A single major OEM approval generates years of pricing power.",
            "For economy HDEO (DEO 3000, Dieselube), the F5 uplift is 0% - these compete purely on price. "
            "Any price above the F5 baseline in the economy tier must be supported by brand or distribution advantages, not spec.",
            "Monitor the economy baseline (Kixx pricing) closely. If Korean brands discount aggressively, "
            "the F5 baseline drops and PSO's spec premium erodes unless parallel brand investment is maintained.",
            "Consider displaying API specification prominently on packaging and POS materials at petrol stations. "
            "Consumers who understand spec buy on spec, not just on price - this shifts F5 weight from 10% toward 20%.",
        ]
    ),
    "F6": dict(
        n=11,
        how=[
            "Uses PSO's own FY25 lubricants volume data (Lubes Data Final.xlsx) to compute regional YoY growth signals "
            "- the percentage difference in volume growth between each region and the national average.",
            "Regions growing faster than the national average can support a price premium (e.g., Central DEO: +5%). "
            "Declining regions require a discount to defend volume (North MCO: -8%).",
            "Formula: Regional Price = F2_National_Price x (1 + Regional_Signal%)",
            "South (Karachi/Hyderabad) is the national baseline (0% adjustment). "
            "Adjustments are implemented as quarterly promotional price variations, NOT permanent list price changes.",
        ],
        rationale=[
            "Pakistan's lubricants market is deeply regional: Central Punjab drives HDEO (agricultural/transport), "
            "Karachi dominates PCMO, KPK/Northern corridor has distinct MCO patterns tied to two-wheeler density.",
            "A flat national price leaves revenue on the table in strong-growth regions (where demand is relatively inelastic) "
            "and loses volume in weak regions (where buyers switch to cheaper local alternatives).",
            "Geographic pricing is standard practice for CPG and lubricants companies in South Asia. "
            "Castrol, Shell, and Total all maintain regional trade price lists in Pakistan.",
            "Using PSO's own volume data (not external estimates) makes the signal accurate and tied to "
            "PSO's distribution reality - it reflects where PSO is genuinely growing or losing ground.",
        ],
        implications=[
            "For HDEO in Central Punjab (+5%): this translates to Rs 75-90/L higher price on a 15W-40 4L pack. "
            "Recoverable through regional promotional pricing without disturbing the national list.",
            "For MCO in North (-8%): this is a defensive measure. North MCO decline suggests PSO is losing to "
            "local blenders or two-wheeler brand partnerships. Pricing alone will not recover volume - conduct a distribution audit.",
            "Grey market risk: if the South-to-North price gap on the same SKU exceeds Rs 50-70/L, "
            "distributors will buy cheap in South and sell in North. Cap regional variation at 5% on any high-volume SKU.",
            "Update regional signals every 6 months using the latest PSO sales data. "
            "The FY25 data in this workbook will be stale by Q3 FY26 and must be refreshed.",
        ]
    ),
    "F7": dict(
        n=8,
        how=[
            "Takes the base oil import parity cost for the grade's required group "
            "(Group I: Rs 280/L, Group II: Rs 420/L, Group III: Rs 680/L).",
            "Adds the grade-specific additive package cost and a fixed blending/overhead allocation of Rs 50/L.",
            "Multiplies total manufacturing cost by 2.8x - the standard retail lubricants markup in Pakistan "
            "(equivalent to approximately 64% gross margin).",
            "Formula: F7 Floor = (Base_Oil_Cost + Additive_Cost + Rs 50) x 2.8",
            "F7 is a standalone sanity check and audit tool, not a synthesis input.",
        ],
        rationale=[
            "Base oil is the primary raw material in any lubricant (55-75% of total product cost). "
            "Its price tracks crude oil with a 6-8 week lag, creating predictable cost movements.",
            "The 2.8x markup is the industry standard for retail lubricants in Pakistan: derived from distributor margin (15%), "
            "dealer/pump margin (8%), PSO overhead (12%), and net profit target (10%).",
            "F7 provides an independent floor that does not depend on competitor data (like F2) or margin targets (like F4). "
            "Even if competitive data is stale or targets are mis-set, F7 catches structural underpricing.",
            "F7 is especially valuable during crude oil price shocks: when base oil import costs spike, "
            "F7 immediately shows which PSO products are now priced below manufacturing cost.",
        ],
        implications=[
            "Any PSO SKU priced below its F7 floor is being sold at a structural loss. "
            "This should trigger an immediate price revision or product exit decision.",
            "If F7 > F4 for any SKU, the F4 margin target is set too low - revise the margin assumption upward "
            "for that tier before the next pricing cycle.",
            "Monitor base oil import prices monthly (PARCO, NRL, and Korean import indices). "
            "A 10% crude oil increase typically translates to 8-12% increase in F7 floor values within 6-8 weeks.",
            "Group III (5W-40, 5W-30, 0W-20) is the most volatile and entirely imported. "
            "Consider hedging 30-60 days of base oil inventory to protect Super Premium pricing from cost shocks.",
        ]
    ),
    "F8": dict(
        n=9,
        how=[
            "Takes the final recommended list price from the weighted synthesis (Summary sheet) as the starting point.",
            "Applies a structured discount cascade for each trade channel: Retail Pump = 0%, "
            "Workshop = -12%, Distributor = -15%, Fleet Contract = -20%.",
            "Formula: Channel Price = List Price x (1 - Discount%)",
            "The fleet contract price is the critical floor: it must never fall below the F4 cost waterfall floor.",
        ],
        rationale=[
            "Different trade channels have fundamentally different cost structures, volume commitments, and credit terms. "
            "A single price for all channels either overcharges workshops (losing trade) or undercharges fleet (losing margin).",
            "Workshop -12%: they stock product, drive recommendations at service, but have no volume commitment. "
            "12% is the minimum trade allowance to secure shelf placement and mechanic recommendation.",
            "Distributor -15%: they take inventory risk, provide credit to dealers, manage last-mile logistics. "
            "15% covers their operating costs and provides a viable business model.",
            "Fleet -20%: they commit to annual volumes and provide predictable revenue, but negotiate aggressively "
            "and pay in 45-90 day cycles. 20% compensates for working capital cost and volume commitment.",
            "A structured discount matrix prevents channel conflict: each account type knows their price "
            "and cannot arbitrage because the discount is tied to account type, not transaction size.",
        ],
        implications=[
            "Critical constraint: Fleet contract price (list x 0.80) must always exceed the F4 pocket price floor. "
            "If not, PSO is subsidising fleet customers from its own margin - unsustainable.",
            "No channel should receive more than 20% off list without board-level approval. "
            "Any 'special deal' deeper than fleet discount erodes the entire price architecture.",
            "Workshops are PSO's most strategic brand advocacy channel. The 12% discount is a relationship investment. "
            "Consider tiering: 12% standard, 14% for top-volume workshops, 16% for PSO-branded workshop partners.",
            "Fleet contracts must be reviewed annually against the UPDATED synthesis price, not the prior year's list. "
            "If list prices rise 8% due to base oil inflation but fleet contracts are locked at prior year list, "
            "PSO absorbs the entire cost increase at its most margin-dilutive channel.",
        ]
    ),
}

# ---- Match sheets by F-number prefix and apply ----

wb = load_workbook(PATH)

for ws in wb.worksheets:
    name = ws.title
    matched = None
    for key in SHEETS:
        if name.startswith(key):
            matched = key
            break
    if not matched:
        continue
    data = SHEETS[matched]
    start = ws.max_row + 2
    add_methodology(ws, start, data["n"], data["how"], data["rationale"], data["implications"])
    print(f"  Added: {name} (row {start})")

wb.save(PATH)
print(f"\nSaved: {PATH}")
