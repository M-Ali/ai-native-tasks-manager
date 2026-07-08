import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

wb = load_workbook("output/reports/PSO_Pricing_Strategy_Workbook_v2.xlsx")
issues = []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (v and isinstance(v, str)): continue
            for ch in v:
                if 0x80 <= ord(ch) <= 0x9F:
                    issues.append((ws.title, cell.coordinate, v[:60]))
                    break

print(f"Encoding issues: {len(issues)}")
for sheet, coord, text in issues[:10]:
    print(f"  [{sheet}] {coord}: {text}")

formula_count = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1

print(f"Live formulas: {formula_count}")
print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
