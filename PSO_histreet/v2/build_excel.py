"""
PSO Lubricants â€” Pricing Strategy Excel Workbook Builder
Produces PSO_Pricing_Strategy_Workbook.xlsx with:
  â€¢ Competitor Data  â€” 365 scraped Daraz.pk listings
  â€¢ F1â€“F8 sheets     â€” parameters + live Excel formulas + methodology
  â€¢ Summary          â€” weighted synthesis, channel prices, market signals
"""
import sqlite3, math
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# â”€â”€ Paths â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
DB_PATH   = Path("db/prices.db")
OUT_PATH  = Path("output/reports/PSO_Pricing_Strategy_Workbook.xlsx")
LUBES_FILE = Path("../Lubes Data Final.xlsx")
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# â”€â”€ Palette â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
C_DARK  = "023520"
C_MED   = "00703C"
C_GOLD  = "C8960C"
C_LIGHT = "E8F5EE"
C_WHITE = "FFFFFF"
C_GRAY  = "F0F0F0"
C_RED   = "C0392B"
C_AMBER = "E67E22"
C_BLUE  = "2980B9"

# â”€â”€ Domain constants â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
PSO_TIER = {
    "PSO Carient Ultra": "super_premium",
    "PSO Carient FS":    "premium",
    "PSO Carient Plus":  "mainstream",
    "PSO Carient SPRO":  "economy",
    "PSO DEO Max":       "super_premium",
    "PSO DEO 8000":      "super_premium",
    "PSO DEO 6000":      "premium",
    "PSO DEO 5000":      "premium",
    "PSO DEO 3000":      "mainstream",
    "PSO Dieselube":     "economy",
    "PSO Blaze Xtreme":  "premium",
    "PSO Blaze 4T":      "mainstream",
}
TIER_LABEL = {"super_premium": "Super Premium", "premium": "Premium",
              "mainstream": "Mainstream", "economy": "Economy"}

SHELL_TIER_BRANDS = {
    "super_premium": ["Shell Helix Ultra", "Shell"],
    "premium":       ["Shell Helix HX7", "ZIC X9"],
    "mainstream":    ["ZIC X7", "Total Quartz 9000"],
    "economy":       ["Kixx", "Kixx G1"],
}

PSO_BRAND_DISCOUNT = {
    "super_premium": 0.08, "premium": 0.06, "mainstream": 0.04, "economy": 0.02,
}
API_SPEC_PREMIUM = {
    "super_premium": 0.28, "premium": 0.18, "mainstream": 0.08, "economy": 0.00,
}
PPA_MULTIPLIER = {
    0.8: 1.55, 1.0: 1.40, 1.5: 1.30, 2.0: 1.20, 3.0: 1.12,
    3.5: 1.10, 4.0: 1.00, 5.0: 0.97, 7.0: 0.92, 10.0: 0.88, 20.0: 0.78,
}
BASE_OIL_GROUP = {
    "0W-20": "III", "5W-20": "III", "5W-30": "III", "5W-40": "III",
    "10W-30": "II", "10W-40": "II", "15W-40": "II", "20W-50": "I",
}
BASE_OIL_COST  = {"I": 280, "II": 420, "III": 680}
ADDITIVE_COST  = {
    "0W-20": 320, "5W-20": 300, "5W-30": 280, "5W-40": 260,
    "10W-30": 160, "10W-40": 140, "15W-40": 120, "20W-50": 90,
}
MIN_MARKUP = 2.8
TIER_MARGIN = {"super_premium": 0.45, "premium": 0.38, "mainstream": 0.30, "economy": 0.22}
CHANNEL_DISCOUNT = {
    "Retail (Pump)": 0.00, "Workshop": 0.12, "Distributor": 0.15, "Fleet Contract": 0.20,
}
REGIONAL_SIGNAL = {
    "pcmo": {"Central":  0.02, "North": -0.03, "South": 0.00},
    "hdeo": {"Central":  0.05, "North": -0.02, "South": 0.00},
    "mco":  {"Central":  0.00, "North": -0.08, "South": 0.00},
}

BRAND_GRADES = {
    "PSO Carient Ultra": ["0W-20", "5W-20", "5W-30", "5W-40"],
    "PSO Carient FS":    ["5W-30", "5W-40", "10W-40"],
    "PSO Carient Plus":  ["10W-40", "15W-40", "20W-50"],
    "PSO Carient SPRO":  ["15W-40", "20W-50"],
    "PSO DEO Max":       ["10W-40", "15W-40"],
    "PSO DEO 8000":      ["10W-40", "15W-40"],
    "PSO DEO 6000":      ["15W-40"],
    "PSO DEO 5000":      ["10W-40", "15W-40"],
    "PSO DEO 3000":      ["15W-40", "20W-50"],
    "PSO Dieselube":     ["15W-40", "20W-50"],
    "PSO Blaze 4T":      ["10W-40", "20W-50"],
    "PSO Blaze Xtreme":  ["10W-40"],
}
BRAND_MAP = {
    "Carient SPRO":  ("PSO Carient SPRO",  "pcmo"),
    "Carient FS":    ("PSO Carient FS",    "pcmo"),
    "Carient Ultra": ("PSO Carient Ultra", "pcmo"),
    "Carient Plus":  ("PSO Carient Plus",  "pcmo"),
    "Blaze 4T":      ("PSO Blaze 4T",      "mco"),
    "Blaze Xtreme":  ("PSO Blaze Xtreme",  "mco"),
    "DEO 3000":      ("PSO DEO 3000",      "hdeo"),
    "DEO 5000":      ("PSO DEO 5000",      "hdeo"),
    "DEO 6000":      ("PSO DEO 6000",      "hdeo"),
    "DEO 8000":      ("PSO DEO 8000",      "hdeo"),
    "DEO Max":       ("PSO DEO Max",       "hdeo"),
    "Dieselube":     ("PSO Dieselube",     "hdeo"),
}

FRAMEWORK_WEIGHTS = [("F1_VBT", 0.30), ("F2_CRP", 0.30), ("F3_PPA", 0.20),
                     ("F5_Spec", 0.10), ("F4_WF", 0.10)]

# â”€â”€ Style helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def _font(bold=False, size=9, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _right():
    return Alignment(horizontal="right", vertical="center")

def cref(row, col_n):
    return f"${get_column_letter(col_n)}${row}"

# â”€â”€ Cell writers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def title_row(ws, row, text, n_cols, bg=C_DARK, size=13):
    ws.row_dimensions[row].height = 28
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=size, color=C_WHITE, name="Calibri")
    c.fill = _fill(bg)
    c.alignment = _left(wrap=False)

def section_row(ws, row, text, n_cols, bg=C_MED):
    ws.row_dimensions[row].height = 18
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=9, color=C_WHITE, name="Calibri")
    c.fill = _fill(bg)
    c.alignment = _left(wrap=False)

def desc_row(ws, row, text, n_cols, bg="FFFFFF", h=40):
    ws.row_dimensions[row].height = h
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=8.5, italic=True, color="555555", name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    if bg != "FFFFFF":
        c.fill = _fill(bg)

def header_row(ws, row, headers, start_col=1, bg=C_DARK):
    ws.row_dimensions[row].height = 17
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(bold=True, size=8.5, color=C_WHITE, name="Calibri")
        c.fill = _fill(bg)
        c.alignment = _center()
        c.border = _border()

def sub_header_row(ws, row, headers, start_col=1):
    ws.row_dimensions[row].height = 16
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(bold=True, size=8.5, color=C_DARK, name="Calibri")
        c.fill = _fill(C_LIGHT)
        c.alignment = _center()
        c.border = _border()

def data_cell(ws, row, col_n, value, fmt=None, bold=False, bg=None, align="right"):
    c = ws.cell(row=row, column=col_n, value=value)
    c.font = Font(bold=bold, size=8.5, name="Calibri")
    c.border = _border()
    c.alignment = _left() if align == "left" else _right()
    if fmt == "pkr":   c.number_format = '#,##0'
    elif fmt == "pct": c.number_format = '0%'
    elif fmt == "pct1":c.number_format = '0.0%'
    elif fmt == "dec": c.number_format = '0.00'
    elif fmt == "dec2":c.number_format = '0.00'
    if bg: c.fill = _fill(bg)
    return c

def formula_cell(ws, row, col_n, formula, fmt="pkr", bg=None, bold=False):
    c = ws.cell(row=row, column=col_n, value=formula)
    c.font = Font(bold=bold, size=8.5, color=C_BLUE if bg is None else "000000", name="Calibri")
    c.border = _border()
    c.alignment = _right()
    if fmt == "pkr":   c.number_format = '#,##0'
    elif fmt == "pct": c.number_format = '0%'
    elif fmt == "dec": c.number_format = '0.00'
    if bg: c.fill = _fill(bg)
    return c

def result_cell(ws, row, col_n, formula, bold=True):
    c = ws.cell(row=row, column=col_n, value=formula)
    c.font = Font(bold=bold, size=9, color=C_DARK, name="Calibri")
    c.fill = _fill(C_LIGHT)
    c.border = _border()
    c.number_format = '#,##0'
    c.alignment = _right()
    return c

def blank_row(ws, row, h=8):
    ws.row_dimensions[row].height = h

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# â”€â”€ Data loaders â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def load_comp_df():
    con = sqlite3.connect(DB_PATH)
    rows = con.execute("""
        SELECT brand_detected, grade_detected, pack_size_l, oil_type,
               price_per_litre, price, title, platform
        FROM scraped_products
        WHERE price_per_litre > 100 AND grade_detected IS NOT NULL
          AND brand_detected IS NOT NULL
          AND DATE(scraped_at) = (SELECT DATE(MAX(scraped_at)) FROM scraped_products)
        ORDER BY oil_type, grade_detected, brand_detected, pack_size_l
    """).fetchall()
    con.close()
    return pd.DataFrame(rows, columns=["Brand","Grade","Pack_L","Oil_Type",
                                        "Price_L","Total_Price","Title","Platform"])

def load_pso_skus():
    import openpyxl as xl
    wb = xl.load_workbook(str(LUBES_FILE), data_only=True)
    ws = wb["SKU Wise"]
    skus, current = [], None
    section_hdrs = {"Vol vs Val (PCMO)","Vol vs Val (MCO)","Vol vs Val (DEO)","Vol vs Val (Industrial)"}
    for row in ws.iter_rows(values_only=True):
        v0 = str(row[0]).strip() if row[0] else ""
        if v0 in section_hdrs: current = None; continue
        if v0 in ("Brand","None"): continue
        if v0: current = v0
        if current not in BRAND_MAP: continue
        brand, oil_type = BRAND_MAP[current]
        try:
            pack_l = float(row[1]) if row[1] else None
            vol    = float(row[3]) if row[3] else 0.0
        except: continue
        if not pack_l or pack_l > 4.0 or vol <= 0: continue
        for grade in BRAND_GRADES.get(brand, []):
            skus.append((brand, grade, pack_l, oil_type))
    seen = set(); unique = []
    for s in skus:
        if s not in seen: seen.add(s); unique.append(s)
    return unique

def mkt_stats(df, grade, pack_l=None, excl_pso=True):
    m = df["Grade"] == grade
    if pack_l: m &= df["Pack_L"] == pack_l
    if excl_pso: m &= ~df["Brand"].str.contains("PSO|Carient", na=False)
    vals = df.loc[m, "Price_L"].dropna()
    widened = False
    if vals.empty and pack_l:
        m2 = df["Grade"] == grade
        if excl_pso: m2 &= ~df["Brand"].str.contains("PSO|Carient", na=False)
        vals = df.loc[m2, "Price_L"].dropna()
        widened = True
    if vals.empty: return None, None, None, 0, widened
    return (round(float(vals.min()),0), round(float(vals.median()),0),
            round(float(vals.max()),0), len(vals), widened)

def tier_med(df, tier, grade, pack_l=None):
    brands = SHELL_TIER_BRANDS.get(tier, [])
    m = df["Grade"] == grade
    if pack_l: m &= df["Pack_L"] == pack_l
    m &= df["Brand"].isin(brands)
    vals = df.loc[m, "Price_L"].dropna()
    if vals.empty and pack_l:
        m2 = (df["Grade"] == grade) & df["Brand"].isin(brands)
        vals = df.loc[m2, "Price_L"].dropna()
    return round(float(vals.median()),0) if not vals.empty else None

# â”€â”€ Sheet builders â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def build_competitor_data(wb, df):
    ws = wb.create_sheet("Competitor Data")
    N = 12
    title_row(ws, 1, "COMPETITOR PRICING DATA â€” Daraz.pk Scrape (July 2026)", N)
    desc_row(ws, 2, (
        "365 motor oil listings scraped from Daraz.pk. "
        "Normalised to Price/Litre (PKR) for cross-pack comparison. "
        "Source of truth for all F1, F2, F3, and F5 framework calculations. "
        "PSO listings are flagged but excluded from market benchmark statistics."
    ), N, h=35)
    blank_row(ws, 3)
    header_row(ws, 4, ["#","Brand","Grade","Pack (L)","Oil Type",
                        "Price/L (PKR)","Total Price (PKR)","Product Title","Platform"])
    for idx, row in df.iterrows():
        r = idx + 5
        ws.row_dimensions[r].height = 15
        is_pso = "PSO" in str(row["Brand"]) or "Carient" in str(row["Brand"])
        bg = "FFF8E6" if is_pso else None
        data_cell(ws, r, 1, idx+1, align="right", bg=bg)
        data_cell(ws, r, 2, row["Brand"], align="left", bg=bg)
        data_cell(ws, r, 3, row["Grade"], align="left", bg=bg)
        data_cell(ws, r, 4, row["Pack_L"], fmt="dec", bg=bg)
        data_cell(ws, r, 5, row["Oil_Type"].upper(), align="left", bg=bg)
        data_cell(ws, r, 6, row["Price_L"], fmt="pkr", bg=bg,
                  bold=(row["Price_L"] == df.loc[df["Grade"]==row["Grade"],"Price_L"].median()))
        data_cell(ws, r, 7, row["Total_Price"], fmt="pkr", bg=bg)
        data_cell(ws, r, 8, row["Title"], align="left", bg=bg)
        data_cell(ws, r, 9, row["Platform"].title(), align="left", bg=bg)

    # Summary stats below data
    last = len(df) + 5
    blank_row(ws, last)
    section_row(ws, last+1, "MARKET STATISTICS BY GRADE (excluding PSO)", N)
    header_row(ws, last+2, ["Grade","Oil Type","# Listings","Min Price/L","Median Price/L",
                             "Max Price/L","Std Dev"], start_col=1)
    r = last + 3
    for grade in ["0W-20","5W-20","5W-30","5W-40","10W-40","15W-40","20W-50"]:
        mn, med, mx, cnt, _ = mkt_stats(df, grade)
        oil = "PCMO" if grade in ["0W-20","5W-20","5W-30","5W-40","10W-40"] else "HDEO/PCMO"
        m = df["Grade"]==grade
        m &= ~df["Brand"].str.contains("PSO|Carient", na=False)
        std = round(float(df.loc[m,"Price_L"].std()), 0) if cnt > 1 else None
        data_cell(ws, r, 1, grade, align="left")
        data_cell(ws, r, 2, oil, align="left")
        data_cell(ws, r, 3, cnt)
        data_cell(ws, r, 4, mn, fmt="pkr")
        data_cell(ws, r, 5, med, fmt="pkr", bold=True)
        data_cell(ws, r, 6, mx, fmt="pkr")
        data_cell(ws, r, 7, std, fmt="pkr")
        r += 1

    set_widths(ws, [5, 22, 9, 8, 10, 12, 14, 65, 10])
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = True


def build_f1(wb, df, skus):
    ws = wb.create_sheet("F1 â€“ Value Tiering")
    N = 10
    title_row(ws, 1, "F1 â€” VALUE-BASED TIERING  (Weight: 30%)", N)
    desc_row(ws, 2, (
        "PSO Carient is benchmarked against the equivalent Shell tier "
        "(Super Premium â†’ Shell Helix Ultra; Premium â†’ HX7/ZIC X9; Mainstream â†’ ZIC X7/Total 9000; Economy â†’ Kixx). "
        "A brand-perception discount is applied because PSO brand equity is below Shell in the premium segments. "
        "As brand investment grows, the discount should narrow. "
        "FORMULA: F1 Price = Shell_Tier_Median Ã— (1 âˆ’ Brand_Discount)"
    ), N, h=50)
    blank_row(ws, 3)

    # â”€â”€ PARAM TABLE 1: Shell tier benchmarks â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    section_row(ws, 4, "PARAMETER TABLE 1 â€” Shell Tier Benchmark Prices (Daraz.pk medians, PKR/L)", N)
    grades_pcmo = ["0W-20","5W-20","5W-30","5W-40","10W-40","15W-40","20W-50"]
    header_row(ws, 5, ["Tier","Shell Benchmark Brands"]+grades_pcmo, start_col=1)

    tier_bench = {}
    r = 6
    for tier, brands in SHELL_TIER_BRANDS.items():
        row_data = [TIER_LABEL[tier], ", ".join(brands)]
        prices = {}
        for g in grades_pcmo:
            p = tier_med(df, tier, g)
            row_data.append(p)
            prices[g] = p
        tier_bench[tier] = prices
        data_cell(ws, r, 1, row_data[0], align="left", bold=True)
        data_cell(ws, r, 2, row_data[1], align="left")
        for i, p in enumerate(row_data[2:], 3):
            data_cell(ws, r, i, p, fmt="pkr")
        r += 1

    blank_row(ws, r); r += 1

    # â”€â”€ PARAM TABLE 2: Brand discount â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    section_row(ws, r, "PARAMETER TABLE 2 â€” PSO Brand Perception Discount vs Shell", N)
    r += 1
    header_row(ws, r, ["Tier","Discount %","Rationale"])
    r += 1
    disc_start = r
    disc_rows = {}  # tier â†’ row
    for tier, disc in PSO_BRAND_DISCOUNT.items():
        rationale = {
            "super_premium": "Significant gap â€” fully synthetic brand communication is nascent in Pakistan",
            "premium":       "Moderate gap â€” HX7-class perception; closing as fleet specs improve",
            "mainstream":    "Small gap â€” mineral lubricants mostly price-competitive",
            "economy":       "Minimal gap â€” budget buyers are price-led, brand matters less",
        }[tier]
        disc_rows[tier] = r
        data_cell(ws, r, 1, TIER_LABEL[tier], align="left", bold=True)
        data_cell(ws, r, 2, disc, fmt="pct")
        data_cell(ws, r, 3, rationale, align="left")
        r += 1

    blank_row(ws, r); r += 1

    # â”€â”€ CALCULATION TABLE â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    section_row(ws, r, "CALCULATION TABLE â€” F1 Price per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Pack (L)","Tier",
                        "Shell Benchmark (PKR/L)","Brand Discount","F1 Price (PKR/L)"])
    r += 1

    for brand, grade, pack_l, oil_type in skus:
        tier = PSO_TIER.get(brand, "mainstream")
        shell_p = tier_med(df, tier, grade, pack_l)
        disc = PSO_BRAND_DISCOUNT.get(tier, 0.04)
        f1 = round(shell_p * (1 - disc), 0) if shell_p else None
        bench_col = 6
        disc_col  = 7
        f1_col    = 8
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, pack_l, fmt="dec")
        data_cell(ws, r, 5, TIER_LABEL[tier], align="left")
        data_cell(ws, r, bench_col, shell_p, fmt="pkr")
        data_cell(ws, r, disc_col, disc, fmt="pct")
        if shell_p:
            formula_cell(ws, r, f1_col,
                         f"={cref(r,bench_col)}*(1-{cref(r,disc_col)})", fmt="pkr")
        else:
            data_cell(ws, r, f1_col, "N/A", align="right")
        r += 1

    desc_row(ws, r+1,
             "ðŸ”µ Blue cells are live Excel formulas. Change the Benchmark or Discount values above and the F1 Price recalculates automatically.",
             N, bg=C_LIGHT, h=22)
    set_widths(ws, [28, 9, 8, 8, 15, 20, 15, 18])
    ws.freeze_panes = "A6"


def build_f2(wb, df, skus):
    ws = wb.create_sheet("F2 â€“ Market Ref Price")
    N = 12
    title_row(ws, 1, "F2 â€” COMPETITIVE REFERENCE PRICE  (Weight: 30%)", N)
    desc_row(ws, 2, (
        "F2 anchors PSO prices to the independent market median â€” the price at which half the competitors charge more and half charge less. "
        "PSO is positioned 3% below median (best-value-in-class, not the cheapest). "
        "When no exact pack size has sufficient listings, the grade-level median is used (widened scope, flagged). "
        "FORMULA: F2 Price = Market_Median Ã— 0.97"
    ), N, h=50)
    blank_row(ws, 3)

    section_row(ws, 4, "MARKET STATISTICS TABLE (source: Daraz.pk scrape, July 2026 â€” PSO/Carient excluded)", N)
    header_row(ws, 5, ["Grade","Pack (L)","# Listings","Min (PKR/L)","Median (PKR/L)",
                        "Max (PKR/L)","Scope","F2 = Median Ã— 0.97"])
    r = 6
    # Track (grade, pack_l) â†’ row for formula referencing in SKU table
    stat_row = {}
    # Build unique (grade, pack_l) combos from skus
    combos = sorted(set((g, p) for _, g, p, _ in skus))
    for grade, pack_l in combos:
        mn, med, mx, cnt, widened = mkt_stats(df, grade, pack_l)
        stat_row[(grade, pack_l)] = (r, 5)  # row, median column
        data_cell(ws, r, 1, grade, align="left")
        data_cell(ws, r, 2, pack_l, fmt="dec")
        data_cell(ws, r, 3, cnt)
        data_cell(ws, r, 4, mn, fmt="pkr")
        data_cell(ws, r, 5, med, fmt="pkr", bold=True)
        data_cell(ws, r, 6, mx, fmt="pkr")
        data_cell(ws, r, 7, "Grade-wide" if widened else "Exact pack", align="left")
        if med:
            formula_cell(ws, r, 8, f"={cref(r,5)}*0.97", fmt="pkr")
        else:
            data_cell(ws, r, 8, "N/A", align="right")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "BRAND-LEVEL RESULTS â€” F2 Price per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Pack (L)",
                        "Market Median (PKR/L)","F2 CRP (PKR/L)","vs Median"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        stat_r, med_col = stat_row.get((grade, pack_l), (None, None))
        _, med, _, _, _ = mkt_stats(df, grade, pack_l)
        f2 = round(med * 0.97, 0) if med else None
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, pack_l, fmt="dec")
        if stat_r:
            formula_cell(ws, r, 5, f"='F2 â€“ Market Ref Price'!{cref(stat_r, med_col)}", fmt="pkr")
            formula_cell(ws, r, 6, f"={cref(r,5)}*0.97", fmt="pkr")
            formula_cell(ws, r, 7, f"=({cref(r,6)}-{cref(r,5)})/{cref(r,5)}", fmt="pct1")
        else:
            data_cell(ws, r, 5, med, fmt="pkr")
            data_cell(ws, r, 6, f2, fmt="pkr")
            data_cell(ws, r, 7, "-0.03" if f2 else None, fmt="pct1")
        r += 1

    set_widths(ws, [28, 9, 8, 8, 18, 18, 12])
    ws.freeze_panes = "A6"


def build_f3(wb, df, skus):
    ws = wb.create_sheet("F3 â€“ Pack Architecture")
    N = 9
    title_row(ws, 1, "F3 â€” PRICE-PACK ARCHITECTURE  (Weight: 20%)", N)
    desc_row(ws, 2, (
        "Consumers pay a per-litre premium for smaller packs (convenience) and a discount for larger packs (value). "
        "The 4L pack is the anchor (multiplier = 1.00Ã—). All other pack sizes are priced relative to the 4L market median. "
        "This creates a legible, consistent price ladder that avoids irrational cross-pack pricing. "
        "FORMULA: F3 Price = 4L_Market_Median Ã— PPA_Multiplier"
    ), N, h=55)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE â€” PPA Multipliers (international retail lubricant norms)", N)
    header_row(ws, 5, ["Pack Size (L)","Multiplier vs 4L Anchor",
                        "% Premium/(Discount)","Rationale"])
    r = 6
    mult_rows = {}  # pack_l â†’ row
    for pack_l, mult in PPA_MULTIPLIER.items():
        mult_rows[pack_l] = r
        rationale = {
            0.8:  "Top-up sachet â€” convenience + dispensing premium",
            1.0:  "Retail single-change â€” most common passenger car change",
            1.5:  "Partial change plus reserve",
            2.0:  "DIY market â€” two-thirds change",
            3.0:  "Near-full PCMO change (3L sump common)",
            3.5:  "Slight discount â€” niche size",
            4.0:  "ANCHOR â€” full 4L engine sump change",
            5.0:  "Slight bulk discount â€” light commercial",
            7.0:  "Bulk / workshop supply",
            10.0: "Distributor / fleet bulk",
            20.0: "20L drum â€” workshop / small fleet",
        }.get(pack_l, "")
        data_cell(ws, r, 1, pack_l, fmt="dec", bold=(pack_l == 4.0))
        data_cell(ws, r, 2, mult, fmt="dec", bold=(pack_l == 4.0),
                  bg=C_LIGHT if pack_l == 4.0 else None)
        data_cell(ws, r, 3, f"{(mult-1)*100:+.0f}%", align="right")
        data_cell(ws, r, 4, rationale, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "4L ANCHOR PRICES BY GRADE (from F2 market medians)", N)
    r += 1
    header_row(ws, r, ["Grade","4L Market Median (PKR/L)","Data Points","Note"])
    r += 1
    anchor_rows = {}  # grade â†’ (row, col)
    for grade in sorted(set(g for _, g, _, _ in skus)):
        _, med4, _, cnt4, widened = mkt_stats(df, grade, 4.0)
        anchor_rows[grade] = (r, 2)
        data_cell(ws, r, 1, grade, align="left", bold=True)
        data_cell(ws, r, 2, med4, fmt="pkr", bold=True)
        data_cell(ws, r, 3, cnt4)
        data_cell(ws, r, 4, "Widened to all packs (4L unavailable)" if widened else "Exact 4L match",
                  align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "CALCULATION TABLE â€” F3 Price per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Pack (L)",
                        "4L Anchor (PKR/L)","Multiplier","F3 Price (PKR/L)"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        anch_r, anch_c = anchor_rows.get(grade, (None, None))
        mult = PPA_MULTIPLIER.get(pack_l, 1.0)
        _, med4, _, _, _ = mkt_stats(df, grade, 4.0)
        f3 = round(med4 * mult, 0) if med4 else None
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, pack_l, fmt="dec")
        if anch_r:
            formula_cell(ws, r, 5,
                         f"='F3 â€“ Pack Architecture'!{cref(anch_r, anch_c)}", fmt="pkr")
        else:
            data_cell(ws, r, 5, med4, fmt="pkr")
        data_cell(ws, r, 6, mult, fmt="dec")
        if med4:
            formula_cell(ws, r, 7, f"={cref(r,5)}*{cref(r,6)}", fmt="pkr")
        else:
            data_cell(ws, r, 7, "N/A", align="right")
        r += 1

    set_widths(ws, [28, 9, 8, 8, 20, 12, 18])
    ws.freeze_panes = "A6"


def build_f4(wb, df, skus):
    ws = wb.create_sheet("F4 â€“ Cost Waterfall")
    N = 12
    title_row(ws, 1, "F4 â€” McKINSEY PRICE WATERFALL  (Hard Floor Constraint)", N)
    desc_row(ws, 2, (
        "The waterfall calculates the minimum viable list price from the cost structure up. "
        "It is NOT a pricing target â€” it is a hard floor. If F1/F2/F3 synthesis falls below F4, "
        "the recommendation is raised to F4. Three cost layers: (1) base oil import cost by Group, "
        "(2) additive package by grade, (3) packaging by pack size. "
        "FORMULA: Pocket Floor = Total_Cost Ã· (1 âˆ’ Target_Margin%)  "
        "â†’  List Floor = Pocket_Floor Ã· (1 âˆ’ Fleet_Discount%)"
    ), N, h=60)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE 1 â€” Base Oil Cost by Group (PKR/L, Pakistan import parity, 2025)", N)
    header_row(ws, 5, ["Base Oil Group","Cost (PKR/L)","Typical Grades","Source"])
    r = 6
    bo_rows = {}
    for grp, cost in BASE_OIL_COST.items():
        bo_rows[grp] = (r, 2)
        grades_in_grp = [g for g, gr in BASE_OIL_GROUP.items() if gr == grp]
        src = {"I": "SN150 / BS150 â€” mainly domestic",
               "II": "VHVI â€” Korean/Middle East import",
               "III": "PAO/GTL â€” European/Korean import"}[grp]
        data_cell(ws, r, 1, f"Group {grp}", align="left", bold=True)
        data_cell(ws, r, 2, cost, fmt="pkr")
        data_cell(ws, r, 3, ", ".join(grades_in_grp), align="left")
        data_cell(ws, r, 4, src, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "PARAMETER TABLE 2 â€” Additive Package Cost by Grade (PKR/L)", N)
    r += 1
    header_row(ws, r, ["Grade","Base Oil Group","Additive Cost (PKR/L)","Notes"])
    r += 1
    add_rows = {}
    for grade, add_cost in ADDITIVE_COST.items():
        bog = BASE_OIL_GROUP.get(grade, "II")
        add_rows[grade] = (r, 3)
        note = ("ACEA A5/SN+ additive â€” expensive DI package"
                if bog == "III" else
                "SN/SL additive package" if bog == "II" else
                "SM/SL economy additive")
        data_cell(ws, r, 1, grade, align="left")
        data_cell(ws, r, 2, f"Group {bog}", align="left")
        data_cell(ws, r, 3, add_cost, fmt="pkr")
        data_cell(ws, r, 4, note, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "PARAMETER TABLE 3 â€” Packaging Cost by Pack Size (PKR/L)", N)
    r += 1
    header_row(ws, r, ["Pack Size","Packaging Cost (PKR/L)","Rationale"])
    r += 1
    pkg_rules = [(1, 60, "â‰¤1L"), (5, 35, "1Lâ€“5L"), (99, 20, ">5L drum")]
    pkg_row_map = {}
    for max_p, cost, label in pkg_rules:
        pkg_row_map[label] = (r, 2)
        data_cell(ws, r, 1, label, align="left")
        data_cell(ws, r, 2, cost, fmt="pkr")
        data_cell(ws, r, 3, "Small packs â€” higher per-L moulding & labelling cost" if max_p==1
                  else "Standard HDPE bottle" if max_p==5
                  else "Drum / gallon â€” lower per-L container cost", align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "PARAMETER TABLE 4 â€” Gross Margin Targets by Tier", N)
    r += 1
    header_row(ws, r, ["Tier","Target Pocket Margin %","Rationale"])
    r += 1
    margin_rows = {}
    for tier, margin in TIER_MARGIN.items():
        margin_rows[tier] = (r, 2)
        rationale = {
            "super_premium": "Fully synthetic â€” high spec, brand premium, lower vol â†’ needs 45% to fund A&P",
            "premium":       "Semi-synthetic â€” balanced vol/value, 38% industry norm for mid-tier",
            "mainstream":    "Mineral + partial synthetic â€” volume-driven, 30% sustains distribution margin",
            "economy":       "Commodity mineral â€” price-led, 22% floor preserves channel viability",
        }[tier]
        data_cell(ws, r, 1, TIER_LABEL[tier], align="left", bold=True)
        data_cell(ws, r, 2, margin, fmt="pct")
        data_cell(ws, r, 3, rationale, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "WATERFALL CALCULATION â€” per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Grade","Pack (L)","Tier",
                        "BO Cost","Additive","Pkg","Total Cost",
                        "Target Margin","Pocket Floor","Fleet Disc (20%)","List Price FLOOR"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        tier = PSO_TIER.get(brand, "mainstream")
        bog = BASE_OIL_GROUP.get(grade, "II")
        bo_c = BASE_OIL_COST.get(bog, 420)
        add_c = ADDITIVE_COST.get(grade, 140)
        pkg_c = 60 if pack_l <= 1 else 35 if pack_l <= 5 else 20
        total = bo_c + add_c + pkg_c
        margin = TIER_MARGIN.get(tier, 0.30)
        pocket = round(total / (1 - margin), 0)
        list_floor = round(pocket / 0.80, 0)   # 1 - 0.20 fleet discount

        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, grade, align="left")
        data_cell(ws, r, 3, pack_l, fmt="dec")
        data_cell(ws, r, 4, TIER_LABEL[tier], align="left")
        data_cell(ws, r, 5, bo_c, fmt="pkr")
        data_cell(ws, r, 6, add_c, fmt="pkr")
        data_cell(ws, r, 7, pkg_c, fmt="pkr")
        formula_cell(ws, r, 8, f"={cref(r,5)}+{cref(r,6)}+{cref(r,7)}", fmt="pkr")
        data_cell(ws, r, 9, margin, fmt="pct")
        formula_cell(ws, r, 10, f"={cref(r,8)}/(1-{cref(r,9)})", fmt="pkr")
        data_cell(ws, r, 11, 0.20, fmt="pct")
        formula_cell(ws, r, 12, f"={cref(r,10)}/(1-{cref(r,11)})", fmt="pkr", bold=True)
        r += 1

    desc_row(ws, r+1,
             "The LIST PRICE FLOOR (col L) is the absolute minimum. PSO must never price below this at the retail pump. "
             "Fleet contract prices will automatically be 20% below this floor â€” which means fleet pricing must negotiate off a higher list.",
             N, bg="FFF8E6", h=30)

    set_widths(ws, [26, 8, 8, 15, 10, 10, 8, 12, 13, 13, 12, 16])
    ws.freeze_panes = "A13"


def build_f5(wb, df, skus):
    ws = wb.create_sheet("F5 â€“ Spec Premium")
    N = 9
    title_row(ws, 1, "F5 â€” OEM / API SPECIFICATION PREMIUM  (Weight: 10%)", N)
    desc_row(ws, 2, (
        "Engine oil prices should reflect API/ACEA specification level. "
        "A fully-synthetic API SP product justifies a higher price than a mineral API SN product of the same viscosity grade. "
        "This framework anchors to the economy-tier (Kixx/Kixx G1) market price for each grade, "
        "then applies a spec uplift multiplier for PSO's tier. "
        "FORMULA: F5 Price = Economy_Baseline Ã— (1 + Spec_Uplift%)"
    ), N, h=55)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE 1 â€” Economy Baseline Prices by Grade (Kixx/Kixx G1 medians, PKR/L)", N)
    header_row(ws, 5, ["Grade","Economy Baseline (PKR/L)","# Listings","Notes"])
    r = 6
    eco_rows = {}
    for grade in ["0W-20","5W-20","5W-30","5W-40","10W-40","15W-40","20W-50"]:
        eco_brands = SHELL_TIER_BRANDS["economy"]
        m = (df["Grade"] == grade) & df["Brand"].isin(eco_brands)
        vals = df.loc[m, "Price_L"].dropna()
        if vals.empty:
            mn2, med2, _, cnt2, _ = mkt_stats(df, grade)
            baseline = mn2
        else:
            baseline = round(float(vals.median()), 0)
            cnt2 = len(vals)
        eco_rows[grade] = (r, 2)
        data_cell(ws, r, 1, grade, align="left")
        data_cell(ws, r, 2, baseline, fmt="pkr", bold=True)
        data_cell(ws, r, 3, cnt2)
        data_cell(ws, r, 4, "Kixx median" if not vals.empty else "Market min (economy proxy)", align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "PARAMETER TABLE 2 â€” Spec Uplift by Tier", N)
    r += 1
    header_row(ws, r, ["Tier","API/ACEA Spec","Uplift %","Rationale"])
    r += 1
    spec_rows = {}
    spec_desc = {
        "super_premium": ("API SP / ACEA A5", 0.28,
                          "Fully synthetic Group III â€” OEM fuel economy approvals (VW 504, BMW LL-04) command 28% premium"),
        "premium":       ("API SN+ / ACEA A3", 0.18,
                          "Semi-synthetic Group II/III blend â€” common fleet spec approval, 18% above economy"),
        "mainstream":    ("API SN / ACEA A3", 0.08,
                          "Mineral / part-synthetic Group I/II â€” basic OEM approval, 8% above pure economy"),
        "economy":       ("API SN / SM", 0.00,
                          "Mineral Group I â€” baseline, no spec uplift"),
    }
    for tier, (spec, uplift, rat) in spec_desc.items():
        spec_rows[tier] = (r, 3)
        data_cell(ws, r, 1, TIER_LABEL[tier], align="left", bold=True)
        data_cell(ws, r, 2, spec, align="left")
        data_cell(ws, r, 3, uplift, fmt="pct")
        data_cell(ws, r, 4, rat, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "CALCULATION TABLE â€” F5 Price per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Tier","Economy Baseline","Spec Uplift","F5 Price"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        tier = PSO_TIER.get(brand, "mainstream")
        eco_r, eco_c = eco_rows.get(grade, (None, None))
        spec_r, spec_c = spec_rows.get(tier, (None, None))
        eco_m = (df["Grade"] == grade) & df["Brand"].isin(SHELL_TIER_BRANDS["economy"])
        eco_vals = df.loc[eco_m, "Price_L"].dropna()
        if eco_vals.empty:
            eco_b, *_ = mkt_stats(df, grade)
        else:
            eco_b = round(float(eco_vals.median()), 0)
        uplift = API_SPEC_PREMIUM.get(tier, 0)
        f5 = round(eco_b * (1 + uplift), 0) if eco_b else None

        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, TIER_LABEL[tier], align="left")
        if eco_r:
            formula_cell(ws, r, 5, f"='F5 â€“ Spec Premium'!{cref(eco_r, eco_c)}", fmt="pkr")
        else:
            data_cell(ws, r, 5, eco_b, fmt="pkr")
        if spec_r:
            formula_cell(ws, r, 6, f"='F5 â€“ Spec Premium'!{cref(spec_r, spec_c)}", fmt="pct")
        else:
            data_cell(ws, r, 6, uplift, fmt="pct")
        formula_cell(ws, r, 7, f"={cref(r,5)}*(1+{cref(r,6)})", fmt="pkr", bold=True)
        r += 1

    set_widths(ws, [28, 9, 8, 15, 20, 12, 18])
    ws.freeze_panes = "A6"


def build_f6(wb, df, skus):
    ws = wb.create_sheet("F6 â€“ Geographic")
    N = 10
    title_row(ws, 1, "F6 â€” GEOGRAPHIC PRICE SEGMENTATION", N)
    desc_row(ws, 2, (
        "PSO's retail fuels and lubricants data (FY25) shows strong regional volume variance. "
        "DEO Central is +18.6% YoY; MCO North is âˆ’27.3% YoY. "
        "These signals translate into price adjustments: growing regions can carry a small premium; "
        "declining regions require a discount to defend volume. "
        "Adjustments are PROMOTIONAL â€” applied quarterly, not as permanent list price changes. "
        "FORMULA: Regional Price = F2_National_Price Ã— (1 + Regional_Signal%)"
    ), N, h=60)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE â€” Regional Growth Signals (source: PSO Lubes Data Final.xlsx FY25 vs FY24)", N)
    header_row(ws, 5, ["Oil Category","Region","Volume Signal","Price Adjustment","Source / Rationale"])
    r = 6
    signal_rows = {}  # (oil_type, region) â†’ (row, adj_col)
    for oil_type, regions in REGIONAL_SIGNAL.items():
        for region, adj in regions.items():
            vol_str = {
                ("hdeo","Central"): "+18.6% YoY DEO Central",
                ("mco","North"):    "-27.3% YoY MCO North",
                ("pcmo","North"):   "Soft YoY PCMO North",
            }.get((oil_type, region), "Regional baseline")
            signal_rows[(oil_type, region)] = (r, 4)
            data_cell(ws, r, 1, oil_type.upper(), align="left")
            data_cell(ws, r, 2, region, align="left")
            data_cell(ws, r, 3, vol_str, align="left")
            data_cell(ws, r, 4, adj, fmt="pct1")
            data_cell(ws, r, 5,
                      f"{'Raise price by' if adj>0 else 'Discount by' if adj<0 else 'No adjustment â€”'} "
                      f"{abs(adj*100):.0f}% {'vs national' if adj!=0 else '(South is national baseline)'}",
                      align="left")
            r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "CALCULATION TABLE â€” Regional Prices per PSO SKU", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Pack (L)",
                        "F2 National (PKR/L)","South Adj","South Price",
                        "Central Adj","Central Price","North Adj","North Price"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        _, med, _, _, _ = mkt_stats(df, grade, pack_l)
        f2 = round(med * 0.97, 0) if med else None
        signals = REGIONAL_SIGNAL.get(oil_type.lower(), {})
        s_adj  = signals.get("South", 0)
        c_adj  = signals.get("Central", 0)
        n_adj  = signals.get("North", 0)
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, pack_l, fmt="dec")
        data_cell(ws, r, 5, f2, fmt="pkr")
        data_cell(ws, r, 6, s_adj, fmt="pct1")
        formula_cell(ws, r, 7,  f"={cref(r,5)}*(1+{cref(r,6)})", fmt="pkr")
        data_cell(ws, r, 8, c_adj, fmt="pct1")
        formula_cell(ws, r, 9,  f"={cref(r,5)}*(1+{cref(r,8)})", fmt="pkr")
        data_cell(ws, r, 10, n_adj, fmt="pct1")
        formula_cell(ws, r, 11, f"={cref(r,5)}*(1+{cref(r,10)})", fmt="pkr")
        r += 1

    desc_row(ws, r+1,
             "South (Karachi/Hyderabad/Sukkur) = national baseline. "
             "Central = Lahore/Faisalabad/Multan. North = Islamabad/Peshawar/Rawalpindi. "
             "Adjust Central and North price quarterly via promotional price list.", N, bg=C_LIGHT, h=25)

    set_widths(ws, [26, 9, 8, 8, 18, 10, 16, 12, 16, 10, 16])
    ws.freeze_panes = "A6"


def build_f7(wb, df, skus):
    ws = wb.create_sheet("F7 â€“ Base Oil Floor")
    N = 8
    title_row(ws, 1, "F7 â€” BASE OIL INDEX PRICE FLOOR", N)
    desc_row(ws, 2, (
        "F7 provides an independent sanity check: the minimum viable price derived purely from base oil import costs. "
        "If a PSO price falls below F7, it either implies subsidised base oil (unsustainable) or accounting error. "
        "Markup of 2.8Ã— is the industry standard retail lubricant multiple (yields ~64% gross margin). "
        "FORMULA: F7 Floor = (Base_Oil_Cost + Additive_Cost + 50_overhead) Ã— 2.8"
    ), N, h=55)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER INPUTS", N)
    header_row(ws, 5, ["Parameter","Value","Unit","Notes"])
    params = [
        ("Retail Markup Multiple", MIN_MARKUP, "Ã—", "Industry norm: 2.8Ã— manufacturing cost = ~64% gross margin"),
        ("Blending/overhead add-on", 50, "PKR/L", "Blending, QC, filling, fixed overhead allocation"),
        ("Group I base oil cost", 280, "PKR/L", "Domestic SN150/BS150, Pakistan market 2025"),
        ("Group II base oil cost", 420, "PKR/L", "VHVI import â€” Korean/Middle East, Pakistan 2025"),
        ("Group III base oil cost", 680, "PKR/L", "PAO/GTL import â€” Europe/Korea, Pakistan 2025"),
    ]
    param_rows = {}
    r = 6
    for param, val, unit, note in params:
        data_cell(ws, r, 1, param, align="left", bold=True)
        data_cell(ws, r, 2, val, fmt="dec")
        data_cell(ws, r, 3, unit, align="left")
        data_cell(ws, r, 4, note, align="left")
        param_rows[param] = (r, 2)
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "CALCULATION TABLE â€” F7 Floor per Grade", N)
    r += 1
    header_row(ws, r, ["Grade","Base Oil Group","BO Cost (PKR/L)","Additive (PKR/L)",
                        "Overhead","Total Mfg Cost","Ã—2.8 Markup","F7 Floor (PKR/L)"])
    r += 1
    floor_rows = {}  # grade â†’ (row, col 8)
    for grade in sorted(set(g for _, g, _, _ in skus)):
        bog = BASE_OIL_GROUP.get(grade, "II")
        bo_c = BASE_OIL_COST.get(bog, 420)
        add_c = ADDITIVE_COST.get(grade, 140)
        total = bo_c + add_c + 50
        f7 = round(total * MIN_MARKUP, 0)
        floor_rows[grade] = (r, 8)
        data_cell(ws, r, 1, grade, align="left", bold=True)
        data_cell(ws, r, 2, f"Group {bog}", align="left")
        data_cell(ws, r, 3, bo_c, fmt="pkr")
        data_cell(ws, r, 4, add_c, fmt="pkr")
        data_cell(ws, r, 5, 50, fmt="pkr")
        formula_cell(ws, r, 6, f"={cref(r,3)}+{cref(r,4)}+{cref(r,5)}", fmt="pkr")
        data_cell(ws, r, 7, MIN_MARKUP, fmt="dec")
        formula_cell(ws, r, 8, f"={cref(r,6)}*{cref(r,7)}", fmt="pkr", bold=True)
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "SKU-LEVEL FLOOR CHECK", N)
    r += 1
    header_row(ws, r, ["Brand","Grade","Pack (L)","F7 Floor (PKR/L)","Note"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        fl_r, fl_c = floor_rows.get(grade, (None, None))
        bog = BASE_OIL_GROUP.get(grade, "II")
        total = BASE_OIL_COST.get(bog, 420) + ADDITIVE_COST.get(grade, 140) + 50
        f7 = round(total * MIN_MARKUP, 0)
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, grade, align="left")
        data_cell(ws, r, 3, pack_l, fmt="dec")
        if fl_r:
            formula_cell(ws, r, 4, f"='F7 â€“ Base Oil Floor'!{cref(fl_r, fl_c)}", fmt="pkr")
        else:
            data_cell(ws, r, 4, f7, fmt="pkr")
        data_cell(ws, r, 5, "Any price below this indicates cost subsidisation or margin destruction", align="left")
        r += 1

    set_widths(ws, [26, 14, 14, 14, 12, 16, 12, 18])
    ws.freeze_panes = "A6"


def build_f8(wb, df, skus, final_prices):
    ws = wb.create_sheet("F8 â€“ Channel Pricing")
    N = 9
    title_row(ws, 1, "F8 â€” CHANNEL PRICING MATRIX", N)
    desc_row(ws, 2, (
        "The recommended list price (from weighted synthesis) cascades to four trade channels via a structured discount waterfall. "
        "Retail pump pays full list price. Workshops get 12% trade discount. "
        "Distributors get 15%. Fleet contract customers â€” who commit to volume â€” get 20%. "
        "Channel prices must ALWAYS stay above the F4 waterfall floor, even for fleet. "
        "FORMULA: Channel Price = List_Price Ã— (1 âˆ’ Channel_Discount%)"
    ), N, h=55)
    blank_row(ws, 3)

    section_row(ws, 4, "PARAMETER TABLE â€” Channel Discount Structure", N)
    header_row(ws, 5, ["Channel","Discount Off List","Rationale","Typical Customer"])
    r = 6
    ch_desc = {
        "Retail (Pump)": (0.00, "Full list price â€” pump is PSO's most valuable channel for margin",
                           "Individual car owner at petrol station"),
        "Workshop":       (0.12, "12% trade allowance â€” compensates stocking and application labour",
                           "Auto workshop, lube shop, tyre shop"),
        "Distributor":    (0.15, "15% margin for regional distributors who take volume risk",
                           "Regional wholesale distributor"),
        "Fleet Contract": (0.20, "20% for committed annual volume â€” volume + predictability trade-off",
                           "Trucks, buses, ride-hailing fleets (>50 vehicles)"),
    }
    ch_rows = {}
    for ch, (disc, rat, cust) in ch_desc.items():
        ch_rows[ch] = (r, 2)
        data_cell(ws, r, 1, ch, align="left", bold=True)
        data_cell(ws, r, 2, disc, fmt="pct")
        data_cell(ws, r, 3, rat, align="left")
        data_cell(ws, r, 4, cust, align="left")
        r += 1

    blank_row(ws, r); r += 1
    section_row(ws, r, "CALCULATION TABLE â€” Channel Prices per PSO SKU (based on Final Recommended List Price)", N)
    r += 1
    header_row(ws, r, ["Brand","Oil Type","Grade","Pack (L)","List Price (PKR/L)",
                        "Retail Pump","Workshop (âˆ’12%)","Distributor (âˆ’15%)","Fleet (âˆ’20%)"])
    r += 1
    for brand, grade, pack_l, oil_type in skus:
        key = (brand, grade, pack_l)
        list_p = final_prices.get(key)
        data_cell(ws, r, 1, brand, align="left")
        data_cell(ws, r, 2, oil_type.upper(), align="left")
        data_cell(ws, r, 3, grade, align="left")
        data_cell(ws, r, 4, pack_l, fmt="dec")
        data_cell(ws, r, 5, list_p, fmt="pkr", bold=True, bg=C_LIGHT)
        formula_cell(ws, r, 6, f"={cref(r,5)}*(1-0.00)", fmt="pkr")
        formula_cell(ws, r, 7, f"={cref(r,5)}*(1-0.12)", fmt="pkr")
        formula_cell(ws, r, 8, f"={cref(r,5)}*(1-0.15)", fmt="pkr")
        formula_cell(ws, r, 9, f"={cref(r,5)}*(1-0.20)", fmt="pkr")
        r += 1

    desc_row(ws, r+1,
             "The List Price column (E) is the final recommended PKR/L from the Summary sheet. "
             "Changing list price in column E will automatically recalculate all channel prices. "
             "NEVER allow fleet price (col I) to fall below the F4 waterfall floor.",
             N, bg="FFF8E6", h=30)

    set_widths(ws, [26, 9, 8, 8, 18, 16, 18, 18, 16])
    ws.freeze_panes = "A7"


def build_summary(wb, df, skus):
    ws = wb.create_sheet("Summary â€“ Final Results")
    N = 20
    title_row(ws, 1, "PSO LUBRICANTS â€” PRICING STRATEGY SUMMARY  (Frameworks F1â€“F8, July 2026)", N)
    desc_row(ws, 2, (
        "Final recommended price per litre for all PSO Carient, DEO, and Blaze SKUs. "
        "Synthesis: weighted average of F1 (30%) + F2 (30%) + F3 (20%) + F5 (10%) + F4 (10%), "
        "with F4 waterfall as hard floor. Confidence: HIGH if â‰¥4 frameworks have data; MEDIUM if â‰¥2; LOW otherwise."
    ), N, h=40)
    blank_row(ws, 3)

    # Framework weights explanation row
    section_row(ws, 4, "FRAMEWORK WEIGHTS â€” Synthesis formula", N, bg=C_GOLD)
    ws.cell(row=4, column=1).font = Font(bold=True, size=9, color="000000", name="Calibri")
    header_row(ws, 5, [
        "Brand","Oil Type","Grade","Pack (L)","Tier",
        "Mkt Min","Mkt Med","Mkt Max",
        "F1 VBT\n30%","F2 CRP\n30%","F3 PPA\n20%","F4 Floor\n(hard)","F5 Spec\n10%",
        "F7 BOI\nFloor","Synth\n(wt. avg)","F4 Floor\nOverride?",
        "Final\nRec/L","Final\nPkg PKR","Confidence","vs Market"
    ])

    final_prices = {}
    r = 6
    for brand, grade, pack_l, oil_type in skus:
        tier = PSO_TIER.get(brand, "mainstream")

        # Market stats
        mn, med, mx, cnt, _ = mkt_stats(df, grade, pack_l)

        # F1
        shell_p = tier_med(df, tier, grade, pack_l)
        disc = PSO_BRAND_DISCOUNT.get(tier, 0.04)
        f1 = round(shell_p * (1 - disc), 0) if shell_p else None

        # F2
        f2 = round(med * 0.97, 0) if med else None

        # F3
        _, med4, _, _, _ = mkt_stats(df, grade, 4.0)
        mult = PPA_MULTIPLIER.get(pack_l, 1.0)
        f3 = round(med4 * mult, 0) if med4 else None

        # F4
        bog = BASE_OIL_GROUP.get(grade, "II")
        bo_c = BASE_OIL_COST.get(bog, 420)
        add_c = ADDITIVE_COST.get(grade, 140)
        pkg_c = 60 if pack_l <= 1 else 35 if pack_l <= 5 else 20
        margin = TIER_MARGIN.get(tier, 0.30)
        pocket = (bo_c + add_c + pkg_c) / (1 - margin)
        f4 = round(pocket / 0.80, 0)

        # F5
        eco_m = (df["Grade"] == grade) & df["Brand"].isin(SHELL_TIER_BRANDS["economy"])
        eco_vals = df.loc[eco_m, "Price_L"].dropna()
        if eco_vals.empty:
            eco_b = mn
        else:
            eco_b = round(float(eco_vals.median()), 0)
        uplift = API_SPEC_PREMIUM.get(tier, 0)
        f5 = round(eco_b * (1 + uplift), 0) if eco_b else None

        # F7
        f7 = round((bo_c + add_c + 50) * MIN_MARKUP, 0)

        # Synthesis
        fw_vals = [(f1, 0.30), (f2, 0.30), (f3, 0.20), (f5, 0.10), (f4, 0.10)]
        avail = [(v, w) for v, w in fw_vals if v]
        if avail:
            tw = sum(w for _, w in avail)
            synth = round(sum(v * w for v, w in avail) / tw, 0)
        else:
            synth = None
        overridden = synth is not None and synth < f4
        rec = max(synth, f4) if synth else f4
        confidence = ("HIGH" if sum(1 for v,_ in fw_vals if v) >= 4 else
                      "MEDIUM" if sum(1 for v,_ in fw_vals if v) >= 2 else "LOW")
        final_prices[(brand, grade, pack_l)] = rec

        # vs market
        vs_med = round((rec - med) / med * 100, 1) if rec and med else None
        signal = ("PREMIUM" if vs_med and vs_med > 10 else
                  "AT MARKET" if vs_med and vs_med > -10 else
                  "VALUE POSITION" if vs_med is not None else "INSUFF. DATA")

        # Write row
        bg_sig = {"PREMIUM": "FFF8E6", "AT MARKET": C_LIGHT,
                  "VALUE POSITION": "FFE6E6", "INSUFF. DATA": C_GRAY}.get(signal, None)

        data_cell(ws, r,  1, brand, align="left")
        data_cell(ws, r,  2, oil_type.upper(), align="left")
        data_cell(ws, r,  3, grade, align="left")
        data_cell(ws, r,  4, pack_l, fmt="dec")
        data_cell(ws, r,  5, TIER_LABEL[tier], align="left")
        data_cell(ws, r,  6, mn,  fmt="pkr")
        data_cell(ws, r,  7, med, fmt="pkr")
        data_cell(ws, r,  8, mx,  fmt="pkr")
        data_cell(ws, r,  9, f1,  fmt="pkr")
        data_cell(ws, r, 10, f2,  fmt="pkr")
        data_cell(ws, r, 11, f3,  fmt="pkr")
        data_cell(ws, r, 12, f4,  fmt="pkr")
        data_cell(ws, r, 13, f5,  fmt="pkr")
        data_cell(ws, r, 14, f7,  fmt="pkr")
        data_cell(ws, r, 15, synth, fmt="pkr")
        data_cell(ws, r, 16, "YES â†’ raised to F4" if overridden else "No",
                  align="left", bg="FFE6E6" if overridden else None)
        data_cell(ws, r, 17, rec, fmt="pkr", bold=True, bg=C_LIGHT)
        data_cell(ws, r, 18,
                  round(rec * pack_l, 0) if rec else None,
                  fmt="pkr", bold=True, bg=C_LIGHT)
        data_cell(ws, r, 19, confidence, align="left",
                  bg={"HIGH":"D5F5E3","MEDIUM":C_LIGHT,"LOW":"FFE6E6"}.get(confidence))
        data_cell(ws, r, 20,
                  f"{vs_med:+.1f}%" if vs_med is not None else "N/A",
                  align="left", bg=bg_sig)
        r += 1

    # Footnotes
    blank_row(ws, r)
    desc_row(ws, r+1,
             "SYNTHESIS FORMULA: Weighted average where available frameworks contribute proportionally. "
             "F1+F2+F3 together = 80% weight â€” the market-facing signals dominate. "
             "F4 is both a 10%-weight input AND a hard floor â€” if weighted avg < F4, rec is set to F4. "
             "F6 (geographic) and F8 (channel) apply AFTER synthesis as overlays, not inputs.",
             N, bg="FEFEFE", h=40)

    desc_row(ws, r+2,
             "MARKET SIGNAL: PREMIUM = PSO price >10% above market median (justify with spec/brand). "
             "AT MARKET = within Â±10% (correct positioning). "
             "VALUE POSITION = PSO >10% below median (revenue upside â€” consider raising).",
             N, bg=C_LIGHT, h=30)

    set_widths(ws, [26, 9, 8, 7, 14, 9, 9, 9, 10, 10, 10, 10, 10, 10, 10, 16, 12, 12, 10, 13])
    ws.freeze_panes = "A6"
    return final_prices


def build_readme(wb):
    ws = wb.create_sheet("README", 0)
    N = 8
    title_row(ws, 1, "PSO LUBRICANTS â€” PRICING STRATEGY WORKBOOK", N, size=15)
    desc_row(ws, 2, "Built by: Claude Code (claude-sonnet-4-6) | Data: Daraz.pk scrape + PSO Lubes Data Final.xlsx | Date: July 2026", N)
    blank_row(ws, 3)
    section_row(ws, 4, "HOW TO USE THIS WORKBOOK", N)
    rows_readme = [
        ("README (this sheet)", "Overview and navigation guide"),
        ("Competitor Data", "All 365 motor oil listings scraped from Daraz.pk. Basis for all market benchmarks."),
        ("F1 â€“ Value Tiering", "Shell-equivalent benchmarking. PSO price = Shell tier median Ã— (1 âˆ’ brand discount). Weight: 30%."),
        ("F2 â€“ Market Ref Price", "Market median price Â± 3% positioning. Strongest market-side signal. Weight: 30%."),
        ("F3 â€“ Pack Architecture", "Price ladder by pack size â€” 4L anchor Ã— multiplier. Ensures legible price steps. Weight: 20%."),
        ("F4 â€“ Cost Waterfall", "McKinsey waterfall â€” cost floor up. This is a HARD FLOOR, not a target. Overrides if synthesis is below."),
        ("F5 â€“ Spec Premium", "API/ACEA specification uplift. Economy baseline Ã— spec multiplier. Weight: 10%."),
        ("F6 â€“ Geographic", "Regional price adjustments from PSO volume data. Promotional overlay â€” not a permanent list change."),
        ("F7 â€“ Base Oil Floor", "Sanity check floor: manufacturing cost Ã— 2.8Ã— markup. If price < F7, cost structure is broken."),
        ("F8 â€“ Channel Pricing", "Channel cascade: List â†’ Workshop (âˆ’12%) â†’ Distributor (âˆ’15%) â†’ Fleet (âˆ’20%)."),
        ("Summary â€“ Final Results", "All SKUs with every framework price, weighted synthesis, and final recommendation."),
    ]
    header_row(ws, 5, ["Sheet","Description"])
    for i, (sheet, desc) in enumerate(rows_readme, 6):
        ws.row_dimensions[i].height = 22
        data_cell(ws, i, 1, sheet, align="left", bold=True, bg=C_LIGHT if i%2==0 else None)
        data_cell(ws, i, 2, desc, align="left", bg=C_LIGHT if i%2==0 else None)

    blank_row(ws, len(rows_readme)+7)
    section_row(ws, len(rows_readme)+8, "KEY FORMULA LOGIC", N)
    formulas = [
        ("F1 Price/L", "= Shell_Tier_Median Ã— (1 âˆ’ Brand_Discount)", "Shell benchmark from Daraz medians"),
        ("F2 Price/L", "= Market_Median Ã— 0.97", "3% below market = best value, not cheapest"),
        ("F3 Price/L", "= 4L_Market_Anchor Ã— PPA_Multiplier", "Multiplier encodes pack-size premium/discount"),
        ("F4 Floor/L", "= (BO_Cost + Additive + Packaging) Ã· (1 âˆ’ Margin%) Ã· (1 âˆ’ 20%)", "Pocket price â†’ List price via fleet discount"),
        ("F5 Price/L", "= Economy_Baseline Ã— (1 + Spec_Uplift%)", "Uplift: Super Prem=28%, Premium=18%, Mainstream=8%"),
        ("F6 Price/L", "= F2_National Ã— (1 + Regional_Signal%)", "Central +2â€“5%, North âˆ’2â€“8%, South = baseline"),
        ("F7 Floor/L", "= (BO_Cost + Additive + 50) Ã— 2.8", "2.8Ã— = industry retail lubricants markup multiple"),
        ("F8 Channels", "= List_Price Ã— (1 âˆ’ Channel_Discount%)", "0% / 12% / 15% / 20% off list"),
        ("Synthesis",   "= Î£(F_i Ã— Weight_i) Ã· Î£(Weight_i)", "F1=30%, F2=30%, F3=20%, F5=10%, F4=10%"),
        ("Final Rec",   "= MAX(Synthesis, F4_Floor)", "Hard floor ensures margin is never destroyed"),
    ]
    r = len(rows_readme) + 9
    header_row(ws, r, ["Framework","Excel Formula Pattern","Description"])
    r += 1
    for fw, formula, desc in formulas:
        ws.row_dimensions[r].height = 18
        data_cell(ws, r, 1, fw, align="left", bold=True)
        data_cell(ws, r, 2, formula, align="left",
                  bg="EEF7FF")
        data_cell(ws, r, 3, desc, align="left")
        r += 1

    set_widths(ws, [28, 65, 55])
    ws.freeze_panes = "A2"


# â”€â”€ Main â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def main():
    print("Loading data...")
    df = load_comp_df()
    print(f"  Competitor listings: {len(df)}")
    skus = load_pso_skus()
    print(f"  PSO SKUs: {len(skus)}")

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    print("Building sheets...")
    build_readme(wb)
    build_competitor_data(wb, df)
    print("  âœ“ Competitor Data")

    # Summary first pass to get final prices
    final_prices = build_summary(wb, df, skus)
    print("  âœ“ Summary")

    build_f1(wb, df, skus)
    print("  âœ“ F1")
    build_f2(wb, df, skus)
    print("  âœ“ F2")
    build_f3(wb, df, skus)
    print("  âœ“ F3")
    build_f4(wb, df, skus)
    print("  âœ“ F4")
    build_f5(wb, df, skus)
    print("  âœ“ F5")
    build_f6(wb, df, skus)
    print("  âœ“ F6")
    build_f7(wb, df, skus)
    print("  âœ“ F7")
    build_f8(wb, df, skus, final_prices)
    print("  âœ“ F8")

    # Strip self-referencing cross-sheet prefixes that cause Excel repair prompts
    import re as _re
    fixed = 0
    for ws in wb.worksheets:
        pat = _re.compile(r"='" + _re.escape(ws.title) + r"'!", _re.IGNORECASE)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    new = pat.sub("=", cell.value)
                    if new != cell.value:
                        cell.value = new
                        fixed += 1
    if fixed:
        print(f"  Fixed {fixed} self-referencing formulas.")

    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")
    print(f"Sheets: {len(wb.sheetnames)}")
    print(f"SKUs: {len(skus)} | Competitor listings: {len(df)}")


if __name__ == "__main__":
    main()

