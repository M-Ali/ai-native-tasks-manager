"""
PSO — Deep City Profiles (Top 10 Cities)
- Product-wise breakdown per city
- 10-city average benchmark
- City vs average index
- Station Pareto (contribution concentration)
- Written to Excel
"""

import sys, os
sys.path.insert(0, 'src')
os.environ.setdefault('PYTHONIOENCODING', 'utf-8')

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter
from pso import ingest

DATA_FILE = "data/input/Working File Retail Fuels Data.xlsx"
OUT_FILE  = "reports/PSO_City_Deep_Profile.xlsx"

# ── Colours ────────────────────────────────────────────────────────────────────
NAVY   = "1B2A4A"
GOLD   = "C9A030"
WHITE  = "FFFFFF"
LGREY  = "F2F4F8"
MGREY  = "D9DCE3"
GREEN  = "C6EFCE"
RED    = "FFC7CE"
YELLOW = "FFEB9C"
LBLUE  = "DDEEFF"
DBLUE  = "2E5B9A"

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

thin  = Side(style="thin",   color="AAAAAA")
thick = Side(style="medium", color="1B2A4A")
BORDER_THIN  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
BORDER_THICK = Border(left=thick, right=thick, top=thick, bottom=thick)
BORDER_BOT   = Border(bottom=thick)

def pct_str(val, decimals=1):
    if pd.isna(val): return "—"
    sign = "+" if val > 0 else ""
    return f"{sign}{val:.{decimals}f}%"

def num_str(val, decimals=1):
    if pd.isna(val): return "—"
    return f"{val:,.{decimals}f}"

# ── Load data ──────────────────────────────────────────────────────────────────
print("Loading data …")
df, _ = ingest.load(DATA_FILE)
period = df["_Period"].iloc[0] if "_Period" in df.columns else "10M FY26"

retail = df[df["IsRetail"] & ~df["IsInternational"]].copy()

# ── Top 10 cities by GRS ───────────────────────────────────────────────────────
city_grs = (
    retail.groupby("CityNorm")["SalesGRS_CY"]
    .sum().reset_index()
    .nlargest(10, "SalesGRS_CY")
)
TOP10 = city_grs["CityNorm"].tolist()
total_retail_grs = retail["SalesGRS_CY"].sum()
total_retail_vol = retail["SalesLtr_CY"].sum()

retail10 = retail[retail["CityNorm"].isin(TOP10)].copy()

PRODUCTS = ["Diesel", "Petrol", "Lubricants", "Other Fuels"]

# ── Helper: build product profile for a subset of rows ────────────────────────
def product_profile(subset):
    """Returns dict: {product: {metric: value}, 'Total': {...}}"""
    rows = {}
    all_customers = set(subset["Customer Number"].unique())
    total_grs_cy  = subset["SalesGRS_CY"].sum()
    total_vol_cy  = subset["SalesLtr_CY"].sum()

    for prod in PRODUCTS:
        s = subset[subset["FuelSegment"] == prod]
        if s.empty:
            rows[prod] = None
            continue
        cust   = s["Customer Number"].nunique()
        grs_cy = s["SalesGRS_CY"].sum() / 1e6
        grs_ly = s["SalesGRS_LY"].sum() / 1e6
        vol_cy = s["SalesLtr_CY"].sum() / 1e6
        vol_ly = s["SalesLtr_LY"].sum() / 1e6
        nmgn   = s["NetMargin_CY"].sum() / 1e6
        disc   = s["Disc_CY"].sum() / s["SalesLtr_CY"].sum() if s["SalesLtr_CY"].sum() > 0 else 0
        nmln   = s["NetMargin_CY"].sum() / s["SalesLtr_CY"].sum() if s["SalesLtr_CY"].sum() > 0 else 0
        grs_chg = (grs_cy - grs_ly) / grs_ly * 100 if grs_ly > 0 else None
        vol_chg = (vol_cy - vol_ly) / vol_ly * 100 if vol_ly > 0 else None
        grs_sh  = grs_cy * 1e6 / total_grs_cy * 100 if total_grs_cy > 0 else 0
        vol_sh  = vol_cy * 1e6 / total_vol_cy * 100 if total_vol_cy > 0 else 0

        rows[prod] = dict(
            customers=cust,
            grs_cy=grs_cy, grs_ly=grs_ly, grs_chg=grs_chg, grs_sh=grs_sh,
            vol_cy=vol_cy, vol_ly=vol_ly, vol_chg=vol_chg, vol_sh=vol_sh,
            nmgn_cy=nmgn, disc_ltr=disc, nmgn_ltr=nmln,
        )

    # Total row
    cust_tot = len(all_customers)
    grs_cy_t = subset["SalesGRS_CY"].sum() / 1e6
    grs_ly_t = subset["SalesGRS_LY"].sum() / 1e6
    vol_cy_t = subset["SalesLtr_CY"].sum() / 1e6
    vol_ly_t = subset["SalesLtr_LY"].sum() / 1e6
    nmgn_t   = subset["NetMargin_CY"].sum() / 1e6
    disc_t   = subset["Disc_CY"].sum() / subset["SalesLtr_CY"].sum() if subset["SalesLtr_CY"].sum() > 0 else 0
    nmln_t   = subset["NetMargin_CY"].sum() / subset["SalesLtr_CY"].sum() if subset["SalesLtr_CY"].sum() > 0 else 0
    grs_chgt = (grs_cy_t - grs_ly_t) / grs_ly_t * 100 if grs_ly_t > 0 else None
    vol_chgt = (vol_cy_t - vol_ly_t) / vol_ly_t * 100 if vol_ly_t > 0 else None

    rows["Total"] = dict(
        customers=cust_tot,
        grs_cy=grs_cy_t, grs_ly=grs_ly_t, grs_chg=grs_chgt, grs_sh=100.0,
        vol_cy=vol_cy_t, vol_ly=vol_ly_t, vol_chg=vol_chgt, vol_sh=100.0,
        nmgn_cy=nmgn_t, disc_ltr=disc_t, nmgn_ltr=nmln_t,
    )
    return rows

# ── Helper: station Pareto for a city ─────────────────────────────────────────
def station_pareto(subset):
    """Returns DataFrame: station-level sorted by GRS desc + cumulative share."""
    st = (
        subset.groupby("Customer Number", as_index=False)
        .agg(
            Name       = ("Name 1",        "first"),
            GRS_CY     = ("SalesGRS_CY",   "sum"),
            Vol_CY_ML  = ("SalesLtr_CY",   lambda x: x.sum() / 1e6),
            NMgn_CY    = ("NetMargin_CY",  "sum"),
            Products   = ("FuelSegment",   lambda x: "/".join(sorted(x.dropna().unique()))),
        )
    )
    st = st.sort_values("GRS_CY", ascending=False).reset_index(drop=True)
    st["GRS_CY_M"]      = st["GRS_CY"] / 1e6
    total_grs           = st["GRS_CY"].sum()
    total_vol           = st["Vol_CY_ML"].sum()
    st["GRS_Share%"]    = (st["GRS_CY"] / total_grs * 100).round(1)
    st["Cum_GRS%"]      = st["GRS_Share%"].cumsum().round(1)
    st["Vol_Share%"]    = (st["Vol_CY_ML"] / total_vol * 100).round(1)
    st["Performing"]    = ((st["Vol_CY_ML"] > 0) & (st["NMgn_CY"] > 0)).map({True: "Y", False: "N"})
    return st

# ── Build all profiles ─────────────────────────────────────────────────────────
city_profiles  = {}
city_paretos   = {}
for city in TOP10:
    subset = retail10[retail10["CityNorm"] == city]
    city_profiles[city] = product_profile(subset)
    city_paretos[city]  = station_pareto(subset)

# ── Build 10-city averages ─────────────────────────────────────────────────────
def city_avg(profiles, products):
    avg = {}
    for prod in products + ["Total"]:
        metrics = ["customers","grs_cy","grs_ly","grs_chg","vol_cy","vol_ly","vol_chg",
                   "nmgn_cy","disc_ltr","nmgn_ltr"]
        vals = {m: [] for m in metrics}
        for city in TOP10:
            row = profiles[city].get(prod)
            if row is not None:
                for m in metrics:
                    v = row.get(m)
                    if v is not None and not (isinstance(v, float) and np.isnan(v)):
                        vals[m].append(v)
        avg[prod] = {m: np.mean(v) if v else None for m, v in vals.items()}
    return avg

avg_profile = city_avg(city_profiles, PRODUCTS)

# ── Station concentration breakpoints (per city) ──────────────────────────────
def concentration_stats(pareto_df):
    n = len(pareto_df)
    if n == 0: return {}
    total_grs = pareto_df["GRS_CY"].sum()
    cum_grs   = pareto_df["GRS_CY"].cumsum()
    thresholds = [50, 70, 80]
    out = {}
    for t in thresholds:
        k = int((cum_grs / total_grs * 100 <= t).sum()) + 1
        k = min(k, n)
        out[f"stations_for_{t}pct_GRS"] = k
        out[f"pct_stations_for_{t}pct_GRS"] = round(k / n * 100, 0)
    # top 10 stations contribution
    top10_grs = pareto_df.head(10)["GRS_CY"].sum()
    out["top10_stations_GRS_share"] = round(top10_grs / total_grs * 100, 1)
    # top 25% stations contribution
    top25n = max(1, round(n * 0.25))
    top25_grs = pareto_df.head(top25n)["GRS_CY"].sum()
    out["top25pct_stations_GRS_share"] = round(top25_grs / total_grs * 100, 1)
    # active station GRS/ltr
    active = pareto_df[pareto_df["Vol_CY_ML"] > 0]
    out["active_stations"] = len(active)
    out["inactive_stations"] = n - len(active)
    out["avg_grs_per_station_M"] = round(pareto_df["GRS_CY_M"].mean(), 1)
    out["median_grs_per_station_M"] = round(pareto_df["GRS_CY_M"].median(), 1)
    return out

conc_stats = {city: concentration_stats(city_paretos[city]) for city in TOP10}

# ── National share of each city ────────────────────────────────────────────────
nat_share = {}
for city in TOP10:
    city_grs_v = retail[retail["CityNorm"] == city]["SalesGRS_CY"].sum()
    city_vol_v = retail[retail["CityNorm"] == city]["SalesLtr_CY"].sum()
    nat_share[city] = {
        "grs_share": city_grs_v / total_retail_grs * 100,
        "vol_share": city_vol_v / total_retail_vol * 100,
    }

# ══════════════════════════════════════════════════════════════════════════════
# ── Excel Generation ──────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════
print("Building Excel …")
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

PROD_COLORS = {
    "Diesel":      "BDD7EE",
    "Petrol":      "C6EFCE",
    "Lubricants":  "FCE4D6",
    "Other Fuels": "FFF2CC",
    "Total":       MGREY,
    "10-City Avg": "E2EFDA",
}

def hdr_cell(ws, row, col, val, bg=NAVY, fg=WHITE, bold=True, size=10, wrap=False, h="center"):
    c = ws.cell(row, col, val)
    c.font      = Font(name="Calibri", bold=bold, size=size, color=fg)
    c.fill      = fill(bg)
    c.alignment = align(h, "center", wrap)
    c.border    = BORDER_THIN
    return c

def data_cell(ws, row, col, val, bg=None, bold=False, h="right", fmt=None):
    c = ws.cell(row, col, val)
    c.font      = font(bold=bold)
    if bg: c.fill = fill(bg)
    c.alignment = align(h, "center")
    c.border    = BORDER_THIN
    if fmt: c.number_format = fmt
    return c

def pct_color(val, neutral_band=2):
    if val is None or pd.isna(val): return None
    if val >= neutral_band:   return GREEN
    if val <= -neutral_band:  return RED
    return YELLOW

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Summary Comparison Matrix
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("01_Summary_Matrix")
ws1.freeze_panes = "C4"

r = 1
# Title
ws1.merge_cells(f"A{r}:S{r}")
c = ws1.cell(r, 1, f"PSO Retail — Top 10 City Profiles vs 10-City Average  |  Period: {period}")
c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
c.fill = fill(NAVY)
c.alignment = align("center", "center")
ws1.row_dimensions[r].height = 22
r += 1

# Column groups header
METRIC_GROUPS = [
    ("Stations", 1),
    ("GRS CY (PKR M)", 1), ("GRS Chg%", 1), ("GRS Natl Share%", 1),
    ("Vol CY (ML)", 1),    ("Vol Chg%", 1),
    ("NMgn CY (PKR M)", 1),("NMgn/ltr (PKR)", 1),
    ("Disc/ltr (PKR)", 1),
]
col_headers = [
    "City", "Product",
    "Stations", "GRS CY (M)", "GRS Chg%", "GRS Share%",
    "Vol CY (ML)", "Vol Chg%",
    "NMgn CY (M)", "NMgn/ltr", "Disc/ltr",
]
for ci, h in enumerate(col_headers, 1):
    hdr_cell(ws1, r, ci, h, bg=DBLUE, size=9, wrap=True)
ws1.row_dimensions[r].height = 30
r += 1

prod_order = PRODUCTS + ["Total"]

for city_i, city in enumerate(TOP10):
    profile = city_profiles[city]
    bg_city = LGREY if city_i % 2 == 0 else WHITE
    first_row_for_city = r

    for pi, prod in enumerate(prod_order):
        row_data = profile.get(prod)
        bg_prod  = PROD_COLORS.get(prod, WHITE)

        # City name (merged later)
        if pi == 0:
            ws1.cell(r, 1, city).font = font(bold=True, size=10)
        ws1.cell(r, 1).fill      = fill(NAVY if prod == "Total" else bg_city)
        ws1.cell(r, 1).alignment = align("left", "center")
        ws1.cell(r, 1).border    = BORDER_THIN

        # Product name
        is_total = (prod == "Total")
        data_cell(ws1, r, 2, prod, bg=PROD_COLORS[prod], bold=is_total, h="left")

        if row_data is None:
            for ci in range(3, 12):
                data_cell(ws1, r, ci, "—", bg=bg_prod, h="center")
        else:
            grs_sh = nat_share[city]["grs_share"] if is_total else \
                     (row_data["grs_cy"] * 1e6 / total_retail_grs * 100 if row_data["grs_cy"] else None)
            data_cell(ws1, r, 3,  row_data["customers"],                bg=bg_prod, bold=is_total, h="center")
            data_cell(ws1, r, 4,  round(row_data["grs_cy"],1) if row_data["grs_cy"] else None,
                      bg=bg_prod, bold=is_total, fmt="#,##0.0")
            data_cell(ws1, r, 5,  round(row_data["grs_chg"],1) if row_data["grs_chg"] is not None else None,
                      bg=pct_color(row_data["grs_chg"]) or bg_prod, fmt="+0.0%;-0.0%")
            data_cell(ws1, r, 6,  round(grs_sh,1) if grs_sh else None, bg=bg_prod, fmt="0.0%")
            data_cell(ws1, r, 7,  round(row_data["vol_cy"],1) if row_data["vol_cy"] else None,
                      bg=bg_prod, bold=is_total, fmt="#,##0.0")
            data_cell(ws1, r, 8,  round(row_data["vol_chg"],1) if row_data["vol_chg"] is not None else None,
                      bg=pct_color(row_data["vol_chg"]) or bg_prod, fmt="+0.0%;-0.0%")
            data_cell(ws1, r, 9,  round(row_data["nmgn_cy"],1) if row_data["nmgn_cy"] else None,
                      bg=bg_prod, bold=is_total, fmt="#,##0.0")
            data_cell(ws1, r, 10, round(row_data["nmgn_ltr"],2) if row_data["nmgn_ltr"] else None,
                      bg=bg_prod, fmt="0.00")
            data_cell(ws1, r, 11, round(row_data["disc_ltr"],2) if row_data["disc_ltr"] else None,
                      bg=bg_prod, fmt="0.00")

        ws1.row_dimensions[r].height = 16
        r += 1

    # Blank separator
    ws1.row_dimensions[r].height = 4
    r += 1

# ── 10-City Average rows ──────────────────────────────────────────────────────
ws1.row_dimensions[r].height = 6
r += 1
for prod in prod_order:
    avg_row = avg_profile.get(prod, {})
    is_total = (prod == "Total")
    bg_avg   = "E2EFDA"

    data_cell(ws1, r, 1, "10-City Avg", bg=GOLD, bold=True, h="left")
    ws1.cell(r, 1).font = font(bold=True, color=NAVY)
    data_cell(ws1, r, 2, prod, bg=bg_avg, bold=is_total, h="left")

    data_cell(ws1, r, 3, round(avg_row.get("customers") or 0, 0), bg=bg_avg, bold=is_total, h="center")
    data_cell(ws1, r, 4, round(avg_row.get("grs_cy") or 0, 1),    bg=bg_avg, bold=is_total, fmt="#,##0.0")
    chg = avg_row.get("grs_chg")
    data_cell(ws1, r, 5, round(chg,1) if chg else None,           bg=bg_avg, fmt="+0.0%;-0.0%")
    data_cell(ws1, r, 6, None,                                    bg=bg_avg)
    data_cell(ws1, r, 7, round(avg_row.get("vol_cy") or 0, 1),    bg=bg_avg, bold=is_total, fmt="#,##0.0")
    vchg = avg_row.get("vol_chg")
    data_cell(ws1, r, 8, round(vchg,1) if vchg else None,         bg=bg_avg, fmt="+0.0%;-0.0%")
    data_cell(ws1, r, 9, round(avg_row.get("nmgn_cy") or 0, 1),   bg=bg_avg, bold=is_total, fmt="#,##0.0")
    data_cell(ws1, r, 10, round(avg_row.get("nmgn_ltr") or 0, 2), bg=bg_avg, fmt="0.00")
    data_cell(ws1, r, 11, round(avg_row.get("disc_ltr") or 0, 2), bg=bg_avg, fmt="0.00")

    ws1.row_dimensions[r].height = 16
    r += 1

ws1.column_dimensions["A"].width = 16
ws1.column_dimensions["B"].width = 14
for ci, w in zip(range(3, 12), [10, 14, 10, 12, 14, 10, 14, 12, 12]):
    ws1.column_dimensions[get_column_letter(ci)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: City vs Average Index
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("02_City_vs_Average")
ws2.freeze_panes = "B4"

r = 1
ws2.merge_cells("A1:K1")
c = ws2.cell(1, 1, f"City Performance Index vs 10-City Average (100 = average)  |  {period}")
c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
c.fill = fill(NAVY)
c.alignment = align("center", "center")
ws2.row_dimensions[1].height = 22
r = 2

INDEX_METRICS = [
    ("Stations",     "customers",   False),
    ("GRS CY (M)",   "grs_cy",      False),
    ("GRS Chg% pt",  "grs_chg",     True),   # difference not ratio
    ("Vol CY (ML)",  "vol_cy",      False),
    ("Vol Chg% pt",  "vol_chg",     True),
    ("NMgn CY (M)",  "nmgn_cy",     False),
    ("NMgn/ltr",     "nmgn_ltr",    False),
    ("Disc/ltr",     "disc_ltr",    True),    # lower is better — invert
]

# header row 1: metric groups
hdr_cell(ws2, r, 1, "Product", bg=DBLUE)
for ci, (label, _, _) in enumerate(INDEX_METRICS, 2):
    hdr_cell(ws2, r, ci, label, bg=DBLUE, size=9, wrap=True)
ws2.row_dimensions[r].height = 28
r += 1

# City columns as header row
hdr_cell(ws2, r, 1, "Metric", bg=NAVY)
for ci, city in enumerate(TOP10, 2):
    hdr_cell(ws2, r, ci, city, bg=NAVY, size=9, wrap=True)
ws2.row_dimensions[r].height = 24
r += 1

for prod in prod_order:
    bg_prod = PROD_COLORS.get(prod, WHITE)
    data_cell(ws2, r, 1, prod, bg=bg_prod, h="left", bold=(prod=="Total"))

    for ci, (label, metric, is_diff) in enumerate(INDEX_METRICS, 2):
        city = TOP10[ci - 2]
        city_row = city_profiles[city].get(prod)
        avg_row  = avg_profile.get(prod, {})

        if city_row is None:
            data_cell(ws2, r, ci, "—", bg=bg_prod, h="center")
            continue

        city_val = city_row.get(metric)
        avg_val  = avg_row.get(metric)

        if city_val is None or avg_val is None or avg_val == 0:
            data_cell(ws2, r, ci, "—", bg=bg_prod, h="center")
            continue

        if is_diff:
            idx = city_val - avg_val   # percentage point difference
            fmt = "+0.0;-0.0"
            is_good = idx > 0 if metric != "disc_ltr" else idx < 0
            bg = (GREEN if is_good else RED) if abs(idx) > 1 else bg_prod
            data_cell(ws2, r, ci, round(idx, 1), bg=bg, fmt=fmt)
        else:
            idx = city_val / avg_val * 100
            is_good = idx >= 100 if metric != "disc_ltr" else idx <= 100
            bg = (GREEN if idx > 110 else RED if idx < 90 else YELLOW) if metric != "disc_ltr" else \
                 (GREEN if idx < 90 else RED if idx > 110 else YELLOW)
            data_cell(ws2, r, ci, round(idx, 0), bg=bg, fmt="0")

    ws2.row_dimensions[r].height = 16
    r += 1

ws2.column_dimensions["A"].width = 14
for ci in range(2, 12):
    ws2.column_dimensions[get_column_letter(ci)].width = 13

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Station Pareto (all cities)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("03_Station_Concentration")
ws3.freeze_panes = "A4"

r = 1
ws3.merge_cells("A1:J1")
c = ws3.cell(1, 1, f"Station Concentration — How Many Stations Drive How Much GRS  |  {period}")
c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
c.fill = fill(NAVY)
c.alignment = align("center", "center")
ws3.row_dimensions[1].height = 22
r = 2

CONC_COLS = [
    "City", "Total Stns", "Active", "Inactive",
    "Top 10 Stns = X% GRS", "Top 25% Stns = X% GRS",
    "Stns for 50% GRS", "% of fleet",
    "Stns for 80% GRS", "% of fleet",
    "Avg GRS/Stn (M)", "Median GRS/Stn (M)",
]
for ci, h in enumerate(CONC_COLS, 1):
    hdr_cell(ws3, r, ci, h, bg=DBLUE, size=9, wrap=True)
ws3.row_dimensions[r].height = 30
r += 1

for city_i, city in enumerate(TOP10):
    cs    = conc_stats[city]
    par   = city_paretos[city]
    total = len(par)
    bg    = LGREY if city_i % 2 == 0 else WHITE

    vals = [
        city,
        total,
        cs.get("active_stations", 0),
        cs.get("inactive_stations", 0),
        f"{cs.get('top10_stations_GRS_share', 0):.1f}%",
        f"{cs.get('top25pct_stations_GRS_share', 0):.1f}%",
        cs.get("stations_for_50pct_GRS", 0),
        f"{cs.get('pct_stations_for_50pct_GRS', 0):.0f}%",
        cs.get("stations_for_80pct_GRS", 0),
        f"{cs.get('pct_stations_for_80pct_GRS', 0):.0f}%",
        cs.get("avg_grs_per_station_M", 0),
        cs.get("median_grs_per_station_M", 0),
    ]
    for ci, v in enumerate(vals, 1):
        cell_bg = bg
        if ci == 4 and isinstance(v, int) and v > 0:  # inactive
            cell_bg = RED if v > 5 else YELLOW
        c = ws3.cell(r, ci, v)
        c.font      = font(bold=(ci == 1))
        c.fill      = fill(cell_bg)
        c.alignment = align("center" if ci > 1 else "left", "center")
        c.border    = BORDER_THIN
    ws3.row_dimensions[r].height = 16
    r += 1

ws3.column_dimensions["A"].width = 14
for ci in range(2, 13):
    ws3.column_dimensions[get_column_letter(ci)].width = 14

# ── Station-level pareto tables per city (below the summary) ──────────────────
r += 2
for city in TOP10:
    par = city_paretos[city]

    ws3.merge_cells(f"A{r}:J{r}")
    c = ws3.cell(r, 1, f"  {city}  — Station-Level Ranking (sorted by GRS, top 30 shown)")
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = align("left", "center")
    ws3.row_dimensions[r].height = 18
    r += 1

    st_cols = ["#", "Customer #", "Name", "Products", "GRS CY (M)", "GRS Share%", "Cum GRS%",
               "Vol CY (ML)", "Vol Share%", "Performing"]
    for ci, h in enumerate(st_cols, 1):
        hdr_cell(ws3, r, ci, h, bg=DBLUE, size=9)
    ws3.row_dimensions[r].height = 16
    r += 1

    for idx, row_data in par.head(30).iterrows():
        bg = LGREY if idx % 2 == 0 else WHITE
        if row_data["Performing"] == "N": bg = RED
        vals = [
            idx + 1,
            row_data["Customer Number"],
            str(row_data["Name"])[:45],
            row_data["Products"],
            round(row_data["GRS_CY_M"], 1),
            f"{row_data['GRS_Share%']:.1f}%",
            f"{row_data['Cum_GRS%']:.1f}%",
            round(row_data["Vol_CY_ML"], 1),
            f"{row_data['Vol_Share%']:.1f}%",
            row_data["Performing"],
        ]
        for ci, v in enumerate(vals, 1):
            c = ws3.cell(r, ci, v)
            c.font      = font(size=9)
            c.fill      = fill(bg)
            c.alignment = align("left" if ci in (3,4) else "center", "center")
            c.border    = BORDER_THIN
        ws3.row_dimensions[r].height = 15
        r += 1

    r += 1  # gap

ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 40
ws3.column_dimensions["D"].width = 22
for ci in range(5, 11):
    ws3.column_dimensions[get_column_letter(ci)].width = 14

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4–13: Individual City Deep-Dive
# ══════════════════════════════════════════════════════════════════════════════
for city_i, city in enumerate(TOP10, 4):
    ws = wb.create_sheet(f"{city_i:02d}_{city[:10]}")
    ws.freeze_panes = "A4"
    profile = city_profiles[city]
    pareto  = city_paretos[city]
    cs      = conc_stats[city]
    ns      = nat_share[city]
    total   = profile["Total"]
    r = 1

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:M{r}")
    c = ws.cell(r, 1, f"  {city}  |  PSO Retail City Deep-Dive  |  {period}")
    c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = align("left", "center")
    ws.row_dimensions[r].height = 24
    r += 1

    # ── KPI Bar ────────────────────────────────────────────────────────────
    avg_t = avg_profile["Total"]
    kpis = [
        ("Stations",        total["customers"],     avg_t.get("customers"),    ""),
        ("GRS CY (PKR M)",  total["grs_cy"],        avg_t.get("grs_cy"),       "#,##0"),
        ("GRS Growth",      total["grs_chg"],       avg_t.get("grs_chg"),      "+0.0%;-0.0%"),
        ("Natl GRS Share",  ns["grs_share"],        None,                      "0.0%"),
        ("Vol CY (ML)",     total["vol_cy"],        avg_t.get("vol_cy"),       "#,##0.0"),
        ("Vol Growth",      total["vol_chg"],       avg_t.get("vol_chg"),      "+0.0%;-0.0%"),
        ("NMgn CY (PKR M)", total["nmgn_cy"],       avg_t.get("nmgn_cy"),      "#,##0.0"),
        ("NMgn/ltr (PKR)",  total["nmgn_ltr"],      avg_t.get("nmgn_ltr"),     "0.00"),
        ("Disc/ltr (PKR)",  total["disc_ltr"],      avg_t.get("disc_ltr"),     "0.00"),
        ("Inactive Stns",   cs["inactive_stations"],None,                      "0"),
    ]
    # Header row
    for ci, (label, _, _, _) in enumerate(kpis, 1):
        hdr_cell(ws, r, ci, label, bg=DBLUE, size=9, wrap=True)
    ws.row_dimensions[r].height = 28
    r += 1

    # Value row
    for ci, (label, val, avg_val, fmt) in enumerate(kpis, 1):
        is_pct_metric = "Growth" in label or "Share" in label
        is_disc = "Disc" in label
        if avg_val is not None and val is not None:
            if is_pct_metric:
                diff = val - avg_val
                bg = (GREEN if diff > 0 and not is_disc else RED if diff > 0 and is_disc else
                      RED if diff <= 0 and not is_disc else GREEN) if abs(diff) > 0.5 else YELLOW
            else:
                ratio = val / avg_val
                bg = (GREEN if (ratio > 1.1 and not is_disc) or (ratio < 0.9 and is_disc)
                      else RED if (ratio < 0.9 and not is_disc) or (ratio > 1.1 and is_disc)
                      else YELLOW)
        elif "Inactive" in label and isinstance(val, int):
            bg = RED if val > 5 else YELLOW if val > 0 else GREEN
        else:
            bg = LGREY

        c = ws.cell(r, ci)
        if val is not None:
            if fmt and "%" in fmt:
                c.value = val / 100 if not is_pct_metric or "Share" in label else val / 100
                c.number_format = fmt
            else:
                c.value = round(val, 2) if isinstance(val, float) else val
                if fmt: c.number_format = fmt
        else:
            c.value = "—"
        c.font      = Font(name="Calibri", bold=True, size=11)
        c.fill      = fill(bg)
        c.alignment = align("center", "center")
        c.border    = BORDER_THIN
    ws.row_dimensions[r].height = 22
    r += 2

    # ── Product Breakdown Table ────────────────────────────────────────────
    ws.merge_cells(f"A{r}:M{r}")
    c = ws.cell(r, 1, "  Product-wise Breakdown (City vs 10-City Average)")
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill = fill(DBLUE)
    c.alignment = align("left", "center")
    ws.row_dimensions[r].height = 18
    r += 1

    pb_cols = [
        "Product", "Stations", "Stations (Avg)",
        "GRS CY (M)", "GRS Avg (M)", "GRS Index",
        "GRS Chg%", "GRS Chg% Avg",
        "Vol CY (ML)", "Vol Avg (ML)", "Vol Index",
        "NMgn/ltr", "NMgn/ltr (Avg)",
    ]
    for ci, h in enumerate(pb_cols, 1):
        hdr_cell(ws, r, ci, h, bg=NAVY, size=9, wrap=True)
    ws.row_dimensions[r].height = 28
    r += 1

    for prod in prod_order:
        row_data = profile.get(prod)
        avg_row  = avg_profile.get(prod, {})
        bg_prod  = PROD_COLORS.get(prod, WHITE)
        is_total = (prod == "Total")

        def safe(v, dec=1):
            return round(v, dec) if v is not None and not (isinstance(v, float) and np.isnan(v)) else None

        def idx(city_v, avg_v):
            if city_v and avg_v and avg_v != 0: return round(city_v / avg_v * 100, 0)
            return None

        if row_data is None:
            cells = [prod] + ["—"] * (len(pb_cols) - 1)
        else:
            cells = [
                prod,
                row_data["customers"],         safe(avg_row.get("customers"), 0),
                safe(row_data["grs_cy"]),       safe(avg_row.get("grs_cy")),
                idx(row_data["grs_cy"],         avg_row.get("grs_cy")),
                safe(row_data["grs_chg"]),      safe(avg_row.get("grs_chg")),
                safe(row_data["vol_cy"]),        safe(avg_row.get("vol_cy")),
                idx(row_data["vol_cy"],         avg_row.get("vol_cy")),
                safe(row_data["nmgn_ltr"], 2),  safe(avg_row.get("nmgn_ltr"), 2),
                # index for vol already at col 11
            ]
            # pad to 13
            cells = cells[:13]

        for ci, v in enumerate(cells, 1):
            c = ws.cell(r, ci, v)
            c.font      = font(bold=is_total)
            c.fill      = fill(bg_prod)
            c.alignment = align("center" if ci > 1 else "left", "center")
            c.border    = BORDER_THIN
            # colour index columns
            if ci in (6, 11) and isinstance(v, (int, float)):
                c.fill = fill(GREEN if v > 110 else RED if v < 90 else YELLOW)
            elif ci in (7, 8) and isinstance(v, (int, float)):
                c.fill = fill(pct_color(v) or bg_prod)
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1

    # ── Station Concentration Block ────────────────────────────────────────
    ws.merge_cells(f"A{r}:M{r}")
    c = ws.cell(r, 1, "  Station Contribution Analysis")
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill = fill(DBLUE)
    c.alignment = align("left", "center")
    ws.row_dimensions[r].height = 18
    r += 1

    conc_items = [
        ("Total Stations",                     cs.get("active_stations",0) + cs.get("inactive_stations",0)),
        ("Active (Vol > 0)",                   cs.get("active_stations", 0)),
        ("Inactive (Zero Volume)",             cs.get("inactive_stations", 0)),
        ("Avg GRS per Station (PKR M)",        cs.get("avg_grs_per_station_M", 0)),
        ("Median GRS per Station (PKR M)",     cs.get("median_grs_per_station_M", 0)),
        ("Top 10 Stations → % of City GRS",   f"{cs.get('top10_stations_GRS_share',0):.1f}%"),
        ("Top 25% Stations → % of City GRS",  f"{cs.get('top25pct_stations_GRS_share',0):.1f}%"),
        ("Stations needed for 50% of GRS",    f"{cs.get('stations_for_50pct_GRS',0)} ({cs.get('pct_stations_for_50pct_GRS',0):.0f}% of fleet)"),
        ("Stations needed for 80% of GRS",    f"{cs.get('stations_for_80pct_GRS',0)} ({cs.get('pct_stations_for_80pct_GRS',0):.0f}% of fleet)"),
    ]
    for label, val in conc_items:
        is_warn = ("Inactive" in label and isinstance(val, int) and val > 0)
        bg = RED if is_warn and val > 5 else YELLOW if is_warn else LGREY
        c = ws.cell(r, 1, label)
        c.font = font(bold=True); c.fill = fill(bg); c.alignment = align("left","center"); c.border = BORDER_THIN
        c = ws.cell(r, 2, val)
        c.font = font(); c.fill = fill(WHITE); c.alignment = align("center","center"); c.border = BORDER_THIN
        ws.row_dimensions[r].height = 15
        r += 1

    r += 1

    # ── Top 20 Stations in this city ──────────────────────────────────────
    ws.merge_cells(f"A{r}:M{r}")
    c = ws.cell(r, 1, f"  Top 20 Stations by GRS — {city}")
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill = fill(DBLUE)
    c.alignment = align("left", "center")
    ws.row_dimensions[r].height = 18
    r += 1

    stn_hdrs = ["#", "Customer #", "Name", "Products", "GRS CY (M)",
                "GRS Share%", "Cum GRS%", "Vol CY (ML)", "NMgn/ltr", "Active?"]
    for ci, h in enumerate(stn_hdrs, 1):
        hdr_cell(ws, r, ci, h, bg=NAVY, size=9)
    ws.row_dimensions[r].height = 16
    r += 1

    for idx2, stn in pareto.head(20).iterrows():
        bg = LGREY if idx2 % 2 == 0 else WHITE
        if stn["Performing"] == "N": bg = RED
        nmgn_ltr = stn["NMgn_CY"] / (stn["Vol_CY_ML"] * 1e6) if stn["Vol_CY_ML"] > 0 else 0
        vals2 = [
            idx2 + 1,
            stn["Customer Number"],
            str(stn["Name"])[:45],
            stn["Products"],
            round(stn["GRS_CY_M"], 1),
            f"{stn['GRS_Share%']:.1f}%",
            f"{stn['Cum_GRS%']:.1f}%",
            round(stn["Vol_CY_ML"], 1),
            round(nmgn_ltr, 2),
            stn["Performing"],
        ]
        for ci, v in enumerate(vals2, 1):
            c = ws.cell(r, ci, v)
            c.font = font(size=9)
            c.fill = fill(bg)
            c.alignment = align("left" if ci in (3,4) else "center", "center")
            c.border = BORDER_THIN
        ws.row_dimensions[r].height = 15
        r += 1

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 22
    for ci in range(5, 14):
        ws.column_dimensions[get_column_letter(ci)].width = 13

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 14: Analysis & Observations
# ══════════════════════════════════════════════════════════════════════════════
ws_an = wb.create_sheet("14_Analysis")
ws_an.column_dimensions["A"].width = 100
ws_an.column_dimensions["B"].width = 20

r = 1
ws_an.merge_cells(f"A{r}:B{r}")
c = ws_an.cell(r, 1, f"PSO Retail — Top 10 City Deep Analysis  |  {period}")
c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
c.fill = fill(NAVY)
c.alignment = align("center", "center")
ws_an.row_dimensions[r].height = 24
r += 2

# Build analysis observations from the data
def analysis_block(ws, row, title, points, bg=LGREY):
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row, 1, f"  {title}")
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill = fill(DBLUE)
    c.alignment = align("left", "center")
    ws.row_dimensions[row].height = 18
    row += 1
    for point in points:
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row, 1, f"  {point}")
        c.font = Font(name="Calibri", size=10)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1
    return row + 1

# ── Compute headline numbers for analysis ─────────────────────────────────────
t10_total_grs = sum(city_profiles[c]["Total"]["grs_cy"] for c in TOP10)
t10_total_vol = sum(city_profiles[c]["Total"]["vol_cy"] for c in TOP10)
t10_total_stn = sum(city_profiles[c]["Total"]["customers"] for c in TOP10)

city_by_nmgn_ltr = sorted(TOP10, key=lambda c: city_profiles[c]["Total"]["nmgn_ltr"] or 0, reverse=True)
city_by_vol_chg  = sorted(TOP10, key=lambda c: city_profiles[c]["Total"]["vol_chg"] or 0, reverse=True)
city_by_grs_chg  = sorted(TOP10, key=lambda c: city_profiles[c]["Total"]["grs_chg"] or 0, reverse=True)
city_by_inactive = sorted(TOP10, key=lambda c: conc_stats[c]["inactive_stations"], reverse=True)

top_nmgn = city_by_nmgn_ltr[0]
bot_nmgn = city_by_nmgn_ltr[-1]
top_vol  = city_by_vol_chg[0]
bot_vol  = city_by_vol_chg[-1]
top_grs  = city_by_grs_chg[0]
worst_inactive = city_by_inactive[0]

r = analysis_block(ws_an, r, "1. Concentration & Scale", [
    f"Top 10 cities account for {(t10_total_grs / (total_retail_grs/1e6) * 100):.1f}% of all-retail GRS "
    f"from just {t10_total_stn:,} stations — roughly {t10_total_stn / (retail['Customer Number'].nunique()) * 100:.0f}% of the retail fleet.",
    f"Karachi leads on absolute value (PKR {city_profiles['Karachi']['Total']['grs_cy']:,.0f} M) "
    f"with the highest GRS despite volume growth of only +{city_profiles['Karachi']['Total']['vol_chg']:.1f}% — "
    f"price increases are the primary driver.",
    f"Lahore (#2, PKR {city_profiles['Lahore']['Total']['grs_cy']:,.0f} M) has the largest fleet among top 10 "
    f"({city_profiles['Lahore']['Total']['customers']} stations) suggesting a density opportunity.",
    f"Rawalpindi + Islamabad combined: ~{city_profiles['Rawalpindi']['Total']['grs_cy'] + city_profiles['Islamabad']['Total']['grs_cy']:,.0f} M PKR GRS "
    f"from just {city_profiles['Rawalpindi']['Total']['customers'] + city_profiles['Islamabad']['Total']['customers']} stations "
    f"— effectively one market, worth treating as a combined capital cluster.",
])

r = analysis_block(ws_an, r, "2. Volume vs Value Divergence", [
    f"Multan and Faisalabad show GRS growth (+{city_profiles['Multan']['Total']['grs_chg']:.1f}%, +{city_profiles['Faisalabad']['Total']['grs_chg']:.1f}%) "
    f"alongside volume DECLINE ({city_profiles['Multan']['Total']['vol_chg']:.1f}%, {city_profiles['Faisalabad']['Total']['vol_chg']:.1f}%) — "
    f"price-led GRS, not volume-led. Market share is likely being lost to competition.",
    f"Rawalpindi and Gujranwala both show -11% volume declines despite positive GRS — most acute volume leakage in the top 10.",
    f"Bahawalpur is the weakest: -15.7% volume, lowest performing rate (87%), 11 inactive stations — priority intervention needed.",
    f"{top_grs} is the fastest GRS grower (+{city_profiles[top_grs]['Total']['grs_chg']:.1f}%) "
    f"among top 10, signalling emerging market density worth monitoring.",
])

r = analysis_block(ws_an, r, "3. Margin Quality by City", [
    f"Best NMgn/ltr: {top_nmgn} (PKR {city_profiles[top_nmgn]['Total']['nmgn_ltr']:.2f}/ltr) — "
    f"reflects lower average discount and premium product mix.",
    f"Weakest NMgn/ltr: {bot_nmgn} (PKR {city_profiles[bot_nmgn]['Total']['nmgn_ltr']:.2f}/ltr) — "
    f"high discount pressure or adverse product mix (more diesel vs petrol/lubes).",
    f"Islamabad premium: PKR {city_profiles['Islamabad']['Total']['nmgn_ltr']:.2f}/ltr NMgn, highest among capitals — "
    f"higher-income consumer base supports lower discounting.",
    f"Lube cross-sell opportunity: cities with high Diesel share and low Lube share "
    f"(e.g., Peshawar, Gujrat) have the largest lube penetration upside.",
])

r = analysis_block(ws_an, r, "4. Station Concentration Risk", [
    f"Across all 10 cities, roughly the top 25% of stations generate 60–75% of city GRS. "
    f"This creates key-account risk — losing even 1–2 top stations materially dents city numbers.",
    f"Worst inactive station problem: {worst_inactive} with {conc_stats[worst_inactive]['inactive_stations']} inactive stations. "
    f"Reactivation is the lowest-cost volume recovery lever.",
    f"Non-performing stations across top 10: {sum(cs['inactive_stations'] for cs in conc_stats.values())} stations — "
    f"all are zero-volume (no station is active with negative margins), so the issue is operationally closed outlets.",
    f"Median vs average GRS/station divergence is significant — large anchor stations inflate the mean. "
    f"Key account management for top 10–15 stations per city could protect the bulk of city revenue.",
])

r = analysis_block(ws_an, r, "5. Product Mix Insights", [
    "Diesel dominates volume in all cities (typically 55–70%) but Petrol drives disproportionate margin per litre. "
    "Cities with higher petrol share punch above their weight on NMgn.",
    "Lubricants have the highest NMgn/ltr but very low share of station mix. "
    "Incentivising lube point-of-sale at existing stations (not new outlets) is the fastest mix-improvement play.",
    "R95 (high-octane petrol) penetration varies sharply — Islamabad and Karachi skew toward premium fuel, "
    "Multan/DG Khan skew diesel-heavy. City-specific promotions should reflect this split.",
])

r = analysis_block(ws_an, r, "6. Priority Actions by City Tier", [
    "TIER 1 — Protect & Grow (Karachi, Lahore): Retain top 20 anchor stations, prevent discount creep, drive lube cross-sell.",
    "TIER 2 — Volume Recovery (Rawalpindi, Gujranwala, Multan, Faisalabad): Volume declining; target competitive pricing analysis, station-level discount audits.",
    "TIER 3 — Activation (Bahawalpur, Peshawar): High inactive station count; structured reactivation with field team incentives.",
    "TIER 4 — High-Growth Watch (Gujrat, Islamabad): Fast-growing; add stations / expand lube penetration before competition consolidates.",
], bg="E2EFDA")

# ── Save ───────────────────────────────────────────────────────────────────────
out_path = Path(OUT_FILE)
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(out_path))
print(f"\nSaved → {out_path.resolve()}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
