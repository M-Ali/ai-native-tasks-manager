"""CLI entry point — orchestrates all agents."""

from __future__ import annotations

import os
import sys
from pathlib import Path

import click
from dotenv import load_dotenv
from rich.console import Console
from rich.panel import Panel
from rich.table import Table as RichTable

load_dotenv()
console = Console()


@click.group()
def cli():
    """PSO OMC Analytics Pipeline — powered by uv + Claude AI."""
    pass


@cli.command()
@click.option("--input", "-i", "input_path",
              default="data/input/Working File Retail Fuels Data.xlsx",
              show_default=True, help="Path to the source Excel file.")
@click.option("--output", "-o", "out_dir",
              default="reports", show_default=True, help="Output directory for reports.")
@click.option("--ai/--no-ai", default=True, show_default=True,
              help="Enable/disable Claude AI narrative generation.")
def run(input_path: str, out_dir: str, ai: bool):
    """Full pipeline: Ingest > Analyze > Lubes > Premium Fuel > AI Insights > Excel Report."""
    from pso import ingest, analyze, lubes_analyze, premium_fuel_analyze, ai_insights, excel_report

    console.print(Panel.fit(
        "[bold white]PSO OMC Analytics Pipeline[/]\n"
        "[dim]Ingest > Analyze > Lubes > Premium Fuel > AI > Excel[/]",
        border_style="cyan",
    ))

    # -- Agent 1: Ingest --------------------------------------------------------
    console.rule("[cyan]Step 1 / 6 — Ingest Agent[/]")
    df, quality = ingest.load(input_path)

    # -- Agent 2: Analysis ------------------------------------------------------
    console.rule("[cyan]Step 2 / 6 — Analysis Agent[/]")
    analysis_tables = analyze.run_all(df)
    _print_table_summary(analysis_tables)

    # -- Agent 3: Lubes --------------------------------------------------------
    console.rule("[cyan]Step 3 / 6 — Lubes Agent[/]")
    lubes_tables = lubes_analyze.run_lubes(df)
    _print_table_summary(lubes_tables)

    # -- Agent 3b: Premium Fuel --------------------------------------------------
    console.rule("[cyan]Step 4 / 6 — Premium Fuel Agent[/]")
    premium_tables = premium_fuel_analyze.run_premium_fuel(df)
    _print_table_summary(premium_tables)

    # -- Agent 4: AI Narrator --------------------------------------------------
    console.rule("[cyan]Step 5 / 6 — AI Narrator Agent[/]")
    period = df["_Period"].iloc[0] if "_Period" in df.columns else "Unknown"
    if ai:
        import json
        from datetime import date
        Path(out_dir).mkdir(parents=True, exist_ok=True)
        ai_json = Path(out_dir) / f"PSO_AI_{period}_{date.today()}.json"
        # Load any previously completed sections so we resume rather than redo
        existing_ai: dict = {}
        if ai_json.exists():
            try:
                existing_ai = json.loads(ai_json.read_text(encoding="utf-8"))
                console.print(f"  [dim]Resuming from existing JSON: {ai_json.name}[/]")
            except Exception:
                pass
        ai_output = ai_insights.run_ai_insights(
            analysis_tables, lubes_tables, period, existing=existing_ai
        )
        with open(ai_json, "w", encoding="utf-8") as f:
            json.dump(ai_output, f, ensure_ascii=False, indent=2)
        console.print(f"  AI text saved: {ai_json.name}")
    else:
        console.print("[yellow]AI skipped (--no-ai flag)[/]")
        ai_output = {}

    # -- Agent 5: Report --------------------------------------------------------
    console.rule("[cyan]Step 6 / 6 — Report Agent[/]")
    report_path = excel_report.build(
        analysis_tables, lubes_tables, premium_tables, ai_output, quality, period, out_dir
    )

    console.print(Panel.fit(
        f"[bold green]Done![/]  Report saved to:\n[white]{report_path}[/]",
        border_style="green",
    ))


@cli.command()
@click.option("--input", "-i", "input_path",
              default="data/input/Working File Retail Fuels Data.xlsx",
              show_default=True)
@click.option("--output", "-o", "out_dir", default="reports", show_default=True)
def lubes(input_path: str, out_dir: str):
    """Run lubricants analysis only — faster iteration."""
    from pso import ingest, lubes_analyze, excel_report
    import pandas as pd
    from openpyxl import Workbook
    from datetime import date

    console.print(Panel.fit("[bold white]PSO Lubes-Only Analysis[/]", border_style="yellow"))

    df, quality = ingest.load(input_path)
    period = df["_Period"].iloc[0] if "_Period" in df.columns else "Unknown"

    lubes_tables = lubes_analyze.run_lubes(df)
    _print_table_summary(lubes_tables)

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = out_dir / f"PSO_Lubes_{period}_{date.today()}.xlsx"

    from pso.excel_report import (
        _write_header, _write_section_title, _write_df, _write_ai_block,
        _sheet_lubes_overview, _sheet_lube_category, PCT_CHG_COLS,
    )

    wb = Workbook()
    wb.remove(wb.active)

    # Reuse the lube sheets from full report
    _sheet_lubes_overview(wb, {}, lubes_tables, period)
    _sheet_lube_category(wb, lubes_tables, "DEO",       "DEO",       period)
    _sheet_lube_category(wb, lubes_tables, "PCMO",      "PCMO",      period)
    _sheet_lube_category(wb, lubes_tables, "MCO",       "MCO",       period)
    _sheet_lube_category(wb, lubes_tables, "LOW GRADE", "LOW_GRADE", period)

    wb.save(fname)
    console.print(f"[bold green]Saved:[/] {fname}")


@cli.command()
@click.option("--input", "-i", "input_path",
              default="data/input/Working File Retail Fuels Data.xlsx",
              show_default=True)
def quality(input_path: str):
    """Run data quality check only — prints report to console."""
    from pso import ingest

    df, q = ingest.load(input_path)

    t = RichTable(title="Data Quality Report", show_header=True, header_style="bold cyan")
    t.add_column("Check", style="white")
    t.add_column("Count", justify="right", style="bold")

    rows = [
        ("Total rows",              str(q["total_rows"])),
        ("Retail Business rows",    str(q["retail_rows"])),
        ("International rows",      str(q["international_rows"])),
        ("Null Sales Org",          str(q["null_org_rows"])),
        ("Null Region (Retail)",    str(q["null_region_retail"])),
        ("Negative Vol CY",         str(q["negative_vol_cy"])),
        ("Zero Vol CY",             str(q["zero_vol_cy_rows"])),
        ("Unnorm cities (distinct)", str(len(q["unnormalized_cities_top30"]))),
    ]
    for label, val in rows:
        t.add_row(label, val)

    console.print(t)

    if q["unnormalized_cities_top30"]:
        console.print("\n[yellow]Top unnormalized cities (add to CITY_NORM in config.py):[/]")
        for city, cnt in list(q["unnormalized_cities_top30"].items())[:20]:
            console.print(f"  [dim]{cnt:>4}x[/]  {city}")


@cli.command()
@click.option("--input", "-i", "input_path",
              default="data/input/Working File Retail Fuels Data.xlsx",
              show_default=True)
def categories(input_path: str):
    """List Sales Org / Category combinations available for run-category, with
    live row counts from the loaded file."""
    from pso import ingest
    from pso.config import (
        SALES_ORG_CATEGORIES, RETAIL_CATEGORY_SEGMENTS,
        COL_ORG, COL_FUEL_SEG, COL_VOL_CY,
    )

    df, _ = ingest.load(input_path)

    t = RichTable(title="Sales Org / Category — Available Combinations",
                  show_header=True, header_style="bold cyan")
    t.add_column("Sales Org", style="white")
    t.add_column("Category", style="white")
    t.add_column("Rows", justify="right")
    t.add_column("Vol CY (ML)", justify="right")

    for org, cats in SALES_ORG_CATEGORIES.items():
        org_df = df[df[COL_ORG] == org]
        if cats is None:
            vol_ml = org_df[COL_VOL_CY].sum() / 1e6
            t.add_row(org, "—", str(len(org_df)), f"{vol_ml:,.2f}")
        else:
            for cat in cats:
                segs = RETAIL_CATEGORY_SEGMENTS.get(cat, set())
                cat_df = org_df[org_df[COL_FUEL_SEG].isin(segs)]
                vol_ml = cat_df[COL_VOL_CY].sum() / 1e6
                t.add_row(org, cat, str(len(cat_df)), f"{vol_ml:,.2f}")

    console.print(t)
    console.print(
        "\n[dim]Run reports for one combination with:[/]\n"
        '  uv run python -m pso.main run-category --org "Retail Business" --category "Fuels"\n'
        '  uv run python -m pso.main run-category --org "Aviation"'
    )


@cli.command(name="run-category")
@click.option("--org", required=True, help='Sales Org, e.g. "Retail Business". See `pso.main categories`.')
@click.option("--category", default=None, help='Category, required only for Orgs that have one (Retail Business).')
@click.option("--input", "-i", "input_path",
              default="data/input/Working File Retail Fuels Data.xlsx",
              show_default=True)
@click.option("--output", "-o", "out_base", default="reports", show_default=True,
              help="Base output directory — the category folder is created under this.")
def run_category(org: str, category: str | None, input_path: str, out_base: str):
    """Run only the reports relevant to one Sales Org (+ Category, if applicable),
    into a folder named after the selection."""
    from pso.config import SALES_ORG_CATEGORIES

    if org not in SALES_ORG_CATEGORIES:
        console.print(f"[red]Unknown Sales Org:[/] {org}")
        console.print(f"Valid options: {', '.join(SALES_ORG_CATEGORIES.keys())}")
        raise SystemExit(1)

    valid_cats = SALES_ORG_CATEGORIES[org]
    if valid_cats is None:
        if category:
            console.print(f"[red]{org} does not take a --category (it has no sub-categories).[/]")
            raise SystemExit(1)
    else:
        if category not in valid_cats:
            console.print(f"[red]{org} requires --category, one of:[/] {', '.join(valid_cats)}")
            raise SystemExit(1)

    folder_name = _safe_name(org) + (f"_{_safe_name(category)}" if category else "")
    out_dir = Path(out_base) / folder_name
    out_dir.mkdir(parents=True, exist_ok=True)

    label = org + (f" — {category}" if category else "")
    console.print(Panel.fit(
        f"[bold white]Run Category[/]\n[dim]{label}[/]\nOutput: {out_dir}",
        border_style="cyan",
    ))

    from pso import ingest
    df, quality = ingest.load(input_path)
    period = df["_Period"].iloc[0] if "_Period" in df.columns else "Unknown"

    root = Path(__file__).resolve().parents[2]   # project root
    env = os.environ.copy()
    env["PSO_INPUT"]  = input_path
    env["PSO_OUTDIR"] = str(out_dir)

    if org == "Retail Business" and category == "Lubricants":
        from pso import lubes_analyze
        lubes_tables = lubes_analyze.run_lubes(df)
        _print_table_summary(lubes_tables)
        _build_lubes_workbook(lubes_tables, period, out_dir)
        _run_workspace_scripts([
            "lubes_report.py", "lubes_stations_analysis.py", "lubes_stations_report.py",
            "city_profiles.py", "city_profiles_volume.py", "lubes_vol_table.py",
            "lubes_vol_uplift.py", "national_vol_slide.py", "frameworks.py",
        ], root, env, out_dir=out_dir)

    elif org == "Retail Business" and category == "Fuels":
        from pso import analyze, premium_fuel_analyze
        analysis_tables = analyze.run_all(df)
        premium_tables = premium_fuel_analyze.run_premium_fuel(df)
        _print_table_summary(analysis_tables)
        _print_table_summary(premium_tables)
        _build_fuels_workbook(analysis_tables, premium_tables, period, out_dir)
        _run_workspace_scripts([
            "fuels_report.py", "fuels_stations_analysis.py", "fuels_stations_report.py",
            "fuels_city_profiles.py", "fuels_city_profiles_volume.py", "fuels_vol_table.py",
            "fuels_vol_uplift.py", "fuels_national_vol_slide.py", "fuels_frameworks.py",
        ], root, env)

    else:
        from pso import org_report
        excel_path, docx_path = org_report.build(df, org, period, out_dir)
        console.print(f"  [green]Saved:[/] {excel_path}")
        console.print(f"  [green]Saved:[/] {docx_path}")

    console.print(Panel.fit(f"[bold green]Done![/]  Reports in:\n[white]{out_dir}[/]", border_style="green"))


def _safe_name(s: str) -> str:
    import re
    return re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_")


def _build_lubes_workbook(lubes_tables: dict, period: str, out_dir: Path) -> None:
    from openpyxl import Workbook
    from pso.excel_report import _sheet_lubes_overview, _sheet_lube_category

    wb = Workbook()
    wb.remove(wb.active)
    _sheet_lubes_overview(wb, {}, lubes_tables, period)
    _sheet_lube_category(wb, lubes_tables, "DEO",       "DEO",       period)
    _sheet_lube_category(wb, lubes_tables, "PCMO",      "PCMO",      period)
    _sheet_lube_category(wb, lubes_tables, "MCO",       "MCO",       period)
    _sheet_lube_category(wb, lubes_tables, "LOW GRADE", "LOW_GRADE", period)
    _sheet_lube_category(wb, lubes_tables, "OTHERS",    "Other",     period)

    fname = out_dir / f"PSO_Lubricants_{period}.xlsx"
    wb.save(fname)
    console.print(f"  [green]Saved:[/] {fname}")


def _build_fuels_workbook(analysis_tables: dict, premium_tables: dict, period: str, out_dir: Path) -> None:
    from openpyxl import Workbook
    from pso.excel_report import _sheet_fuel, _sheet_premium_fuel_overview, _sheet_premium_fuel_stations

    wb = Workbook()
    wb.remove(wb.active)
    _sheet_fuel(wb, analysis_tables, "Diesel", "02_Diesel", period)
    _sheet_fuel(wb, analysis_tables, "Petrol", "03_Petrol", period)
    _sheet_premium_fuel_overview(wb, premium_tables, period)
    _sheet_premium_fuel_stations(wb, premium_tables, period)

    fname = out_dir / f"PSO_Fuels_{period}.xlsx"
    wb.save(fname)
    console.print(f"  [green]Saved:[/] {fname}")


# Some older workspace scripts hardcode their own output subfolder instead of
# respecting PSO_OUTDIR (they predate the run-category feature and are left
# untouched). Map script -> (hardcoded folder relative to project root, name to
# use for the destination subfolder under the category's out_dir) so
# _run_workspace_scripts can relocate their output after the fact.
_LEGACY_OUTPUT_OVERRIDES = {
    "city_profiles.py":        ("reports/city_profiles",        "city_profiles"),
    "city_profiles_volume.py": ("reports/city_profiles_volume", "city_profiles_volume"),
}


def _run_workspace_scripts(scripts: list[str], root: Path, env: dict, out_dir: Path | None = None) -> None:
    import subprocess

    failures = []
    for script in scripts:
        path = root / "workspace" / script
        if not path.exists():
            console.print(f"  [yellow]Skipping (not built yet):[/] {script}")
            continue

        override = _LEGACY_OUTPUT_OVERRIDES.get(script)
        legacy_dir = None
        before: dict[Path, float] = {}
        if override and out_dir is not None:
            legacy_dir = root / override[0]
            if legacy_dir.exists():
                before = {p: p.stat().st_mtime for p in legacy_dir.glob("*") if p.is_file()}

        console.print(f"  [cyan]Running[/] {script} …")
        result = subprocess.run([sys.executable, str(path)], cwd=root, env=env)
        if result.returncode != 0:
            failures.append(script)
            console.print(f"  [red]FAILED:[/] {script} (exit {result.returncode})")
            continue

        if override and legacy_dir is not None and legacy_dir.exists():
            dest = out_dir / override[1]
            moved = 0
            for p in legacy_dir.glob("*"):
                if not p.is_file():
                    continue
                if p not in before or p.stat().st_mtime != before[p]:
                    dest.mkdir(parents=True, exist_ok=True)
                    p.replace(dest / p.name)
                    moved += 1
            if moved:
                console.print(f"  [dim]Relocated {moved} file(s): {legacy_dir} -> {dest}[/]")

    if failures:
        console.print(f"  [red]{len(failures)} script(s) failed:[/] {failures}")


def _print_table_summary(tables: dict) -> None:
    t = RichTable(show_header=True, header_style="bold dim")
    t.add_column("Table", style="cyan")
    t.add_column("Rows", justify="right")
    t.add_column("Cols", justify="right")
    for name, df in tables.items():
        t.add_row(name, str(len(df)), str(len(df.columns)))
    console.print(t)


if __name__ == "__main__":
    cli()
