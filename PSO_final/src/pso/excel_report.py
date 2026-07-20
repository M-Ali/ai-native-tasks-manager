"""Report Agent — multi-sheet formatted Excel workbook."""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers as xl_numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from rich.console import Console

from pso.config import (
    COLOUR_GREEN_FILL, COLOUR_RED_FILL, COLOUR_YELLOW_FILL,
    COLOUR_HEADER_BG, COLOUR_HEADER_FG, COLOUR_SUBHDR_BG, COLOUR_SUBHDR_FG,
    COLOUR_ALT_ROW, GROWTH_THRESHOLD_PCT,
)

console = Console()

# -- style constants ------------------------------------------------------------
_HDR_FONT  = Font(bold=True, color=COLOUR_HEADER_FG, name="Calibri", size=11)
_HDR_FILL  = PatternFill("solid", fgColor=COLOUR_HEADER_BG)
_SUBHDR_FONT  = Font(bold=True, color=COLOUR_SUBHDR_FG, name="Calibri", size=10)
_SUBHDR_FILL  = PatternFill("solid", fgColor=COLOUR_SUBHDR_BG)
_ALT_FILL  = PatternFill("solid", fgColor=COLOUR_ALT_ROW)
_GREEN     = PatternFill("solid", fgColor=COLOUR_GREEN_FILL)
_RED       = PatternFill("solid", fgColor=COLOUR_RED_FILL)
_YELLOW    = PatternFill("solid", fgColor=COLOUR_YELLOW_FILL)
_THIN      = Side(style="thin", color="CCCCCC")
_BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT      = Alignment(horizontal="left", vertical="center")
_RIGHT     = Alignment(horizontal="right", vertical="center")


def _pct_fill(val) -> PatternFill | None:
    if pd.isna(val):
        return None
    if val > GROWTH_THRESHOLD_PCT:
        return _GREEN
    if val < -GROWTH_THRESHOLD_PCT:
        return _RED
    return _YELLOW


def _write_header(ws, title: str, period: str, row: int = 1) -> int:
    ws.merge_cells(f"A{row}:L{row}")
    cell = ws[f"A{row}"]
    cell.value = f"{title}  |  Period: {period}"
    cell.font   = Font(bold=True, size=13, color=COLOUR_HEADER_FG, name="Calibri")
    cell.fill   = _HDR_FILL
    cell.alignment = _CENTER
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_df(
    ws, df: pd.DataFrame, start_row: int, start_col: int = 1,
    pct_cols: list[str] | None = None,
    table_name: str | None = None,
) -> int:
    """Write DataFrame to worksheet starting at (start_row, start_col).
    Returns the next empty row."""
    if df.empty:
        ws.cell(start_row, start_col, "No data")
        return start_row + 2

    pct_cols = pct_cols or []
    cols = df.columns.tolist()
    n_cols = len(cols)
    n_rows = len(df)

    # Header row
    for ci, col in enumerate(cols, start_col):
        cell = ws.cell(start_row, ci)
        cell.value     = col
        cell.font      = _SUBHDR_FONT
        cell.fill      = _SUBHDR_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
    ws.row_dimensions[start_row].height = 16

    # Data rows
    for ri, (_, row_data) in enumerate(df.iterrows(), 1):
        fill = _ALT_FILL if ri % 2 == 0 else None
        for ci, col in enumerate(cols, start_col):
            val = row_data[col]
            cell = ws.cell(start_row + ri, ci)

            if isinstance(val, float) and np.isnan(val):
                cell.value = None
            elif isinstance(val, (int, float)):
                cell.value = round(float(val), 4)
                cell.alignment = _RIGHT
            elif isinstance(val, bool):
                cell.value = "Yes" if val else "No"
                cell.alignment = _CENTER
            else:
                cell.value = str(val) if val is not None else ""
                cell.alignment = _LEFT

            cell.border = _BORDER

            # pct change colouring
            if col in pct_cols:
                pf = _pct_fill(val if not (isinstance(val, float) and np.isnan(val)) else None)
                if pf:
                    cell.fill = pf
            elif fill:
                cell.fill = fill

    end_row = start_row + n_rows

    # Excel Table
    if table_name and n_rows > 0:
        safe_name = re.sub(r"[^A-Za-z0-9_]", "_", table_name)[:30]
        tbl_ref = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(start_col + n_cols - 1)}{end_row}"
        )
        tbl = Table(displayName=safe_name, ref=tbl_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        try:
            ws.add_table(tbl)
        except Exception:
            pass  # duplicate name — skip silently

    # Auto-width (approximate)
    for ci, col in enumerate(cols, start_col):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) else 0,
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 30)

    return end_row + 2


def _write_section_title(ws, title: str, row: int, col: int = 1) -> int:
    cell = ws.cell(row, col)
    cell.value = title
    cell.font  = Font(bold=True, size=11, color=COLOUR_HEADER_FG, name="Calibri")
    cell.fill  = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = _LEFT
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_ai_block(ws, title: str, text: str, row: int) -> int:
    # Section title
    cell = ws.cell(row, 1)
    cell.value = f"-- {title} --"
    cell.font  = Font(bold=True, size=11, color="1F3864", name="Calibri")
    cell.fill  = PatternFill("solid", fgColor="DCE6F1")
    ws.row_dimensions[row].height = 18
    row += 1

    # Excel cells cap at 32,767 chars. Chunk text into <=800-char segments per cell
    # so that long AI responses are never silently truncated.
    CHUNK = 800
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    for line in lines:
        if not line.strip():
            # blank separator
            row += 1
            continue
        # If a single line is very long, break it into sub-chunks
        while len(line) > CHUNK:
            chunk, line = line[:CHUNK], line[CHUNK:]
            cell = ws.cell(row, 1)
            cell.value = chunk
            cell.font  = Font(size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 30
            row += 1
        cell = ws.cell(row, 1)
        cell.value = line
        cell.font  = Font(size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 15
        row += 1
    return row + 1


PCT_CHG_COLS = [
    "GRS_Chg_Pct", "Vol_Chg_Pct", "NMgn_Chg_Pct",
    "Vol_Chg_Pct", "NMgn_per_Ltr_Chg_Pct",
    "GRS_Chg_Pct", "PMgn_per_Ltr_Chg", "Disc_per_Ltr_Chg", "NMgn_per_Ltr_Chg",
]


def build(
    analysis: dict[str, pd.DataFrame],
    lubes: dict[str, pd.DataFrame],
    premium: dict[str, pd.DataFrame],
    ai: dict[str, str],
    quality: dict,
    period: str,
    out_dir: str | Path = "reports",
) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = out_dir / f"PSO_Analytics_{period}_{date.today()}.xlsx"

    console.print(f"[bold cyan]Building Excel report[/] > {filename.name}")

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # -- Sheet 00: Executive Summary --------------------------------------------
    _sheet_summary(wb, analysis, period)

    # -- Sheet 01: Retail Split -------------------------------------------------
    _sheet_retail_split(wb, analysis, period)

    # -- Sheet 02: Diesel ------------------------------------------------------
    _sheet_fuel(wb, analysis, "Diesel", "02_Diesel", period)

    # -- Sheet 03: Petrol ------------------------------------------------------
    _sheet_fuel(wb, analysis, "Petrol", "03_Petrol", period)

    # -- Sheet 04: Lubes Overview ----------------------------------------------
    _sheet_lubes_overview(wb, analysis, lubes, period)

    # -- Sheets 05-09: Per lube category --------------------------------------
    _sheet_lube_category(wb, lubes, "DEO",             "05_Lubes_DEO",       period)
    _sheet_lube_category(wb, lubes, "PCMO",            "06_Lubes_PCMO",      period)
    _sheet_lube_category(wb, lubes, "MCO",             "07_Lubes_MCO",       period)
    _sheet_lube_category(wb, lubes, "LOW GRADE",       "08_Lubes_LowGrade",  period)
    _sheet_lube_category(wb, lubes, "OTHERS",          "09_Lubes_Other",     period)

    # -- Sheet 10: City Pareto -------------------------------------------------
    _sheet_city_pareto(wb, analysis, lubes, period)

    # -- Sheet 11: Regional Gaps -----------------------------------------------
    _sheet_regional_gaps(wb, analysis, period)

    # -- Sheet 12: AI Observations ---------------------------------------------
    _sheet_ai(wb, ai, period)

    # -- Sheet 13: City Opportunity Sizing ------------------------------------
    _sheet_city_opportunity(wb, analysis, lubes, period)

    # -- Sheet 14: Premium Fuel (R95 vs PMG) — Overview -------------------------
    _sheet_premium_fuel_overview(wb, premium, period)

    # -- Sheet 15: Premium Fuel (R95 vs PMG) — Stations ------------------------
    _sheet_premium_fuel_stations(wb, premium, period)

    # -- Sheet 16: Data Quality ------------------------------------------------
    _sheet_quality(wb, quality, period)

    # If the file is open in Excel, try alternate names rather than crashing
    for suffix in ["", "_v2", "_v3", "_v4", "_v5"]:
        candidate = out_dir / f"PSO_Analytics_{period}_{date.today()}{suffix}.xlsx"
        try:
            wb.save(candidate)
            console.print(f"  [bold green]Saved:[/] {candidate}")
            return candidate
        except PermissionError:
            if suffix == "_v5":
                raise
            console.print(f"  [yellow]File locked, trying {candidate.stem}{suffix or ''}...[/]")
    return filename  # unreachable


# -- sheet builders ------------------------------------------------------------

def _sheet_summary(wb: Workbook, analysis: dict, period: str) -> None:
    ws = wb.create_sheet("00_Summary")
    ws.freeze_panes = "A3"
    row = _write_header(ws, "PSO OMC — Portfolio Scorecard", period)
    row = _write_section_title(ws, "All Business Channels — CY vs SPLY", row)
    df = analysis.get("portfolio_summary", pd.DataFrame())
    _write_df(ws, df, row, pct_cols=PCT_CHG_COLS, table_name="tbl_portfolio")


def _sheet_retail_split(wb: Workbook, analysis: dict, period: str) -> None:
    ws = wb.create_sheet("01_Retail_Split")
    ws.freeze_panes = "A3"
    row = _write_header(ws, "Retail Business — Fuels vs Lubricants Split", period)
    row = _write_section_title(ws, "Retail Segment Breakdown", row)
    df = analysis.get("retail_segment_split", pd.DataFrame())
    row = _write_df(ws, df, row, pct_cols=PCT_CHG_COLS, table_name="tbl_retail_split")

    row = _write_section_title(ws, "Region Performance — All Retail Segments", row + 1)
    df2 = analysis.get("region_performance", pd.DataFrame())
    _write_df(ws, df2, row, pct_cols=PCT_CHG_COLS, table_name="tbl_region_perf")


def _sheet_fuel(wb: Workbook, analysis: dict, fuel: str, sheet_name: str, period: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A3"
    key = fuel.lower()
    row = _write_header(ws, f"{fuel} — Performance Analysis", period)

    row = _write_section_title(ws, f"{fuel} by Region", row)
    df = analysis.get(f"{key}_by_region", pd.DataFrame())
    row = _write_df(ws, df, row, pct_cols=PCT_CHG_COLS, table_name=f"tbl_{key}_region")

    row = _write_section_title(ws, f"{fuel} by Product within Region", row + 1)
    df2 = analysis.get(f"{key}_by_product", pd.DataFrame())
    row = _write_df(ws, df2, row, pct_cols=PCT_CHG_COLS, table_name=f"tbl_{key}_product")

    row = _write_section_title(ws, f"{fuel} — City Pareto (volume concentration)", row + 1)
    df3 = analysis.get(f"{key}_pareto", pd.DataFrame())
    _write_df(ws, df3, row, pct_cols=PCT_CHG_COLS, table_name=f"tbl_{key}_pareto")


def _sheet_lubes_overview(wb: Workbook, analysis: dict, lubes: dict, period: str) -> None:
    ws = wb.create_sheet("04_Lubes_Overview")
    ws.freeze_panes = "A3"
    row = _write_header(ws, "Lubricants — Full Overview", period)

    row = _write_section_title(ws, "Category Trend (CY vs SPLY)", row)
    row = _write_df(ws, lubes.get("lube_category_trend", pd.DataFrame()), row,
                    pct_cols=PCT_CHG_COLS + ["NMgn_Chg_Pct"],
                    table_name="tbl_lube_cat_trend")

    row = _write_section_title(ws, "Margin Decomposition — Primary > Discount > Net (per litre)", row + 1)
    row = _write_df(ws, lubes.get("lube_margin_decomp", pd.DataFrame()), row,
                    pct_cols=["NMgn_per_Ltr_Chg_Pct"],
                    table_name="tbl_lube_marg_decomp")

    row = _write_section_title(ws, "Lubes by Region", row + 1)
    row = _write_df(ws, analysis.get("lubes_by_region", pd.DataFrame()), row,
                    pct_cols=PCT_CHG_COLS, table_name="tbl_lubes_region")

    row = _write_section_title(ws, "Net Margin/Ltr — Region × Category (value pockets)", row + 1)
    _write_df(ws, lubes.get("lube_margin_pockets", pd.DataFrame()), row,
              table_name="tbl_lube_mgn_pockets")


def _sheet_lube_category(wb: Workbook, lubes: dict, cat: str, sheet_name: str, period: str) -> None:
    key_map = {
        "DEO": "deo_detail",
        "PCMO": "pcmo_detail",
        "MCO": "mco_detail",
        "LOW GRADE": "low_grade_detail",
        "OTHERS": None,
    }
    detail_key = key_map.get(cat)
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A3"
    row = _write_header(ws, f"Lubricants — {cat} Deep Dive", period)

    if detail_key and detail_key in lubes:
        row = _write_section_title(ws, f"{cat}: Region & City Breakdown", row)
        row = _write_df(ws, lubes[detail_key], row, pct_cols=PCT_CHG_COLS,
                        table_name=f"tbl_{re.sub(r'[^A-Za-z0-9]','_',cat)}_detail")

    # Category-level customer analysis from corp_segments
    corp = lubes.get("lube_corp_segments", pd.DataFrame())
    if not corp.empty and "LubeCategory" in corp.columns:
        cat_corp = corp[corp["LubeCategory"] == cat]
        if not cat_corp.empty:
            row = _write_section_title(ws, f"{cat}: Top Corporate Segments", row + 1)
            _write_df(ws, cat_corp, row, table_name=f"tbl_{re.sub(r'[^A-Za-z0-9]','_',cat)}_corp")


def _sheet_city_pareto(wb: Workbook, analysis: dict, lubes: dict, period: str) -> None:
    ws = wb.create_sheet("10_City_Pareto")
    ws.freeze_panes = "A3"
    row = _write_header(ws, "City Concentration — 50% / 80% Pareto", period)

    row = _write_section_title(ws, "All Retail — City Pareto (volume + GRS cumulative share)", row)
    row = _write_df(ws, analysis.get("city_pareto_all", pd.DataFrame()).head(60), row,
                    pct_cols=PCT_CHG_COLS, table_name="tbl_city_pareto_all")

    row = _write_section_title(ws, "Lubricants — City × Category Matrix (top 30 cities)", row + 1)
    _write_df(ws, lubes.get("lube_city_category", pd.DataFrame()), row,
              table_name="tbl_lube_city_cat")


def _sheet_regional_gaps(wb: Workbook, analysis: dict, period: str) -> None:
    ws = wb.create_sheet("11_Regional_Gaps")
    ws.freeze_panes = "A3"
    row = _write_header(ws, "Regional Performance Gaps", period)

    row = _write_section_title(ws, "Segment × Region Volume Matrix", row)
    row = _write_df(ws, analysis.get("segment_region_matrix", pd.DataFrame()), row,
                    table_name="tbl_seg_region_matrix")

    row = _write_section_title(ws, "Underperforming Segments (Vol AND Margin both declined)", row + 1)
    _write_df(ws, analysis.get("underperforming_regions", pd.DataFrame()), row,
              pct_cols=PCT_CHG_COLS, table_name="tbl_underperf")


def _sheet_ai(wb: Workbook, ai: dict[str, str], period: str) -> None:
    ws = wb.create_sheet("12_AI_Observations")
    ws.column_dimensions["A"].width = 120
    row = _write_header(ws, "AI-Generated Observations & Recommendations", period)
    ws.row_dimensions[1].height = 22

    section_titles = {
        "exec_summary":       "1. Executive Summary",
        "diesel_analysis":    "2. Diesel Analysis",
        "petrol_analysis":    "3. Petrol Analysis",
        "lubes_problem":      "4. Lubricants — Root Cause Analysis",
        "lubes_geography":    "5. Lubricants — Geographic Patterns",
        "regional_gaps":      "6. Regional Performance Gaps",
        "city_concentration": "7. City Concentration",
        "recommendations":    "8. Prioritized Action List for PSO Management",
    }

    if not ai:
        ws.cell(row, 1).value = (
            "AI insights not generated — set OPENAI_API_KEY, GEMINI_API_KEY, or "
            "ANTHROPIC_API_KEY in .env and re-run without --no-ai flag."
        )
        return

    provider = ai.get("_provider", "AI")
    ws.cell(row, 1).value = f"Generated by: {provider}  |  Period: {period}"
    ws.cell(row, 1).font = Font(italic=True, size=9, color="666666", name="Calibri")
    row += 2

    for key, title in section_titles.items():
        text = ai.get(key, "")
        if text:
            row = _write_ai_block(ws, title, text, row)


def _sheet_city_opportunity(wb: Workbook, analysis: dict, lubes: dict, period: str) -> None:
    ws = wb.create_sheet("13_City_Opportunity")
    ws.freeze_panes = "A3"
    row = _write_header(ws, "City Opportunity Sizing — Lube Mix + Fuel Discount", period)

    SOUTH_PETROL_DISC = 0.28   # PKR/ltr South benchmark
    SOUTH_DIESEL_DISC = 0.40
    MGN_GAP_PER_LTR   = 232.1  # PKR/ltr DEO − LOW GRADE margin gap

    # ── A. Lube Mix Opportunity ────────────────────────────────────────────────
    row = _write_section_title(ws, "A. Lube Mix Opportunity — shift LOW GRADE → DEO/PCMO @ PKR 232/ltr", row)

    lcc = lubes.get("lube_city_category", pd.DataFrame()).copy()
    if not lcc.empty:
        city_col = next((c for c in ("CityNorm", "City") if c in lcc.columns), None)
        lowg_col = next((c for c in lcc.columns if "LOW GRADE" in c and "Vol" in c), None)
        tot_col  = next((c for c in lcc.columns if "Total_Vol" in c and "ML" in c.upper()), None)

        if city_col and lowg_col and tot_col:
            lcc[lowg_col] = pd.to_numeric(lcc[lowg_col], errors="coerce").fillna(0)
            lcc[tot_col]  = pd.to_numeric(lcc[tot_col],  errors="coerce").fillna(0)
            lcc["_lowg_ml"]   = lcc[lowg_col] / 1_000_000   # litres → ML
            lcc["_tot_ml"]    = lcc[tot_col]
            lcc["_lowg_pct"]  = (lcc["_lowg_ml"] / lcc["_tot_ml"].replace(0, float("nan")) * 100).round(1)
            lcc["_mix_opp"]   = (lcc["_lowg_ml"] * MGN_GAP_PER_LTR).round(1)

            mix_top = (lcc.nlargest(15, "_mix_opp")
                       [[city_col, "_tot_ml", "_lowg_ml", "_lowg_pct", "_mix_opp"]]
                       .copy())
            mix_top.columns = ["City", "Total Lube (ML)", "LOW GRADE (ML)", "LOW GRADE %", "Opportunity (PKR M)"]
            row = _write_df(ws, mix_top, row, table_name="tbl_mix_opp")

    # ── B. Fuel Discount Opportunity ──────────────────────────────────────────
    fuel_opps = {}
    for seg, tkey, south_disc in [
        ("Petrol", "petrol_by_city", SOUTH_PETROL_DISC),
        ("Diesel", "diesel_by_city", SOUTH_DIESEL_DISC),
    ]:
        row = _write_section_title(
            ws, f"B{'' if seg == 'Petrol' else 'C'}. {seg} Discount Normalisation vs South (PKR {south_disc:.2f}/ltr)", row + 1
        )
        tbl = analysis.get(tkey, pd.DataFrame()).copy()
        if tbl.empty:
            continue

        city_col = "CityNorm" if "CityNorm" in tbl.columns else next(
            (c for c in tbl.columns if "city" in c.lower()), None)
        vol_col  = next((c for c in tbl.columns if "vol" in c.lower() and "cy" in c.lower() and "ml" in c.lower()), None)
        disc_col = next((c for c in tbl.columns if "disc" in c.lower() and "ltr" in c.lower() and "cy" in c.lower()), None)

        if not all([city_col, vol_col, disc_col]):
            continue

        tbl[vol_col]  = pd.to_numeric(tbl[vol_col],  errors="coerce").fillna(0)
        tbl[disc_col] = pd.to_numeric(tbl[disc_col], errors="coerce").fillna(0)
        tbl["_excess"] = (tbl[disc_col] - south_disc).clip(lower=0)
        tbl["_opp"]    = (tbl["_excess"] * tbl[vol_col]).round(1)

        fuel_out = (tbl[tbl["_excess"] > 0]
                    .nlargest(15, "_opp")
                    [[city_col, vol_col, disc_col, "_excess", "_opp"]]
                    .copy())
        fuel_out.columns = ["City", "Vol CY (ML)", "Disc/ltr CY (PKR)", "Excess vs South", "Opportunity (PKR M)"]
        row = _write_df(ws, fuel_out, row, table_name=f"tbl_{seg.lower()}_disc_opp")
        fuel_opps[seg] = tbl[[city_col, "_opp"]].rename(columns={city_col: "City", "_opp": seg})

    # ── D. Combined Ranking ───────────────────────────────────────────────────
    row = _write_section_title(ws, "D. Combined Opportunity Ranking (top 20 cities)", row + 1)

    parts = []
    if not lcc.empty and city_col and "_mix_opp" in lcc.columns:
        lube_agg = lcc[[city_col, "_mix_opp"]].rename(columns={city_col: "City", "_mix_opp": "Lube Mix"})
        parts.append(lube_agg)

    combined = None
    if parts:
        combined = parts[0]
    for seg, tdf in fuel_opps.items():
        combined = combined.merge(tdf, on="City", how="outer") if combined is not None else tdf

    if combined is not None:
        for c in combined.columns[1:]:
            combined[c] = pd.to_numeric(combined[c], errors="coerce").fillna(0)
        combined["Total Opportunity (PKR M)"] = combined.iloc[:, 1:].sum(axis=1)
        combined = combined.sort_values("Total Opportunity (PKR M)", ascending=False).head(20)
        combined = combined.reset_index(drop=True)
        _write_df(ws, combined, row, table_name="tbl_combined_opp")

    ws.column_dimensions["A"].width = 28
    for col_letter in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col_letter].width = 18


def _sheet_premium_fuel_overview(wb: Workbook, premium: dict, period: str) -> None:
    ws = wb.create_sheet("14_Premium_Fuel_R95")
    ws.freeze_panes = "A3"
    row = _write_header(ws, "Premium Fuel — R95 vs PMG (CY vs SPLY)", period)

    row = _write_section_title(ws, "National Trend — R95 vs PMG", row)
    row = _write_df(ws, premium.get("premium_product_trend", pd.DataFrame()), row,
                    pct_cols=PCT_CHG_COLS, table_name="tbl_premium_trend")

    row = _write_section_title(ws, "Margin Decomposition — Primary > Discount > Net (per litre)", row + 1)
    row = _write_df(ws, premium.get("premium_margin_decomp", pd.DataFrame()), row,
                    pct_cols=["Vol_Chg_Pct", "NMgn_per_Ltr_Chg_Pct"],
                    table_name="tbl_premium_margin")

    row = _write_section_title(ws, "R95 Share by Region", row + 1)
    row = _write_df(ws, premium.get("premium_by_region", pd.DataFrame()), row,
                    table_name="tbl_premium_region")

    row = _write_section_title(ws, "R95 Share by City (proof-of-demand ranking)", row + 1)
    row = _write_df(ws, premium.get("premium_by_city", pd.DataFrame()).head(60), row,
                    table_name="tbl_premium_city")

    row = _write_section_title(ws, "Cities — R95 Volume Growing vs SPLY", row + 1)
    row = _write_df(ws, premium.get("premium_growing_markets", pd.DataFrame()), row,
                    pct_cols=["Vol_Chg_Pct"], table_name="tbl_premium_growing")

    row = _write_section_title(ws, "Cities — R95 Volume Declining vs SPLY", row + 1)
    row = _write_df(ws, premium.get("premium_declining_markets", pd.DataFrame()), row,
                    pct_cols=["Vol_Chg_Pct"], table_name="tbl_premium_declining")

    row = _write_section_title(ws, "R95 Customer Segments (Corporate Groups)", row + 1)
    _write_df(ws, premium.get("premium_customer_segments", pd.DataFrame()), row,
              table_name="tbl_premium_corp")


def _sheet_premium_fuel_stations(wb: Workbook, premium: dict, period: str) -> None:
    ws = wb.create_sheet("15_Premium_Fuel_Stations")
    ws.freeze_panes = "A3"
    row = _write_header(ws, "Premium Fuel — Station-Level R95 Mix & Launch Whitespace", period)

    row = _write_section_title(ws, "Station Mix — Stations Already Selling R95 (top 100 by R95 volume)", row)
    row = _write_df(ws, premium.get("premium_station_mix", pd.DataFrame()).head(100), row,
                    table_name="tbl_premium_station_mix")

    row = _write_section_title(
        ws, "Whitespace — Active PMG Stations With Zero R95 (ranked by PMG volume = launch priority)", row + 1
    )
    _write_df(ws, premium.get("premium_whitespace_stations", pd.DataFrame()).head(200), row,
              table_name="tbl_premium_whitespace")


def _sheet_quality(wb: Workbook, quality: dict, period: str) -> None:
    ws = wb.create_sheet("16_Data_Quality")
    row = _write_header(ws, "Data Quality Report", period)

    scalar_rows = [
        ("Total rows in source",             quality.get("total_rows", 0)),
        ("Retail Business rows",             quality.get("retail_rows", 0)),
        ("International city rows (flagged)", quality.get("international_rows", 0)),
        ("Null Sales Org rows",              quality.get("null_org_rows", 0)),
        ("Null Region (Retail rows only)",   quality.get("null_region_retail", 0)),
        ("Negative Volume CY rows",          quality.get("negative_vol_cy", 0)),
        ("Zero Volume CY rows",              quality.get("zero_vol_cy_rows", 0)),
    ]
    for label, val in scalar_rows:
        ws.cell(row, 1).value = label
        ws.cell(row, 2).value = val
        ws.cell(row, 1).font = Font(name="Calibri", size=10)
        ws.cell(row, 2).font = Font(name="Calibri", size=10, bold=True)
        row += 1

    row += 1
    ws.cell(row, 1).value = "City names NOT in normalization table (top 30 by frequency)"
    ws.cell(row, 1).font = Font(bold=True, name="Calibri", size=10)
    row += 1
    ws.cell(row, 1).value = "City Name"
    ws.cell(row, 2).value = "Occurrences"
    ws.cell(row, 1).font = _SUBHDR_FONT
    ws.cell(row, 2).font = _SUBHDR_FONT
    ws.cell(row, 1).fill = _SUBHDR_FILL
    ws.cell(row, 2).fill = _SUBHDR_FILL
    row += 1
    for city, cnt in quality.get("unnormalized_cities_top30", {}).items():
        ws.cell(row, 1).value = city
        ws.cell(row, 2).value = cnt
        row += 1

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 15
